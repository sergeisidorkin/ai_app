import hashlib
import json
from datetime import date
from typing import Optional

from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Count, Max, Q
from django.http import HttpResponse, HttpResponseBadRequest, JsonResponse
from django.shortcuts import get_object_or_404, render
from django.template.loader import render_to_string
from django.urls import NoReverseMatch, reverse
from django.utils import timezone
from django.utils.text import slugify
from django.views.decorators.csrf import ensure_csrf_cookie
from django.views.decorators.http import require_GET, require_POST

from projects_app.models import LegalEntity, Performer, ProjectRegistration
from policy_app.models import TypicalSection
from requests_app.models import RequestItem, RequestTable

from yandexdisk_app.models import YandexDiskAccount, YandexDiskSelection

from .models import (
    ChecklistCommentHistory,
    ChecklistItem,
    ChecklistRequestNote,
    ChecklistStatus,
    ChecklistStatusHistory,
    SharedChecklistLink,
)


def _product_short_label(product) -> str:
    for attr in ("short_name", "short", "code", "name_en", "name"):
        val = getattr(product, attr, None)
        if val:
            return str(val)
    return str(product) if product else ""


def _project_meta(project: Optional[ProjectRegistration], asset_name: Optional[str] = None) -> dict:
    if not project:
        return {"assets": [], "asset_items": [], "asset": "", "sections": []}

    legal_qs = (
        LegalEntity.objects.filter(project=project)
        .select_related("work_item")
        .order_by("position", "id")
    )

    assets_flat = []
    asset_items = []
    seen_flat = set()
    seen_work_items = {}

    for entity in legal_qs:
        wi = entity.work_item
        asset_label = (
            (getattr(wi, "asset_name", "") or "").strip()
            or (getattr(wi, "name", "") or "").strip()
        )
        entity_label = (
            (entity.legal_name or "").strip()
            or (entity.work_name or "").strip()
            or asset_label
        )

        if asset_label and asset_label not in seen_work_items:
            seen_work_items[asset_label] = True
            asset_items.append({"value": f"asset:{asset_label}", "label": asset_label, "type": "asset"})

        if entity_label and entity_label not in seen_flat:
            seen_flat.add(entity_label)
            assets_flat.append(entity_label)
            asset_items.append({"value": entity_label, "label": entity_label, "type": "entity",
                                "asset": asset_label})

    performer_qs = None
    if not assets_flat:
        performer_qs = (
            Performer.objects.filter(registration=project)
            .exclude(asset_name="")
            .select_related("typical_section")
            .order_by("position", "id")
        )
        for perf in performer_qs:
            asset = (perf.asset_name or "").strip()
            if asset and asset not in seen_flat:
                seen_flat.add(asset)
                assets_flat.append(asset)
                asset_items.append({"value": asset, "label": asset, "type": "asset"})

    entity_count_per_asset = {}
    for ai in asset_items:
        if ai["type"] == "entity":
            a = ai.get("asset", "")
            entity_count_per_asset[a] = entity_count_per_asset.get(a, 0) + 1
    single_entity_assets = {a for a, c in entity_count_per_asset.items() if c == 1}
    asset_items = [
        ai for ai in asset_items
        if not (ai["type"] == "entity" and ai.get("asset", "") in single_entity_assets)
    ]

    all_values = set(assets_flat) | {ai["value"] for ai in asset_items} | {"all"}
    selected_asset = ""
    if asset_name and asset_name in all_values:
        selected_asset = asset_name
    else:
        first_asset = next((ai["value"] for ai in asset_items if ai["type"] == "asset"), None)
        selected_asset = first_asset or (assets_flat[0] if assets_flat else "")

    sections = _sections_for_asset(project, selected_asset, performer_qs)

    return {
        "assets": assets_flat,
        "asset_items": asset_items,
        "asset": selected_asset,
        "sections": sections,
    }


def _resolve_asset_name(project, selected_asset):
    """Return the Performer-level asset_name for any kind of asset selector."""
    if not selected_asset or selected_asset == "all":
        return selected_asset or ""
    if selected_asset.startswith("asset:"):
        return selected_asset[6:]
    entity = (
        LegalEntity.objects
        .filter(project=project)
        .filter(
            Q(legal_name__iexact=selected_asset)
            | Q(work_name__iexact=selected_asset)
        )
        .select_related("work_item")
        .first()
    )
    if entity and entity.work_item:
        return (entity.work_item.asset_name or entity.work_item.name or "").strip()
    return selected_asset


def _sections_for_asset(project, selected_asset, performer_qs=None):
    if not selected_asset:
        return []

    effective_asset = _resolve_asset_name(project, selected_asset)

    if performer_qs is None:
        performer_qs = (
            Performer.objects.filter(registration=project)
            .exclude(asset_name="")
            .select_related("typical_section")
            .order_by("position", "id")
        )

    if selected_asset == "all":
        ids = set()
        sections = []
        for perf in performer_qs:
            ts = getattr(perf, "typical_section", None)
            if ts and ts.id not in ids:
                ids.add(ts.id)
                label = str(ts)
                if ts.short_name_ru:
                    label += " " + ts.short_name_ru
                sections.append({"id": ts.id, "name": label})
        return sections

    ids = set()
    sections = []
    for perf in performer_qs:
        perf_asset = (perf.asset_name or "").strip()
        if perf_asset != effective_asset:
            continue
        ts = getattr(perf, "typical_section", None)
        if ts and ts.id not in ids:
            ids.add(ts.id)
            label = str(ts)
            if ts.short_name_ru:
                label += " " + ts.short_name_ru
            sections.append({"id": ts.id, "name": label})
    return sections


def _code_cell_class(status_cells):
    statuses = {
        cell["status"].status
        for cell in status_cells
        if cell.get("status")
    }
    if not statuses:
        return ""
    if statuses == {ChecklistStatus.Status.PROVIDED}:
        return "chk-code--provided"
    if statuses == {ChecklistStatus.Status.PARTIAL}:
        return "chk-code--partial"
    if statuses == {ChecklistStatus.Status.NOT_REQUIRED}:
        return "chk-code--na"
    if ChecklistStatus.Status.PARTIAL in statuses:
        return "chk-code--partial"
    if (
        ChecklistStatus.Status.PROVIDED in statuses
        and ChecklistStatus.Status.MISSING in statuses
    ):
        return "chk-code--partial"
    if (
        ChecklistStatus.Status.PROVIDED in statuses
        and ChecklistStatus.Status.NOT_REQUIRED in statuses
        and len(statuses) == 2
    ):
        return "chk-code--provided"
    return ""


def _project_options():
    regs = ProjectRegistration.objects.select_related("type").order_by("position", "id")
    options = []
    for reg in regs:
        short_uid = (reg.short_uid or "").strip()
        product_short = _product_short_label(getattr(reg, "type", None))
        options.append(
            {
                "id": reg.id,
                "short_uid": short_uid,
                "label": " ".join(x for x in (short_uid, product_short, reg.name) if x),
                "code": short_uid or f"{reg.number}{reg.group}".upper(),
                "product_short": product_short,
                "product_id": getattr(getattr(reg, "type", None), "id", None),
            }
        )
    return options


def _resolve_section(project: ProjectRegistration, section_id: Optional[str], asset_name: Optional[str]):
    if section_id:
        section = TypicalSection.objects.filter(pk=section_id, product=project.type).first()
        if section:
            return section

    meta = _project_meta(project, asset_name)
    first = meta["sections"][0]["id"] if meta["sections"] else None
    if first:
        return TypicalSection.objects.filter(pk=first).first()

    return None


def _legal_entities_for(project: ProjectRegistration, asset_name: Optional[str]):
    qs = (
        LegalEntity.objects.filter(project=project)
        .select_related("work_item")
        .order_by("position", "id")
    )
    if not asset_name or asset_name == "all":
        return list(qs)

    if asset_name.startswith("asset:"):
        effective = asset_name[6:]
        return list(qs.filter(
            Q(work_item__asset_name__iexact=effective)
            | Q(work_item__name__iexact=effective)
        ))

    entity_qs = list(qs.filter(
        Q(legal_name__iexact=asset_name)
        | Q(work_name__iexact=asset_name)
    ))
    if entity_qs:
        return entity_qs

    return list(qs.filter(
        Q(work_item__asset_name__iexact=asset_name)
        | Q(work_item__name__iexact=asset_name)
    ))


# ---------------------------------------------------------------------------
#  Lazy initialization: create ChecklistItems from RequestItems on first access
# ---------------------------------------------------------------------------

def _ensure_checklist_items(project: ProjectRegistration, section: TypicalSection):
    """Create ChecklistItems from RequestItems if none exist for this project+section."""
    if ChecklistItem.objects.filter(project=project, section=section).exists():
        return

    if not project.type:
        return

    table = RequestTable.objects.filter(product=project.type, section=section).first()
    if not table:
        return

    items_to_create = []
    for ri in table.items.all().order_by("position", "id"):
        items_to_create.append(ChecklistItem(
            project=project,
            section=section,
            code=ri.code,
            number=ri.number,
            short_name=ri.short_name,
            name=ri.name,
            position=ri.position,
            source_request_item=ri,
        ))
    if items_to_create:
        ChecklistItem.objects.bulk_create(items_to_create)


def _render_comment_modal_content(
    request,
    *,
    field,
    checklist_item,
    project,
    section,
    asset_name,
    add_comment_url,
    readonly=False,
):
    return render(
        request,
        "checklists_app/components/comment_modal.html",
        _comment_modal_context(
            checklist_item=checklist_item,
            project=project,
            section=section,
            field=field,
            active_scope_value=asset_name,
            add_comment_url=add_comment_url,
            readonly=readonly,
        ),
    )


# ---------------------------------------------------------------------------
#  Build table context (shared between internal and public views)
# ---------------------------------------------------------------------------

def _build_table_context(project, section, asset_name, checklist_items, legal_entities, url_map, include_comment_history=False):
    status_map = {}
    if checklist_items and legal_entities:
        status_qs = ChecklistStatus.objects.filter(
            checklist_item__in=[it.id for it in checklist_items],
            legal_entity__in=[le.id for le in legal_entities],
        ).select_related("checklist_item", "legal_entity")
        status_map = {(st.checklist_item_id, st.legal_entity_id): st for st in status_qs}

    note_map = {}
    if checklist_items:
        note_qs = ChecklistRequestNote.objects.filter(
            checklist_item__in=[it.id for it in checklist_items],
            project=project,
            section=section,
            asset_name=asset_name,
        )
        note_map = {note.checklist_item_id: note for note in note_qs}

    history_map = {}
    if checklist_items and include_comment_history:
        history_qs = (
            ChecklistCommentHistory.objects
            .filter(
                note__checklist_item__in=[it.id for it in checklist_items],
                note__project=project,
                note__section=section,
                note__asset_name=asset_name,
            )
            .select_related("author", "note")
            .order_by("-created_at")
        )
        for entry in history_qs:
            key = entry.note.checklist_item_id
            history_map.setdefault(key, {"imc_comment": [], "customer_comment": []})
            history_map[key][entry.field].append(entry)

    folder_map = {}
    if checklist_items:
        try:
            from checklists_app.models import ChecklistItemFolder
            folder_qs = ChecklistItemFolder.objects.filter(
                checklist_item__in=[it.id for it in checklist_items],
            )
            folder_map = {f.checklist_item_id: f for f in folder_qs}
        except Exception:
            pass

    rows = []
    seen_additional_groups = set()
    for item in checklist_items:
        statuses = [
            {"entity": entity, "status": status_map.get((item.id, entity.id))}
            for entity in legal_entities
        ]
        additional_header = None
        if item.item_type == ChecklistItem.ItemType.ADDITIONAL and item.additional_number and item.additional_date:
            group_key = (item.additional_number, item.additional_date)
            if group_key not in seen_additional_groups:
                seen_additional_groups.add(group_key)
                additional_header = (
                    f"Дополнительный запрос № {item.additional_number}"
                    f" от {item.additional_date.strftime('%d.%m.%Y')}"
                )
        folder = folder_map.get(item.id)
        rows.append({
            "item": item,
            "section_obj": section,
            "statuses": statuses,
            "code_class": _code_cell_class(statuses),
            "note": note_map.get(item.id),
            "history": history_map.get(item.id, {"imc_comment": [], "customer_comment": []}),
            "additional_header": additional_header,
            "file_count": folder.file_count if folder else None,
            "last_upload_at": folder.last_upload_at if folder else None,
        })

    return {
        "project": project,
        "section": section,
        "asset_name": asset_name,
        "legal_entities": legal_entities,
        "rows": rows,
        "status_choices": ChecklistStatus.Status.choices,
        "default_status": ChecklistStatus.Status.MISSING,
        "update_status_url": url_map["update_status"],
        "add_comment_url": url_map["add_comment"],
        "comment_modal_url": url_map["comment_modal"],
        "update_note_url": url_map["update_note"],
        "error": None if checklist_items else "Для выбранного продукта и раздела нет запросов",
    }


def _build_all_sections_context(project, section_items_list, asset_name, legal_entities, url_map, include_comment_history=False):
    all_item_ids = []
    for _sec, items in section_items_list:
        all_item_ids.extend(it.id for it in items)

    status_map = {}
    if all_item_ids and legal_entities:
        status_qs = ChecklistStatus.objects.filter(
            checklist_item__in=all_item_ids,
            legal_entity__in=[le.id for le in legal_entities],
        ).select_related("checklist_item", "legal_entity")
        status_map = {(st.checklist_item_id, st.legal_entity_id): st for st in status_qs}

    note_map = {}
    if all_item_ids:
        note_qs = ChecklistRequestNote.objects.filter(
            checklist_item__in=all_item_ids,
            project=project,
            asset_name=asset_name,
        )
        note_map = {note.checklist_item_id: note for note in note_qs}

    history_map = {}
    if all_item_ids and include_comment_history:
        history_qs = (
            ChecklistCommentHistory.objects
            .filter(note__checklist_item__in=all_item_ids, note__project=project, note__asset_name=asset_name)
            .select_related("author", "note")
            .order_by("-created_at")
        )
        for entry in history_qs:
            key = entry.note.checklist_item_id
            history_map.setdefault(key, {"imc_comment": [], "customer_comment": []})
            history_map[key][entry.field].append(entry)

    folder_map = {}
    if all_item_ids:
        try:
            from checklists_app.models import ChecklistItemFolder
            folder_qs = ChecklistItemFolder.objects.filter(checklist_item__in=all_item_ids)
            folder_map = {f.checklist_item_id: f for f in folder_qs}
        except Exception:
            pass

    rows = []
    for sec, items in section_items_list:
        if not items:
            continue
        rows.append({
            "section_header": sec.name_ru or str(sec),
            "section_obj": sec,
        })
        seen_additional_groups = set()
        for item in items:
            statuses = [
                {"entity": entity, "status": status_map.get((item.id, entity.id))}
                for entity in legal_entities
            ]
            additional_header = None
            if item.item_type == ChecklistItem.ItemType.ADDITIONAL and item.additional_number and item.additional_date:
                group_key = (item.additional_number, item.additional_date)
                if group_key not in seen_additional_groups:
                    seen_additional_groups.add(group_key)
                    additional_header = (
                        f"Дополнительный запрос № {item.additional_number}"
                        f" от {item.additional_date.strftime('%d.%m.%Y')}"
                    )
            folder = folder_map.get(item.id)
            rows.append({
                "item": item,
                "section_obj": sec,
                "statuses": statuses,
                "code_class": _code_cell_class(statuses),
                "note": note_map.get(item.id),
                "history": history_map.get(item.id, {"imc_comment": [], "customer_comment": []}),
                "additional_header": additional_header,
                "file_count": folder.file_count if folder else None,
                "last_upload_at": folder.last_upload_at if folder else None,
            })

    return {
        "project": project,
        "asset_name": asset_name,
        "legal_entities": legal_entities,
        "rows": rows,
        "status_choices": ChecklistStatus.Status.choices,
        "default_status": ChecklistStatus.Status.MISSING,
        "update_status_url": url_map["update_status"],
        "add_comment_url": url_map["add_comment"],
        "comment_modal_url": url_map["comment_modal"],
        "update_note_url": url_map["update_note"],
        "error": None if rows else "Для выбранного продукта нет запросов",
    }


def _status_short_label(value: str) -> str:
    return {
        ChecklistStatus.Status.PROVIDED: "Предоставлено",
        ChecklistStatus.Status.PARTIAL: "Предоставлено частично",
        ChecklistStatus.Status.MISSING: "Не предоставлено",
        ChecklistStatus.Status.NOT_REQUIRED: "Не требуется",
    }.get(value, value)


def _format_status_changed_at(status_obj: Optional[ChecklistStatus]) -> str:
    if not status_obj or not status_obj.status_changed_at:
        return "—"
    return timezone.localtime(status_obj.status_changed_at).strftime("%d.%m.%y %H:%M")


def _comment_flags(note: Optional[ChecklistRequestNote]) -> dict:
    return {
        "imc": bool(note and (note.imc_comment or "").strip()),
        "customer": bool(note and (note.customer_comment or "").strip()),
        "registeredCount": getattr(note, "registered_comment_count", 0) if note else 0,
        "guestCount": getattr(note, "guest_comment_count", 0) if note else 0,
        "lastCommentAt": "",
        "lastCommentByStaff": False,
    }


def _comment_scope_id(field: str, item_id: int, scope_value: str) -> str:
    normalized = slugify(scope_value or "all") or "all"
    digest = hashlib.md5((scope_value or "all").encode("utf-8")).hexdigest()[:8]
    return f"comment-history-{field}-{item_id}-{normalized}-{digest}"


def _comment_scope_tabs(project: ProjectRegistration) -> list[dict]:
    meta = _project_meta(project, None)
    tabs = [{"value": "all", "label": "Все активы", "type": "all"}]
    seen = {"all"}
    for item in meta.get("asset_items", []):
        value = (item.get("value") or "").strip()
        if not value or value in seen:
            continue
        seen.add(value)
        label = (item.get("label") or value).strip()
        item_type = item.get("type") or "entity"
        if item_type == "asset":
            label = f"{label} • \u1d00\u1d0b\u1d1b\u1d0e\u0299"
        tabs.append({
            "value": value,
            "label": label,
            "type": item_type,
        })
    return tabs


def _comment_item_totals(*, checklist_item: ChecklistItem, project: ProjectRegistration, section: TypicalSection, field: str) -> dict:
    note_qs = ChecklistRequestNote.objects.filter(
        checklist_item=checklist_item,
        project=project,
        section=section,
    )
    history_qs = ChecklistCommentHistory.objects.filter(note__in=note_qs, field=field)
    last_comment = history_qs.order_by("-created_at").select_related("author").first()
    last_comment_at = ""
    last_comment_by_staff = False
    if last_comment:
        last_comment_at = timezone.localtime(last_comment.created_at).strftime("%d.%m.%y %H:%M")
        last_comment_by_staff = bool(last_comment.author and last_comment.author.is_staff)
    return {
        "hasComment": note_qs.exclude(**{field: ""}).exists(),
        "registeredCount": history_qs.filter(author__isnull=False).count(),
        "guestCount": history_qs.filter(author__isnull=True).count(),
        "lastCommentAt": last_comment_at,
        "lastCommentByStaff": last_comment_by_staff,
    }


def _comment_summary_map(*, project: ProjectRegistration, item_ids: list[int], field: str) -> dict[int, dict]:
    _empty = {
        "imc": False,
        "customer": False,
        "registeredCount": 0,
        "guestCount": 0,
        "lastCommentAt": "",
        "lastCommentByStaff": False,
    }
    summary_map = {item_id: {**_empty} for item_id in item_ids}
    if not item_ids:
        return summary_map

    note_qs = ChecklistRequestNote.objects.filter(
        checklist_item__in=item_ids,
        project=project,
    ).only("checklist_item_id", "imc_comment", "customer_comment")
    for note in note_qs:
        summary = summary_map.setdefault(note.checklist_item_id, {**_empty})
        summary["imc"] = summary["imc"] or bool((note.imc_comment or "").strip())
        summary["customer"] = summary["customer"] or bool((note.customer_comment or "").strip())

    history_counts = (
        ChecklistCommentHistory.objects
        .filter(
            note__checklist_item__in=item_ids,
            note__project=project,
            field=field,
        )
        .values("note__checklist_item_id")
        .annotate(
            registered_count=Count("id", filter=Q(author__isnull=False)),
            guest_count=Count("id", filter=Q(author__isnull=True)),
        )
    )
    for entry in history_counts:
        summary = summary_map.setdefault(entry["note__checklist_item_id"], {**_empty})
        summary["registeredCount"] = entry["registered_count"]
        summary["guestCount"] = entry["guest_count"]

    latest_ids = (
        ChecklistCommentHistory.objects
        .filter(
            note__checklist_item__in=item_ids,
            note__project=project,
            field=field,
        )
        .values("note__checklist_item_id")
        .annotate(latest_id=Max("id"))
        .values_list("latest_id", flat=True)
    )
    last_comments = (
        ChecklistCommentHistory.objects
        .filter(id__in=latest_ids)
        .select_related("author", "note")
    )
    for entry in last_comments:
        item_id = entry.note.checklist_item_id
        if item_id in summary_map:
            summary_map[item_id]["lastCommentAt"] = timezone.localtime(entry.created_at).strftime("%d.%m.%y %H:%M")
            summary_map[item_id]["lastCommentByStaff"] = bool(entry.author and entry.author.is_staff)

    return summary_map


def _comment_modal_context(
    *,
    checklist_item: ChecklistItem,
    project: ProjectRegistration,
    section: TypicalSection,
    field: str,
    active_scope_value: str,
    add_comment_url: str,
    readonly: bool,
) -> dict:
    tabs = _comment_scope_tabs(project)
    valid_values = {tab["value"] for tab in tabs}
    active_value = active_scope_value if active_scope_value in valid_values else "all"
    note_qs = ChecklistRequestNote.objects.filter(
        checklist_item=checklist_item,
        project=project,
        section=section,
        asset_name__in=list(valid_values),
    )
    note_map = {note.asset_name: note for note in note_qs}
    history_qs = (
        ChecklistCommentHistory.objects
        .filter(note__in=note_qs, field=field)
        .select_related("author", "note")
        .order_by("created_at", "id")
    )
    history_map: dict[str, list[ChecklistCommentHistory]] = {}
    for entry in history_qs:
        history_map.setdefault(entry.note.asset_name, []).append(entry)

    totals = _comment_item_totals(
        checklist_item=checklist_item,
        project=project,
        section=section,
        field=field,
    )
    rendered_tabs = []
    for tab in tabs:
        scope_value = tab["value"]
        histories = history_map.get(scope_value, [])
        rendered_tabs.append({
            **tab,
            "scope_id": _comment_scope_id(field, checklist_item.id, scope_value),
            "active": scope_value == active_value,
            "histories": histories,
            "comment_count": len(histories),
            "has_comment": bool(note_map.get(scope_value) and getattr(note_map[scope_value], field, "").strip()),
        })

    return {
        "field": field,
        "item": checklist_item,
        "project": project,
        "section": section,
        "asset_name": active_value,
        "active_scope_id": _comment_scope_id(field, checklist_item.id, active_value),
        "tabs": rendered_tabs,
        "active_scope_value": active_value,
        "registered_count": totals["registeredCount"],
        "guest_count": totals["guestCount"],
        "has_comment": totals["hasComment"],
        "add_comment_url": add_comment_url,
        "readonly": readonly,
    }


def _comment_update_event_payload(note: ChecklistRequestNote, field: str) -> dict:
    totals = _comment_item_totals(
        checklist_item=note.checklist_item,
        project=note.project,
        section=note.section,
        field=field,
    )
    return {
        "itemId": note.checklist_item_id,
        "field": field,
        "scopeValue": note.asset_name,
        "hasComment": totals["hasComment"],
        "registeredCount": totals["registeredCount"],
        "guestCount": totals["guestCount"],
        "lastCommentAt": totals["lastCommentAt"],
        "lastCommentByStaff": totals["lastCommentByStaff"],
    }


def _resolve_grid_scope(project: ProjectRegistration, asset_name: str, section_id: str) -> dict:
    if not project.type:
        return {
            "error": "У проекта не указан тип продукта",
            "all_mode": False,
            "section": None,
            "section_items_list": [],
            "legal_entities": [],
        }

    all_mode = section_id == "all"
    effective_asset = _resolve_asset_name(project, asset_name)

    if all_mode:
        perf_qs = Performer.objects.filter(registration=project).exclude(asset_name="")
        if effective_asset and asset_name != "all":
            perf_qs = perf_qs.filter(asset_name=effective_asset)
        perf_qs = perf_qs.select_related("typical_section__product").order_by("position", "id")

        seen_ids = set()
        sections = []
        for perf in perf_qs:
            ts = perf.typical_section
            if ts and ts.id not in seen_ids:
                seen_ids.add(ts.id)
                sections.append(ts)

        if not sections:
            return {
                "error": "Нет разделов для данного продукта",
                "all_mode": True,
                "section": None,
                "section_items_list": [],
                "legal_entities": [],
            }

        section_items_list = []
        for sec in sections:
            _ensure_checklist_items(project, sec)
            items = list(ChecklistItem.objects.filter(project=project, section=sec).order_by("position", "id"))
            section_items_list.append((sec, items))

        return {
            "error": None,
            "all_mode": True,
            "section": sections[0],
            "section_items_list": section_items_list,
            "legal_entities": _legal_entities_for(project, asset_name),
        }

    section = _resolve_section(project, section_id, asset_name)
    if not section:
        return {
            "error": "Не удалось определить раздел",
            "all_mode": False,
            "section": None,
            "section_items_list": [],
            "legal_entities": [],
        }

    _ensure_checklist_items(project, section)
    checklist_items = list(ChecklistItem.objects.filter(project=project, section=section).order_by("position", "id"))
    return {
        "error": None,
        "all_mode": False,
        "section": section,
        "section_items_list": [(section, checklist_items)],
        "legal_entities": _legal_entities_for(project, asset_name),
    }


def _build_grid_payload(
    project: ProjectRegistration,
    asset_name: str,
    scope: dict,
    *,
    readonly: bool,
    show_actions: bool,
    create_url: str = "",
    xlsx_url: str = "",
    approve_info_request_url: str = "",
) -> dict:
    section_items_list = scope["section_items_list"]
    legal_entities = scope["legal_entities"]
    all_item_ids = [item.id for _section, items in section_items_list for item in items]

    status_map = {}
    if all_item_ids and legal_entities:
        status_qs = ChecklistStatus.objects.filter(
            checklist_item__in=all_item_ids,
            legal_entity__in=[entity.id for entity in legal_entities],
        ).select_related("checklist_item", "legal_entity")
        status_map = {(st.checklist_item_id, st.legal_entity_id): st for st in status_qs}

    comment_summary_map = _comment_summary_map(
        project=project,
        item_ids=all_item_ids,
        field=ChecklistCommentHistory.Field.IMC,
    )

    folder_map = {}
    if all_item_ids:
        try:
            from checklists_app.models import ChecklistItemFolder
            folder_qs = ChecklistItemFolder.objects.filter(checklist_item__in=all_item_ids)
            folder_map = {f.checklist_item_id: f for f in folder_qs}
        except Exception:
            pass

    status_label_map = dict(ChecklistStatus.Status.choices)
    rows = []
    for sec, items in section_items_list:
        if scope["all_mode"] and items:
            rows.append({
                "kind": "section_header",
                "sectionId": sec.id,
                "sectionName": sec.name_ru or str(sec),
            })

        seen_additional_groups = set()
        for item in items:
            if item.item_type == ChecklistItem.ItemType.ADDITIONAL and item.additional_number and item.additional_date:
                group_key = (item.additional_number, item.additional_date)
                if group_key not in seen_additional_groups:
                    seen_additional_groups.add(group_key)
                    rows.append({
                        "kind": "additional_header",
                        "sectionId": sec.id,
                        "text": (
                            f"Дополнительный запрос № {item.additional_number}"
                            f" от {item.additional_date.strftime('%d.%m.%Y')}"
                        ),
                    })

            row_statuses = []
            cells = []
            for entity in legal_entities:
                status_obj = status_map.get((item.id, entity.id))
                row_statuses.append({"entity": entity, "status": status_obj})
                value = status_obj.status if status_obj else ChecklistStatus.Status.MISSING
                cells.append({
                    "entityId": entity.id,
                    "status": value,
                    "statusLabel": status_label_map[value],
                    "statusShortLabel": _status_short_label(value),
                    "dateDisplay": _format_status_changed_at(status_obj),
                    "dateIso": timezone.localtime(status_obj.status_changed_at).isoformat() if status_obj and status_obj.status_changed_at else "",
                })

            folder = folder_map.get(item.id)
            rows.append({
                "kind": "item",
                "id": item.id,
                "sectionId": sec.id,
                "sectionName": sec.name_ru or str(sec),
                "code": item.code,
                "number": f"{item.number:02d}",
                "shortName": item.short_name,
                "name": item.name,
                "codeClass": _code_cell_class(row_statuses),
                "comments": comment_summary_map.get(item.id, _comment_flags(None)),
                "fileCount": folder.file_count if folder else None,
                "lastUploadAt": (
                    timezone.localtime(folder.last_upload_at).strftime("%d.%m.%y %H:%M")
                    if folder and folder.last_upload_at else None
                ),
                "cells": cells,
                "actions": {
                    "editUrl": reverse("checklists_app:item_form_edit", args=[item.id]) if show_actions else "",
                    "deleteUrl": reverse("checklists_app:item_delete", args=[item.id]) if show_actions else "",
                    "moveUpUrl": reverse("checklists_app:item_move", args=[item.id, "up"]) if show_actions else "",
                    "moveDownUrl": reverse("checklists_app:item_move", args=[item.id, "down"]) if show_actions else "",
                },
            })

    section = scope["section"]
    create_href = ""
    if create_url and section:
        create_href = f"{create_url}?project={project.id}"
        if scope["all_mode"]:
            create_href += "&all_mode=1"
        else:
            create_href += f"&section={section.id}"

    return {
        "project": {
            "id": project.id,
            "shortUid": (project.short_uid or "").strip(),
            "name": project.name,
        },
        "assetName": asset_name,
        "allMode": scope["all_mode"],
        "readonly": readonly,
        "error": scope["error"],
        "section": {
            "id": section.id,
            "name": section.name_ru or str(section),
        } if section else None,
        "entities": [
            {
                "id": entity.id,
                "label": entity.legal_name or entity.work_name or getattr(getattr(entity, "work_item", None), "asset_name", "") or "Юридическое лицо",
            }
            for entity in legal_entities
        ],
        "statusChoices": [
            {"value": value, "label": label, "shortLabel": _status_short_label(value)}
            for value, label in ChecklistStatus.Status.choices
        ],
        "rows": rows,
        "ui": {
            "showActions": show_actions,
            "createUrl": create_href,
            "xlsxUrl": xlsx_url,
            "approveInfoRequestUrl": approve_info_request_url,
            "projectUid": (project.short_uid or "").strip(),
        },
        "virtualization": {
            "enabled": False,
            "rowHeight": 44,
            "window": {
                "start": 0,
                "end": len(rows),
            },
        },
    }


def _status_update_payload(checklist_item, legal_entity, asset_name, status_obj):
    related_entities = _legal_entities_for(legal_entity.project, asset_name or None)
    status_map = {
        st.legal_entity_id: st
        for st in ChecklistStatus.objects.filter(
            checklist_item=checklist_item,
            legal_entity__in=[entity.id for entity in related_entities],
        )
    }
    row_statuses = [
        {"entity": entity, "status": status_map.get(entity.id)}
        for entity in related_entities
    ]
    value = status_obj.status if status_obj else ChecklistStatus.Status.MISSING
    return {
        "checklistItem": checklist_item.id,
        "legalEntity": legal_entity.id,
        "status": value,
        "statusLabel": dict(ChecklistStatus.Status.choices)[value],
        "statusShortLabel": _status_short_label(value),
        "dateDisplay": _format_status_changed_at(status_obj),
        "dateIso": timezone.localtime(status_obj.status_changed_at).isoformat() if status_obj and status_obj.status_changed_at else "",
        "codeClass": _code_cell_class(row_statuses),
    }


def _batch_update_statuses(*, request_user, asset_name, updates, shared_link=None):
    valid_values = {value for value, _ in ChecklistStatus.Status.choices}
    if not isinstance(updates, list) or not updates:
        raise ValueError("Нет изменений для сохранения.")

    results = []
    user = request_user if request_user and getattr(request_user, "is_authenticated", False) else None

    with transaction.atomic():
        for update in updates:
            item_id = str(update.get("checklist_item") or "").strip()
            legal_entity_id = str(update.get("legal_entity") or "").strip()
            status_value = str(update.get("status") or "").strip()

            if not all([item_id, legal_entity_id, status_value]):
                raise ValueError("Недостаточно данных.")
            if status_value not in valid_values:
                raise ValueError("Некорректный статус.")

            checklist_item = get_object_or_404(ChecklistItem, pk=item_id)
            legal_entity = get_object_or_404(LegalEntity, pk=legal_entity_id)
            if shared_link and (
                legal_entity.project_id != shared_link.project_id
                or checklist_item.project_id != shared_link.project_id
            ):
                raise ValueError("Нет доступа к этой записи.")

            existing = ChecklistStatus.objects.filter(
                checklist_item=checklist_item,
                legal_entity=legal_entity,
            ).first()
            if existing and existing.status == status_value:
                status_obj = existing
            elif not existing and status_value == ChecklistStatus.Status.MISSING:
                status_obj = None
            else:
                status_obj, created = ChecklistStatus.objects.select_for_update().get_or_create(
                    checklist_item=checklist_item,
                    legal_entity=legal_entity,
                    defaults={"updated_by": user},
                )
                previous = None if created else status_obj.status
                status_obj.status = status_value
                status_obj.updated_by = user
                status_obj.save(previous_status=previous)

                if previous != status_value:
                    ChecklistStatusHistory.objects.create(
                        checklist_status=status_obj,
                        checklist_item=checklist_item,
                        legal_entity=legal_entity,
                        previous_status=previous or "",
                        new_status=status_value,
                        changed_by=user,
                    )

            results.append(_status_update_payload(checklist_item, legal_entity, asset_name, status_obj))

    return results


def _text_update_payload(item: ChecklistItem) -> dict:
    return {
        "itemId": item.id,
        "shortName": item.short_name or "",
        "name": item.name or "",
    }


# ---------------------------------------------------------------------------
#  Internal (authenticated) views
# ---------------------------------------------------------------------------

@require_GET
def panel(request):
    project_options = _project_options()
    selected_project_uid = project_options[0]["short_uid"] if project_options else None
    selected_project = (
        ProjectRegistration.objects.select_related("type").filter(short_uid=selected_project_uid).first()
    )
    meta = _project_meta(selected_project, None) if selected_project else {"assets": [], "asset": "", "sections": []}

    try:
        table_url = reverse("checklists_app:table_partial")
        grid_url = reverse("checklists_app:grid_data")
        update_url = reverse("checklists_app:update_status")
        batch_update_url = reverse("checklists_app:update_status_batch")
        text_update_url_base = reverse("checklists_app:item_text_update", args=[0])
        note_url = reverse("checklists_app:update_note")
        comment_modal_url = reverse("checklists_app:comment_modal")
        meta_url_base = reverse("checklists_app:project_meta", args=["__uid__"])
        item_form_create_url = reverse("checklists_app:item_form_create")
        batch_edit_url = reverse("checklists_app:item_batch_edit")
    except NoReverseMatch:
        table_url = "/checklists/partial/table/"
        grid_url = "/checklists/grid/data/"
        update_url = "/checklists/status/update/"
        batch_update_url = "/checklists/status/batch-update/"
        text_update_url_base = "/checklists/item/text-update/0/"
        note_url = "/checklists/note/update/"
        comment_modal_url = "/checklists/comment/modal/"
        meta_url_base = "/checklists/project-meta/__uid__/"
        item_form_create_url = "/checklists/item/form/create/"
        batch_edit_url = "/checklists/item/batch-edit/"

    yadisk_url = ""
    if request.user.is_authenticated:
        yadisk_sel = YandexDiskSelection.objects.filter(user=request.user).first()
        if yadisk_sel and yadisk_sel.resource_path:
            path = yadisk_sel.resource_path.strip("/")
            yadisk_url = f"https://disk.yandex.ru/client/disk/{path}" if path else "https://disk.yandex.ru/client/disk"
        elif YandexDiskAccount.objects.filter(user=request.user).exists():
            yadisk_url = "https://disk.yandex.ru/client/disk"

    return render(
        request,
        "checklists_app/panel.html",
        {
            "project_options": project_options,
            "selected_project_uid": selected_project_uid,
            "asset_options": meta["assets"],
            "asset_items": meta.get("asset_items", []),
            "selected_asset": meta["asset"],
            "section_options": meta["sections"],
            "table_partial_url": table_url,
            "grid_data_url": grid_url,
            "update_status_url": update_url,
            "update_status_batch_url": batch_update_url,
            "item_text_update_url_base": text_update_url_base,
            "update_note_url": note_url,
            "comment_modal_url": comment_modal_url,
            "project_meta_url_base": meta_url_base,
            "item_form_create_url": item_form_create_url,
            "batch_edit_url": batch_edit_url,
            "yadisk_url": yadisk_url,
        },
    )


@require_GET
def project_meta(request, uid: str):
    project = get_object_or_404(ProjectRegistration.objects.select_related("type"), short_uid=uid)
    asset = (request.GET.get("asset") or "").strip() or None
    return JsonResponse(_project_meta(project, asset))


@require_GET
def grid_data(request):
    project_uid = (request.GET.get("project_uid") or request.GET.get("project") or "").strip()
    asset_name = (request.GET.get("asset") or "").strip()
    section_id = (request.GET.get("section") or "").strip()

    if not project_uid:
        return JsonResponse({"error": "Выберите проект.", "rows": [], "entities": [], "statusChoices": []})

    project = get_object_or_404(ProjectRegistration.objects.select_related("type"), short_uid=project_uid)
    scope = _resolve_grid_scope(project, asset_name, section_id)
    approve_url = ""
    info_request_approved_at = ""
    hide_edit = False
    if request.user.is_authenticated:
        from notifications_app.models import Notification
        has_pending = Notification.objects.filter(
            notification_type=Notification.NotificationType.PROJECT_INFO_REQUEST_APPROVAL,
            recipient=request.user,
            project=project,
            is_processed=False,
        ).exists()
        if has_pending:
            approve_url = reverse("checklists_app:approve_info_request")
        else:
            processed = Notification.objects.filter(
                notification_type=Notification.NotificationType.PROJECT_INFO_REQUEST_APPROVAL,
                recipient=request.user,
                project=project,
                is_processed=True,
            ).order_by("-action_at").first()
            if processed and processed.action_at:
                from django.utils import timezone
                local_dt = timezone.localtime(processed.action_at)
                info_request_approved_at = local_dt.strftime("%d.%m.%y %H:%M")
                hide_edit = True

    show_actions = bool(request.user.is_authenticated and request.user.is_staff) and not hide_edit

    payload = _build_grid_payload(
        project,
        asset_name,
        scope,
        readonly=False,
        show_actions=show_actions,
        create_url=reverse("checklists_app:item_form_create"),
        xlsx_url=reverse("checklists_app:export_xlsx") + f"?project_uid={project.short_uid}" if scope["all_mode"] else "",
        approve_info_request_url=approve_url,
    )
    if info_request_approved_at:
        payload["ui"]["infoRequestApprovedAt"] = info_request_approved_at
    return JsonResponse(payload)


@login_required
@require_GET
def comment_modal(request):
    field = (request.GET.get("field") or "").strip()
    item_id = request.GET.get("item")
    project_id = request.GET.get("project")
    section_id = request.GET.get("section")
    asset_name = (request.GET.get("asset") or "").strip()

    if field not in {"imc_comment", "customer_comment"}:
        return HttpResponseBadRequest("Неизвестное поле.")
    if not all([item_id, project_id, section_id]):
        return HttpResponseBadRequest("Недостаточно данных.")

    checklist_item = get_object_or_404(ChecklistItem, pk=item_id)
    project = get_object_or_404(ProjectRegistration, pk=project_id)
    section = get_object_or_404(TypicalSection, pk=section_id)

    return _render_comment_modal_content(
        request,
        field=field,
        checklist_item=checklist_item,
        project=project,
        section=section,
        asset_name=asset_name,
        add_comment_url=reverse("checklists_app:add_comment"),
        readonly=False,
    )


@login_required
@require_POST
def update_status_batch(request):
    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except json.JSONDecodeError:
        return JsonResponse({"ok": False, "error": "Некорректный JSON."}, status=400)

    asset_name = str(payload.get("asset_name") or "").strip()
    updates = payload.get("updates") or []
    try:
        results = _batch_update_statuses(
            request_user=request.user,
            asset_name=asset_name,
            updates=updates,
        )
    except ValueError as exc:
        return JsonResponse({"ok": False, "error": str(exc)}, status=400)

    return JsonResponse({"ok": True, "results": results})


@login_required
@require_POST
def item_text_update(request, pk: int):
    item = get_object_or_404(ChecklistItem, pk=pk)
    field = (request.POST.get("field") or "").strip()
    value = (request.POST.get("value") or "").strip()

    if field not in {"short_name", "name"}:
        return JsonResponse({"ok": False, "error": "Неизвестное поле."}, status=400)
    if field == "name" and not value:
        return JsonResponse({"ok": False, "error": "Наименование запроса не может быть пустым."}, status=400)

    setattr(item, field, value)
    item.save(update_fields=[field])
    return JsonResponse({"ok": True, "item": _text_update_payload(item)})


@require_GET
def table_partial(request):
    project_uid = (request.GET.get("project_uid") or request.GET.get("project") or "").strip()
    asset_name = (request.GET.get("asset") or "").strip()
    section_id = (request.GET.get("section") or "").strip()

    if not project_uid:
        return render(request, "checklists_app/status_table.html", {"project": None})

    project = get_object_or_404(
        ProjectRegistration.objects.select_related("type"),
        short_uid=project_uid,
    )
    if not project.type:
        return render(
            request,
            "checklists_app/status_table.html",
            {"project": project, "section": None, "error": "У проекта не указан тип продукта"},
        )

    all_mode = section_id == "all"

    effective_asset = _resolve_asset_name(project, asset_name)

    if all_mode:
        perf_qs = Performer.objects.filter(registration=project).exclude(asset_name="")
        if effective_asset and asset_name != "all":
            perf_qs = perf_qs.filter(asset_name=effective_asset)
        perf_qs = perf_qs.select_related("typical_section__product").order_by("position", "id")

        seen_ids = set()
        sections = []
        for perf in perf_qs:
            ts = perf.typical_section
            if ts and ts.id not in seen_ids:
                seen_ids.add(ts.id)
                sections.append(ts)
        if not sections:
            return render(request, "checklists_app/status_table.html", {
                "project": project, "section": None, "error": "Нет разделов для данного продукта",
            })

        all_items = []
        for sec in sections:
            _ensure_checklist_items(project, sec)
            items = list(
                ChecklistItem.objects.filter(project=project, section=sec).order_by("position", "id")
            )
            all_items.append((sec, items))

        legal_entities = _legal_entities_for(project, asset_name)
        first_section = sections[0]

        url_map = {
            "update_status": reverse("checklists_app:update_status"),
            "add_comment": reverse("checklists_app:add_comment"),
            "comment_modal": reverse("checklists_app:comment_modal"),
            "update_note": reverse("checklists_app:update_note"),
        }
        context = _build_all_sections_context(
            project, all_items, asset_name, legal_entities, url_map,
        )
        context["show_actions"] = True
        context["all_mode"] = True
        context["section"] = first_section
        try:
            context["item_form_create_url"] = reverse("checklists_app:item_form_create")
        except NoReverseMatch:
            context["item_form_create_url"] = ""
        return render(request, "checklists_app/status_table.html", context)

    section = _resolve_section(project, section_id, asset_name)
    if not section:
        return render(
            request,
            "checklists_app/status_table.html",
            {"project": project, "section": None, "error": "Не удалось определить раздел"},
        )

    _ensure_checklist_items(project, section)
    checklist_items = list(
        ChecklistItem.objects.filter(project=project, section=section).order_by("position", "id")
    )
    legal_entities = _legal_entities_for(project, asset_name)

    url_map = {
        "update_status": reverse("checklists_app:update_status"),
        "add_comment": reverse("checklists_app:add_comment"),
        "comment_modal": reverse("checklists_app:comment_modal"),
        "update_note": reverse("checklists_app:update_note"),
    }
    context = _build_table_context(project, section, asset_name, checklist_items, legal_entities, url_map)
    context["show_actions"] = True
    try:
        context["item_form_create_url"] = reverse("checklists_app:item_form_create")
    except NoReverseMatch:
        context["item_form_create_url"] = ""
    return render(request, "checklists_app/status_table.html", context)


def _render_status_cell(request, checklist_item, legal_entity, status_obj, asset_name,
                        update_url_name="checklists_app:update_status"):
    project = legal_entity.project
    related_entities = _legal_entities_for(project, asset_name or None)
    status_map = {
        st.legal_entity_id: st
        for st in ChecklistStatus.objects.filter(
            checklist_item=checklist_item,
            legal_entity__in=[le.id for le in related_entities],
        )
    }
    row_statuses = [
        {"entity": entity, "status": status_map.get(entity.id)}
        for entity in related_entities
    ]
    code_class = _code_cell_class(row_statuses)
    return render(
        request,
        "checklists_app/components/status_cell_fragment.html",
        {
            "item": checklist_item,
            "legal_entity": legal_entity,
            "status": status_obj,
            "status_choices": ChecklistStatus.Status.choices,
            "default_status": ChecklistStatus.Status.MISSING,
            "update_url": update_url_name if update_url_name.startswith("/") else reverse(update_url_name),
            "code_class": code_class,
            "asset_name": asset_name,
        },
    )


@login_required
@require_POST
def update_status(request):
    item_id = request.POST.get("checklist_item")
    legal_entity_id = request.POST.get("legal_entity")
    status_value = request.POST.get("status")
    asset_name = (request.POST.get("asset_name") or "").strip()

    if not all([item_id, legal_entity_id, status_value]):
        return HttpResponseBadRequest("Недостаточно данных.")

    valid_values = {value for value, _ in ChecklistStatus.Status.choices}
    if status_value not in valid_values:
        return HttpResponseBadRequest("Некорректный статус.")

    checklist_item = get_object_or_404(ChecklistItem, pk=item_id)
    legal_entity = get_object_or_404(LegalEntity, pk=legal_entity_id)

    existing = ChecklistStatus.objects.filter(
        checklist_item=checklist_item, legal_entity=legal_entity,
    ).first()
    if existing and existing.status == status_value:
        return _render_status_cell(request, checklist_item, legal_entity, existing, asset_name)
    if not existing and status_value == ChecklistStatus.Status.MISSING:
        return _render_status_cell(request, checklist_item, legal_entity, None, asset_name)

    with transaction.atomic():
        status_obj, created = ChecklistStatus.objects.select_for_update().get_or_create(
            checklist_item=checklist_item,
            legal_entity=legal_entity,
            defaults={"updated_by": request.user},
        )
        previous = None if created else status_obj.status
        status_obj.status = status_value
        status_obj.updated_by = request.user
        status_obj.save(previous_status=previous)

        if previous != status_value:
            ChecklistStatusHistory.objects.create(
                checklist_status=status_obj,
                checklist_item=checklist_item,
                legal_entity=legal_entity,
                previous_status=previous or "",
                new_status=status_value,
                changed_by=request.user,
            )

    return _render_status_cell(request, checklist_item, legal_entity, status_obj, asset_name)


@login_required
@require_POST
def update_note(request):
    field = request.POST.get("field")
    item_id = request.POST.get("checklist_item")
    project_id = request.POST.get("project")
    section_id = request.POST.get("section")
    asset_name = (request.POST.get("asset") or "").strip()
    value = (request.POST.get("value") or "").strip()

    if field not in {"imc_comment", "customer_comment"}:
        return HttpResponseBadRequest("Неизвестное поле.")
    if not all([item_id, project_id, section_id]):
        return HttpResponseBadRequest("Недостаточно данных.")

    project = get_object_or_404(ProjectRegistration, pk=project_id)
    section = get_object_or_404(TypicalSection, pk=section_id)
    checklist_item = get_object_or_404(ChecklistItem, pk=item_id)

    note, _ = ChecklistRequestNote.objects.get_or_create(
        checklist_item=checklist_item,
        project=project,
        section=section,
        asset_name=asset_name,
        defaults={"updated_by": request.user},
    )
    setattr(note, field, value)
    note.updated_by = request.user
    note.save(update_fields=[field, "updated_by", "updated_at"])

    return render(
        request,
        "checklists_app/components/comment_cell.html",
        {
            "field": field,
            "item": checklist_item,
            "project": project,
            "section": section,
            "asset_name": asset_name,
            "note": note,
            "comment_modal_url": reverse("checklists_app:comment_modal"),
        },
    )


@login_required
@require_POST
def add_comment(request):
    field = request.POST.get("field")
    item_id = request.POST.get("checklist_item")
    project_id = request.POST.get("project")
    section_id = request.POST.get("section")
    asset_name = (request.POST.get("asset") or "").strip()
    value = (request.POST.get("value") or "").strip()

    if field not in {"imc_comment", "customer_comment"}:
        return HttpResponseBadRequest("Неизвестное поле.")
    if not all([item_id, project_id, section_id]):
        return HttpResponseBadRequest("Недостаточно данных.")
    if not value:
        return HttpResponseBadRequest("Введите текст комментария.")

    checklist_item = get_object_or_404(ChecklistItem, pk=item_id)
    project = get_object_or_404(ProjectRegistration, pk=project_id)
    section = get_object_or_404(TypicalSection, pk=section_id)

    note, _ = ChecklistRequestNote.objects.get_or_create(
        checklist_item=checklist_item,
        project=project,
        section=section,
        asset_name=asset_name,
        defaults={"updated_by": request.user},
    )

    if field == "imc_comment":
        note.imc_comment = value
    else:
        note.customer_comment = value
    note.updated_by = request.user
    note.save(update_fields=[field, "updated_by", "updated_at"])

    ChecklistCommentHistory.objects.create(
        note=note,
        field=field,
        text=value,
        author=request.user,
    )

    history_qs = note.comment_history.filter(field=field).select_related("author").order_by("created_at", "id")
    history_html = render_to_string(
        "checklists_app/components/comment_thread.html",
        {
            "scope_id": _comment_scope_id(field, checklist_item.id, asset_name),
            "histories": history_qs,
            "oob": True,
        },
        request=request,
    )
    resp = HttpResponse(history_html)
    resp["HX-Trigger"] = json.dumps({
        "checklists:comment-updated": _comment_update_event_payload(note, field)
    })
    return resp


# ---------------------------------------------------------------------------
#  ChecklistItem CRUD
# ---------------------------------------------------------------------------

@login_required
@require_GET
def item_form_create(request):
    project_id = request.GET.get("project")
    section_id = request.GET.get("section")
    all_mode = request.GET.get("all_mode") == "1"
    project = get_object_or_404(ProjectRegistration, pk=project_id) if project_id else None
    section = get_object_or_404(TypicalSection, pk=section_id) if section_id else None

    sections_list = []
    if all_mode and project and project.type:
        perf_qs = (
            Performer.objects.filter(registration=project)
            .exclude(asset_name="")
            .select_related("typical_section__product")
            .order_by("position", "id")
        )
        seen = set()
        for perf in perf_qs:
            ts = perf.typical_section
            if ts and ts.id not in seen:
                seen.add(ts.id)
                sections_list.append(ts)
        if not section and sections_list:
            section = sections_list[0]

    code = section.code if section else ""
    next_num = 1
    has_additional = False
    if project and section:
        agg = ChecklistItem.objects.filter(project=project, section=section).aggregate(m=Max("number"))
        next_num = (agg["m"] or 0) + 1
        has_additional = ChecklistItem.objects.filter(
            project=project, section=section,
            item_type=ChecklistItem.ItemType.ADDITIONAL,
        ).exists()
    forced_type = "additional" if has_additional else ""
    return render(request, "checklists_app/checklist_item_form.html", {
        "project": project, "section": section,
        "code": code, "next_number": next_num,
        "item_type": "additional" if has_additional else "basic",
        "forced_type": forced_type,
        "sections_list": sections_list,
        "all_mode": all_mode,
        "form_action": reverse("checklists_app:item_create"),
    })


@login_required
@require_GET
def item_form_edit(request, pk):
    ci = get_object_or_404(ChecklistItem, pk=pk)
    has_additional = ChecklistItem.objects.filter(
        project=ci.project, section=ci.section,
        item_type=ChecklistItem.ItemType.ADDITIONAL,
    ).exists()
    forced_type = "additional" if has_additional else ""
    lock_number = has_additional and ci.item_type == ChecklistItem.ItemType.BASIC
    return render(request, "checklists_app/checklist_item_form.html", {
        "project": ci.project, "section": ci.section,
        "code": ci.code, "next_number": ci.number,
        "short_name": ci.short_name, "name": ci.name,
        "item_type": ci.item_type,
        "forced_type": forced_type,
        "lock_number": lock_number,
        "edit_mode": True, "item_id": ci.id,
        "form_action": reverse("checklists_app:item_update", args=[ci.pk]),
    })


@login_required
@require_POST
def item_create(request):
    project_id = request.POST.get("project")
    section_id = request.POST.get("section")
    code = (request.POST.get("code") or "").strip()
    number_raw = (request.POST.get("number") or "").strip()
    short_name = (request.POST.get("short_name") or "").strip()
    name = (request.POST.get("name") or "").strip()
    item_type = (request.POST.get("item_type") or "basic").strip()

    if not all([project_id, section_id, code, number_raw, name]):
        return HttpResponseBadRequest("Заполните все обязательные поля.")

    project = get_object_or_404(ProjectRegistration, pk=project_id)
    section = get_object_or_404(TypicalSection, pk=section_id)

    try:
        number = int(number_raw)
    except ValueError:
        return HttpResponseBadRequest("Некорректный номер.")

    max_pos = ChecklistItem.objects.filter(project=project, section=section).aggregate(m=Max("position"))["m"] or 0

    additional_date = None
    additional_number = None

    if item_type == ChecklistItem.ItemType.ADDITIONAL:
        today = date.today()
        existing_today = ChecklistItem.objects.filter(
            project=project, section=section,
            item_type=ChecklistItem.ItemType.ADDITIONAL,
            additional_date=today,
        ).first()
        if existing_today:
            additional_number = existing_today.additional_number
        else:
            max_add = ChecklistItem.objects.filter(
                project=project, section=section,
                item_type=ChecklistItem.ItemType.ADDITIONAL,
            ).aggregate(m=Max("additional_number"))["m"] or 0
            additional_number = max_add + 1
        additional_date = today

    if ChecklistItem.objects.filter(project=project, section=section, number=number).exists():
        return HttpResponseBadRequest(f"Запрос с номером {number} уже существует в данном разделе.")

    ChecklistItem.objects.create(
        project=project, section=section,
        code=code, number=number, short_name=short_name, name=name,
        position=max_pos + 1,
        item_type=item_type,
        additional_date=additional_date,
        additional_number=additional_number,
    )
    resp = HttpResponse(status=204)
    resp["HX-Trigger"] = "checklists:saved"
    return resp


@login_required
@require_POST
def item_update(request, pk):
    ci = get_object_or_404(ChecklistItem, pk=pk)
    code = (request.POST.get("code") or "").strip()
    number_raw = (request.POST.get("number") or "").strip()
    short_name = (request.POST.get("short_name") or "").strip()
    name = (request.POST.get("name") or "").strip()
    new_type = (request.POST.get("item_type") or "basic").strip()

    if not all([code, number_raw, name]):
        return HttpResponseBadRequest("Заполните все обязательные поля.")

    try:
        new_number = int(number_raw)
    except ValueError:
        return HttpResponseBadRequest("Некорректный номер.")

    if new_number != ci.number and ChecklistItem.objects.filter(
        project=ci.project, section=ci.section, number=new_number,
    ).exclude(pk=ci.pk).exists():
        return HttpResponseBadRequest(f"Запрос с номером {new_number} уже существует в данном разделе.")

    ci.number = new_number
    ci.code = code
    ci.short_name = short_name
    ci.name = name

    old_type = ci.item_type
    ci.item_type = new_type

    if new_type == ChecklistItem.ItemType.ADDITIONAL and old_type != ChecklistItem.ItemType.ADDITIONAL:
        today = date.today()
        existing_today = ChecklistItem.objects.filter(
            project=ci.project, section=ci.section,
            item_type=ChecklistItem.ItemType.ADDITIONAL,
            additional_date=today,
        ).exclude(pk=ci.pk).first()
        if existing_today:
            ci.additional_number = existing_today.additional_number
        else:
            max_add = ChecklistItem.objects.filter(
                project=ci.project, section=ci.section,
                item_type=ChecklistItem.ItemType.ADDITIONAL,
            ).exclude(pk=ci.pk).aggregate(m=Max("additional_number"))["m"] or 0
            ci.additional_number = max_add + 1
        ci.additional_date = today
    elif new_type == ChecklistItem.ItemType.BASIC:
        ci.additional_date = None
        ci.additional_number = None

    ci.save(update_fields=[
        "code", "number", "short_name", "name",
        "item_type", "additional_date", "additional_number",
    ])
    resp = HttpResponse(status=204)
    resp["HX-Trigger"] = "checklists:saved"
    return resp


@login_required
@require_GET
def item_check_number(request):
    project_id = request.GET.get("project", "")
    section_id = request.GET.get("section", "")
    number_raw = request.GET.get("number", "").strip()
    exclude_id = request.GET.get("exclude", "").strip()

    if not all([project_id, section_id, number_raw]):
        return JsonResponse({"exists": False})

    try:
        number = int(number_raw)
    except ValueError:
        return JsonResponse({"exists": False})

    qs = ChecklistItem.objects.filter(
        project_id=project_id, section_id=section_id, number=number,
    )
    if exclude_id and exclude_id.isdigit():
        qs = qs.exclude(pk=int(exclude_id))
    return JsonResponse({"exists": qs.exists()})


@login_required
@require_POST
def item_delete(request, pk):
    ci = get_object_or_404(ChecklistItem, pk=pk)
    ci.delete()
    resp = HttpResponse(status=204)
    resp["HX-Trigger"] = "checklists:saved"
    return resp


@login_required
@require_POST
def item_move(request, pk, direction):
    ci = get_object_or_404(ChecklistItem, pk=pk)

    group_filter = Q(project=ci.project, section=ci.section)
    if ci.item_type == ChecklistItem.ItemType.ADDITIONAL and ci.additional_number is not None:
        group_filter &= Q(item_type=ChecklistItem.ItemType.ADDITIONAL, additional_number=ci.additional_number)
    else:
        group_filter &= Q(item_type=ChecklistItem.ItemType.BASIC)

    siblings = ChecklistItem.objects.filter(group_filter).order_by("position", "id")
    items = list(siblings)
    idx = next((i for i, x in enumerate(items) if x.pk == ci.pk), None)
    if idx is None:
        return HttpResponse(status=204)

    if direction == "up" and idx > 0:
        swap = items[idx - 1]
    elif direction == "down" and idx < len(items) - 1:
        swap = items[idx + 1]
    else:
        return HttpResponse(status=204)

    ci.position, swap.position = swap.position, ci.position
    ci.save(update_fields=["position"])
    swap.save(update_fields=["position"])
    resp = HttpResponse(status=204)
    resp["HX-Trigger"] = "checklists:saved"
    return resp


@login_required
@require_POST
def item_batch_edit(request):
    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({"ok": False, "error": "Некорректный JSON."}, status=400)

    deleted_ids = data.get("deleted") or []
    order_list = data.get("order") or []
    text_updates = data.get("text_updates") or []

    with transaction.atomic():
        if deleted_ids:
            ChecklistItem.objects.filter(pk__in=deleted_ids).delete()

        for entry in order_list:
            item_id = entry.get("id")
            position = entry.get("position")
            if item_id is not None and position is not None:
                ChecklistItem.objects.filter(pk=item_id).update(position=position)

        for entry in text_updates:
            item_id = entry.get("id")
            field = (entry.get("field") or "").strip()
            value = (entry.get("value") or "").strip()
            if not item_id or field not in ("short_name", "name"):
                continue
            if field == "name" and not value:
                continue
            ChecklistItem.objects.filter(pk=item_id).update(**{field: value})

    resp = HttpResponse(status=204)
    resp["HX-Trigger"] = "checklists:saved"
    return resp


# ---------------------------------------------------------------------------
#  XLSX export
# ---------------------------------------------------------------------------

def _generate_xlsx(project: ProjectRegistration) -> HttpResponse:
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    meta = _project_meta(project, None)
    assets = meta["assets"] or [""]

    perf_qs = (
        Performer.objects.filter(registration=project)
        .exclude(asset_name="")
        .select_related("typical_section__product")
        .order_by("position", "id")
    )
    seen_sec = set()
    sections = []
    for perf in perf_qs:
        ts = perf.typical_section
        if ts and ts.id not in seen_sec:
            seen_sec.add(ts.id)
            sections.append(ts)

    if not sections:
        return HttpResponseBadRequest("Нет разделов для экспорта.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Статусы запросов"

    header_font = Font(bold=True, size=10)
    header_fill = PatternFill(start_color="E8EEF6", end_color="E8EEF6", fill_type="solid")
    section_font = Font(bold=True, size=10)
    section_fill = PatternFill(start_color="E8EEF6", end_color="E8EEF6", fill_type="solid")
    add_header_font = Font(bold=True, size=10)
    thin_border = Border(
        left=Side(style="thin", color="DEE2E6"),
        right=Side(style="thin", color="DEE2E6"),
        top=Side(style="thin", color="DEE2E6"),
        bottom=Side(style="thin", color="DEE2E6"),
    )
    wrap_align = Alignment(wrap_text=True, vertical="top")
    top_align = Alignment(vertical="top")

    all_legal = {}
    for asset_name in assets:
        entities = _legal_entities_for(project, asset_name)
        for ent in entities:
            if ent.id not in all_legal:
                label = (
                    (ent.legal_name or "").strip()
                    or (ent.work_name or "").strip()
                    or (getattr(ent.work_item, "asset_name", "") or "").strip()
                    or "Юр. лицо"
                )
                all_legal[ent.id] = {"entity": ent, "label": label}

    entity_order = list(all_legal.keys())
    entity_labels = [all_legal[eid]["label"] for eid in entity_order]

    headers = ["Код", "№", "Краткое наименование", "Наименование запроса"]
    for label in entity_labels:
        headers.append(label)
        headers.append("Дата")
    headers.append("Комментарий IMC Montan")
    headers.append("Комментарий Заказчика")

    for ci, val in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=val)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = top_align

    row_idx = 2

    all_item_ids = []
    section_items_map = {}
    for sec in sections:
        _ensure_checklist_items(project, sec)
        items = list(
            ChecklistItem.objects.filter(project=project, section=sec).order_by("position", "id")
        )
        section_items_map[sec.id] = items
        all_item_ids.extend(it.id for it in items)

    status_map = {}
    if all_item_ids and entity_order:
        for st in ChecklistStatus.objects.filter(
            checklist_item__in=all_item_ids,
            legal_entity__in=entity_order,
        ).select_related("checklist_item", "legal_entity"):
            status_map[(st.checklist_item_id, st.legal_entity_id)] = st

    note_map = {}
    if all_item_ids:
        for note in ChecklistRequestNote.objects.filter(checklist_item__in=all_item_ids, project=project):
            note_map[note.checklist_item_id] = note

    for sec in sections:
        items = section_items_map.get(sec.id, [])
        if not items:
            continue

        num_cols = len(headers)
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=num_cols)
        cell = ws.cell(row=row_idx, column=1, value=sec.name_ru or str(sec))
        cell.font = section_font
        cell.fill = section_fill
        cell.border = thin_border
        cell.alignment = top_align
        for ci in range(2, num_cols + 1):
            ws.cell(row=row_idx, column=ci).border = thin_border
            ws.cell(row=row_idx, column=ci).fill = section_fill
        row_idx += 1

        seen_additional_groups = set()
        for item in items:
            if item.item_type == ChecklistItem.ItemType.ADDITIONAL and item.additional_number and item.additional_date:
                group_key = (item.additional_number, item.additional_date)
                if group_key not in seen_additional_groups:
                    seen_additional_groups.add(group_key)
                    add_text = (
                        f"Дополнительный запрос № {item.additional_number}"
                        f" от {item.additional_date.strftime('%d.%m.%Y')}"
                    )
                    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=num_cols)
                    cell = ws.cell(row=row_idx, column=1, value=add_text)
                    cell.font = add_header_font
                    cell.border = thin_border
                    cell.alignment = top_align
                    for ci in range(2, num_cols + 1):
                        ws.cell(row=row_idx, column=ci).border = thin_border
                    row_idx += 1

            col = 1
            ws.cell(row=row_idx, column=col, value=item.code).border = thin_border
            ws.cell(row=row_idx, column=col).alignment = top_align
            col += 1
            ws.cell(row=row_idx, column=col, value=f"{item.number:02d}").border = thin_border
            ws.cell(row=row_idx, column=col).alignment = top_align
            col += 1
            ws.cell(row=row_idx, column=col, value=item.short_name).border = thin_border
            ws.cell(row=row_idx, column=col).alignment = top_align
            col += 1
            c = ws.cell(row=row_idx, column=col, value=item.name)
            c.border = thin_border
            c.alignment = wrap_align
            col += 1

            for eid in entity_order:
                st = status_map.get((item.id, eid))
                status_label = ""
                status_date = ""
                if st:
                    status_label = st.get_status_display()
                    if st.status_changed_at:
                        status_date = timezone.localtime(st.status_changed_at).strftime("%d.%m.%y %H:%M")
                ws.cell(row=row_idx, column=col, value=status_label).border = thin_border
                ws.cell(row=row_idx, column=col).alignment = top_align
                col += 1
                ws.cell(row=row_idx, column=col, value=status_date).border = thin_border
                ws.cell(row=row_idx, column=col).alignment = top_align
                col += 1

            note = note_map.get(item.id)
            c = ws.cell(row=row_idx, column=col, value=(note.imc_comment if note else ""))
            c.border = thin_border
            c.alignment = wrap_align
            col += 1
            c = ws.cell(row=row_idx, column=col, value=(note.customer_comment if note else ""))
            c.border = thin_border
            c.alignment = wrap_align
            row_idx += 1

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 6
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 40
    for i in range(len(entity_labels)):
        status_col = 5 + i * 2
        date_col = 6 + i * 2
        ws.column_dimensions[get_column_letter(status_col)].width = 18
        ws.column_dimensions[get_column_letter(date_col)].width = 14
    if len(headers) >= 5 + len(entity_labels) * 2 + 1:
        ws.column_dimensions[get_column_letter(len(headers) - 1)].width = 30
        ws.column_dimensions[get_column_letter(len(headers))].width = 30

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    project_name = (project.name or project.short_uid or "project").replace(" ", "_")
    filename = f"checklist_{project_name}.xlsx"

    response = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
@require_GET
def export_xlsx(request):
    project_uid = (request.GET.get("project_uid") or "").strip()
    if not project_uid:
        return HttpResponseBadRequest("Не указан проект.")
    project = get_object_or_404(
        ProjectRegistration.objects.select_related("type"), short_uid=project_uid,
    )
    return _generate_xlsx(project)


# ---------------------------------------------------------------------------
#  Public shared links
# ---------------------------------------------------------------------------

def _get_shared_link(token: str) -> Optional[SharedChecklistLink]:
    try:
        link = SharedChecklistLink.objects.select_related("project", "project__type").get(token=token)
    except SharedChecklistLink.DoesNotExist:
        return None
    if link.is_expired:
        return None
    return link


@login_required
@require_POST
def shared_link_create_or_get(request):
    project_uid = (request.POST.get("project_uid") or "").strip()
    if not project_uid:
        return JsonResponse({"ok": False, "error": "Не указан проект."}, status=400)

    project = get_object_or_404(ProjectRegistration, short_uid=project_uid)
    link = SharedChecklistLink.objects.filter(
        project=project, expires_at__gt=timezone.now()
    ).order_by("-created_at").first()

    if not link:
        link = SharedChecklistLink.objects.create(
            project=project,
            created_by=request.user,
        )

    public_url = request.build_absolute_uri(
        reverse("checklists_app:shared_page", args=[link.token])
    )
    return JsonResponse({
        "ok": True,
        "link_id": link.id,
        "token": link.token,
        "url": public_url,
        "permission": link.permission,
        "expires_at": link.expires_at.strftime("%Y-%m-%d"),
    })


@login_required
@require_POST
def shared_link_update(request):
    link_id = request.POST.get("link_id")
    if not link_id:
        return JsonResponse({"ok": False, "error": "Не указан ID ссылки."}, status=400)

    link = get_object_or_404(SharedChecklistLink, pk=link_id)
    permission = request.POST.get("permission")
    expires_at = request.POST.get("expires_at")

    if permission and permission in dict(SharedChecklistLink.Permission.choices):
        link.permission = permission

    if expires_at:
        try:
            link.expires_at = timezone.make_aware(
                timezone.datetime.strptime(expires_at, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
            )
        except (ValueError, TypeError):
            pass

    link.save(update_fields=["permission", "expires_at", "updated_at"])
    return JsonResponse({"ok": True})


@ensure_csrf_cookie
@require_GET
def shared_page(request, token: str):
    link = _get_shared_link(token)
    if not link:
        return render(request, "checklists_app/shared_expired.html", status=410)

    project = link.project
    meta = _project_meta(project, None) if project else {"assets": [], "asset": "", "sections": []}

    project_options = [{
        "id": project.id,
        "short_uid": project.short_uid or "",
        "label": " ".join(
            x for x in (project.short_uid or "", _product_short_label(getattr(project, "type", None)), project.name) if x
        ),
        "code": (project.short_uid or "").strip() or f"{project.number}{project.group}".upper(),
        "product_short": _product_short_label(getattr(project, "type", None)),
        "product_id": getattr(getattr(project, "type", None), "id", None),
    }]

    try:
        table_url = reverse("checklists_app:shared_table_partial", args=[token])
        grid_url = reverse("checklists_app:shared_grid_data", args=[token])
        meta_url_base = reverse("checklists_app:shared_project_meta", args=[token])
        update_url = reverse("checklists_app:shared_update_status", args=[token])
        batch_update_url = reverse("checklists_app:shared_update_status_batch", args=[token])
        text_update_url_base = reverse("checklists_app:shared_item_text_update", args=[token, 0])
        note_url = reverse("checklists_app:shared_update_note", args=[token])
        comment_modal_url = reverse("checklists_app:shared_comment_modal", args=[token])
    except NoReverseMatch:
        table_url = f"/checklists/shared/{token}/table/"
        grid_url = f"/checklists/shared/{token}/grid/"
        meta_url_base = f"/checklists/shared/{token}/meta/"
        update_url = f"/checklists/shared/{token}/status/update/"
        batch_update_url = f"/checklists/shared/{token}/status/batch-update/"
        text_update_url_base = f"/checklists/shared/{token}/item/text-update/0/"
        note_url = f"/checklists/shared/{token}/note/update/"
        comment_modal_url = f"/checklists/shared/{token}/comment/modal/"

    return render(request, "checklists_app/shared_page.html", {
        "link": link,
        "project": project,
        "project_options": project_options,
        "selected_project_uid": project.short_uid or "",
        "asset_options": meta["assets"],
        "asset_items": meta.get("asset_items", []),
        "selected_asset": meta["asset"],
        "section_options": meta["sections"],
        "table_partial_url": table_url,
        "grid_data_url": grid_url,
        "update_status_url": update_url,
        "update_status_batch_url": batch_update_url,
        "item_text_update_url_base": text_update_url_base,
        "update_note_url": note_url,
        "comment_modal_url": comment_modal_url,
        "project_meta_url_base": meta_url_base,
        "can_edit": link.can_edit,
        "token": token,
    })


@require_GET
def shared_project_meta(request, token: str):
    link = _get_shared_link(token)
    if not link:
        return JsonResponse({"error": "Ссылка недействительна."}, status=403)
    asset = (request.GET.get("asset") or "").strip() or None
    return JsonResponse(_project_meta(link.project, asset))


@require_GET
def shared_grid_data(request, token: str):
    link = _get_shared_link(token)
    if not link:
        return JsonResponse({"error": "Ссылка недействительна или срок действия истёк.", "rows": []}, status=403)

    project = link.project
    asset_name = (request.GET.get("asset") or "").strip()
    section_id = (request.GET.get("section") or "").strip()
    scope = _resolve_grid_scope(project, asset_name, section_id)
    payload = _build_grid_payload(
        project,
        asset_name,
        scope,
        readonly=not link.can_edit,
        show_actions=False,
        xlsx_url=reverse("checklists_app:shared_export_xlsx", args=[token]) if scope["all_mode"] else "",
    )
    return JsonResponse(payload)


@require_GET
def shared_comment_modal(request, token: str):
    link = _get_shared_link(token)
    if not link:
        return HttpResponseBadRequest("Ссылка недействительна или срок действия истёк.")

    field = (request.GET.get("field") or "").strip()
    item_id = request.GET.get("item")
    project_id = request.GET.get("project")
    section_id = request.GET.get("section")
    asset_name = (request.GET.get("asset") or "").strip()

    if field not in {"imc_comment", "customer_comment"}:
        return HttpResponseBadRequest("Неизвестное поле.")
    if not all([item_id, project_id, section_id]):
        return HttpResponseBadRequest("Недостаточно данных.")

    checklist_item = get_object_or_404(ChecklistItem, pk=item_id)
    project = get_object_or_404(ProjectRegistration, pk=project_id)
    if project.id != link.project_id or checklist_item.project_id != link.project_id:
        return HttpResponseBadRequest("Нет доступа к этому проекту.")
    section = get_object_or_404(TypicalSection, pk=section_id)

    return _render_comment_modal_content(
        request,
        field=field,
        checklist_item=checklist_item,
        project=project,
        section=section,
        asset_name=asset_name,
        add_comment_url=reverse("checklists_app:shared_add_comment", args=[token]),
        readonly=not link.can_edit,
    )


@require_POST
def shared_update_status_batch(request, token: str):
    link = _get_shared_link(token)
    if not link or not link.can_edit:
        return JsonResponse({"ok": False, "error": "Нет прав на редактирование или ссылка недействительна."}, status=403)

    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except json.JSONDecodeError:
        return JsonResponse({"ok": False, "error": "Некорректный JSON."}, status=400)

    asset_name = str(payload.get("asset_name") or "").strip()
    updates = payload.get("updates") or []
    try:
        results = _batch_update_statuses(
            request_user=request.user,
            asset_name=asset_name,
            updates=updates,
            shared_link=link,
        )
    except ValueError as exc:
        return JsonResponse({"ok": False, "error": str(exc)}, status=400)

    return JsonResponse({"ok": True, "results": results})


@require_POST
def shared_item_text_update(request, token: str, pk: int):
    link = _get_shared_link(token)
    if not link or not link.can_edit:
        return JsonResponse({"ok": False, "error": "Нет прав на редактирование или ссылка недействительна."}, status=403)

    item = get_object_or_404(ChecklistItem, pk=pk)
    if item.project_id != link.project_id:
        return JsonResponse({"ok": False, "error": "Нет доступа к этой записи."}, status=403)

    field = (request.POST.get("field") or "").strip()
    value = (request.POST.get("value") or "").strip()

    if field not in {"short_name", "name"}:
        return JsonResponse({"ok": False, "error": "Неизвестное поле."}, status=400)
    if field == "name" and not value:
        return JsonResponse({"ok": False, "error": "Наименование запроса не может быть пустым."}, status=400)

    setattr(item, field, value)
    item.save(update_fields=[field])
    return JsonResponse({"ok": True, "item": _text_update_payload(item)})


@require_GET
def shared_table_partial(request, token: str):
    link = _get_shared_link(token)
    if not link:
        return HttpResponse('<div class="alert alert-danger">Ссылка недействительна или срок действия истёк.</div>')

    project = link.project
    asset_name = (request.GET.get("asset") or "").strip()
    section_id = (request.GET.get("section") or "").strip()

    if not project.type:
        return render(request, "checklists_app/status_table.html", {
            "project": project, "section": None,
            "error": "У проекта не указан тип продукта",
            "readonly": not link.can_edit,
        })

    all_mode = section_id == "all"

    effective_asset = _resolve_asset_name(project, asset_name)

    if all_mode:
        perf_qs = Performer.objects.filter(registration=project).exclude(asset_name="")
        if effective_asset and asset_name != "all":
            perf_qs = perf_qs.filter(asset_name=effective_asset)
        perf_qs = perf_qs.select_related("typical_section__product").order_by("position", "id")

        seen_ids = set()
        sections = []
        for perf in perf_qs:
            ts = perf.typical_section
            if ts and ts.id not in seen_ids:
                seen_ids.add(ts.id)
                sections.append(ts)

        if not sections:
            return render(request, "checklists_app/status_table.html", {
                "project": project, "section": None,
                "error": "Нет разделов для данного продукта",
                "readonly": not link.can_edit,
            })

        all_items = []
        for sec in sections:
            _ensure_checklist_items(project, sec)
            items = list(
                ChecklistItem.objects.filter(project=project, section=sec).order_by("position", "id")
            )
            all_items.append((sec, items))

        legal_entities = _legal_entities_for(project, asset_name)
        url_map = {
            "update_status": reverse("checklists_app:shared_update_status", args=[token]),
            "add_comment": reverse("checklists_app:shared_add_comment", args=[token]),
            "comment_modal": reverse("checklists_app:shared_comment_modal", args=[token]),
            "update_note": reverse("checklists_app:shared_update_note", args=[token]),
        }
        context = _build_all_sections_context(project, all_items, asset_name, legal_entities, url_map)
        context["readonly"] = not link.can_edit
        context["all_mode"] = True
        context["section"] = sections[0]
        context["xlsx_url"] = reverse("checklists_app:shared_export_xlsx", args=[token])
        return render(request, "checklists_app/status_table.html", context)

    section = _resolve_section(project, section_id, asset_name)
    if not section:
        return render(request, "checklists_app/status_table.html", {
            "project": project, "section": None,
            "error": "Не удалось определить раздел",
            "readonly": not link.can_edit,
        })

    _ensure_checklist_items(project, section)
    checklist_items = list(
        ChecklistItem.objects.filter(project=project, section=section).order_by("position", "id")
    )
    legal_entities = _legal_entities_for(project, asset_name)

    url_map = {
        "update_status": reverse("checklists_app:shared_update_status", args=[token]),
        "add_comment": reverse("checklists_app:shared_add_comment", args=[token]),
        "comment_modal": reverse("checklists_app:shared_comment_modal", args=[token]),
        "update_note": reverse("checklists_app:shared_update_note", args=[token]),
    }
    context = _build_table_context(project, section, asset_name, checklist_items, legal_entities, url_map)
    context["readonly"] = not link.can_edit
    return render(request, "checklists_app/status_table.html", context)


@require_POST
def shared_update_status(request, token: str):
    link = _get_shared_link(token)
    if not link or not link.can_edit:
        return HttpResponseBadRequest("Нет прав на редактирование или ссылка недействительна.")

    item_id = request.POST.get("checklist_item")
    legal_entity_id = request.POST.get("legal_entity")
    status_value = request.POST.get("status")
    asset_name = (request.POST.get("asset_name") or "").strip()

    if not all([item_id, legal_entity_id, status_value]):
        return HttpResponseBadRequest("Недостаточно данных.")

    valid_values = {v for v, _ in ChecklistStatus.Status.choices}
    if status_value not in valid_values:
        return HttpResponseBadRequest("Некорректный статус.")

    checklist_item = get_object_or_404(ChecklistItem, pk=item_id)
    legal_entity = get_object_or_404(LegalEntity, pk=legal_entity_id)

    if legal_entity.project_id != link.project_id:
        return HttpResponseBadRequest("Нет доступа к этому юридическому лицу.")

    update_url = reverse("checklists_app:shared_update_status", args=[token])

    existing = ChecklistStatus.objects.filter(
        checklist_item=checklist_item, legal_entity=legal_entity,
    ).first()
    if existing and existing.status == status_value:
        return _render_status_cell(request, checklist_item, legal_entity, existing, asset_name, update_url)
    if not existing and status_value == ChecklistStatus.Status.MISSING:
        return _render_status_cell(request, checklist_item, legal_entity, None, asset_name, update_url)

    user = request.user if request.user.is_authenticated else None
    with transaction.atomic():
        status_obj, created = ChecklistStatus.objects.select_for_update().get_or_create(
            checklist_item=checklist_item,
            legal_entity=legal_entity,
            defaults={"updated_by": user},
        )
        previous = None if created else status_obj.status
        status_obj.status = status_value
        status_obj.updated_by = user
        status_obj.save(previous_status=previous)

        if previous != status_value:
            ChecklistStatusHistory.objects.create(
                checklist_status=status_obj,
                checklist_item=checklist_item,
                legal_entity=legal_entity,
                previous_status=previous or "",
                new_status=status_value,
                changed_by=user,
            )

    return _render_status_cell(request, checklist_item, legal_entity, status_obj, asset_name, update_url)


@require_POST
def shared_update_note(request, token: str):
    link = _get_shared_link(token)
    if not link or not link.can_edit:
        return HttpResponseBadRequest("Нет прав на редактирование или ссылка недействительна.")

    field = request.POST.get("field")
    item_id = request.POST.get("checklist_item")
    project_id = request.POST.get("project")
    section_id = request.POST.get("section")
    asset_name = (request.POST.get("asset") or "").strip()
    value = (request.POST.get("value") or "").strip()

    if field not in {"imc_comment", "customer_comment"}:
        return HttpResponseBadRequest("Неизвестное поле.")
    if not all([item_id, project_id, section_id]):
        return HttpResponseBadRequest("Недостаточно данных.")

    project = get_object_or_404(ProjectRegistration, pk=project_id)
    if project.id != link.project_id:
        return HttpResponseBadRequest("Нет доступа к этому проекту.")

    section = get_object_or_404(TypicalSection, pk=section_id)
    checklist_item = get_object_or_404(ChecklistItem, pk=item_id)

    note, _ = ChecklistRequestNote.objects.get_or_create(
        checklist_item=checklist_item, project=project, section=section, asset_name=asset_name,
        defaults={"updated_by": request.user if request.user.is_authenticated else None},
    )
    setattr(note, field, value)
    note.updated_by = request.user if request.user.is_authenticated else None
    note.save(update_fields=[field, "updated_by", "updated_at"])

    return render(request, "checklists_app/components/comment_cell.html", {
        "field": field, "item": checklist_item, "project": project,
        "section": section, "asset_name": asset_name, "note": note,
        "comment_modal_url": reverse("checklists_app:shared_comment_modal", args=[token]),
    })


@require_POST
def shared_add_comment(request, token: str):
    link = _get_shared_link(token)
    if not link or not link.can_edit:
        return HttpResponseBadRequest("Нет прав на редактирование или ссылка недействительна.")

    field = request.POST.get("field")
    item_id = request.POST.get("checklist_item")
    project_id = request.POST.get("project")
    section_id = request.POST.get("section")
    asset_name = (request.POST.get("asset") or "").strip()
    value = (request.POST.get("value") or "").strip()

    if field not in {"imc_comment", "customer_comment"}:
        return HttpResponseBadRequest("Неизвестное поле.")
    if not all([item_id, project_id, section_id]):
        return HttpResponseBadRequest("Недостаточно данных.")
    if not value:
        return HttpResponseBadRequest("Введите текст комментария.")

    checklist_item = get_object_or_404(ChecklistItem, pk=item_id)
    project = get_object_or_404(ProjectRegistration, pk=project_id)
    if project.id != link.project_id:
        return HttpResponseBadRequest("Нет доступа к этому проекту.")

    section = get_object_or_404(TypicalSection, pk=section_id)

    note, _ = ChecklistRequestNote.objects.get_or_create(
        checklist_item=checklist_item, project=project, section=section, asset_name=asset_name,
        defaults={"updated_by": request.user if request.user.is_authenticated else None},
    )

    if field == "imc_comment":
        note.imc_comment = value
    else:
        note.customer_comment = value
    note.updated_by = request.user if request.user.is_authenticated else None
    note.save(update_fields=[field, "updated_by", "updated_at"])

    ChecklistCommentHistory.objects.create(
        note=note, field=field, text=value,
        author=request.user if request.user.is_authenticated else None,
    )

    history_qs = note.comment_history.filter(field=field).select_related("author").order_by("created_at", "id")
    history_html = render_to_string(
        "checklists_app/components/comment_thread.html",
        {
            "scope_id": _comment_scope_id(field, checklist_item.id, asset_name),
            "histories": history_qs,
            "oob": True,
        },
        request=request,
    )
    resp = HttpResponse(history_html)
    resp["HX-Trigger"] = json.dumps({
        "checklists:comment-updated": _comment_update_event_payload(note, field)
    })
    return resp


@require_GET
def shared_export_xlsx(request, token: str):
    link = _get_shared_link(token)
    if not link:
        return HttpResponseBadRequest("Ссылка недействительна или срок действия истёк.")
    return _generate_xlsx(link.project)


@login_required
@require_POST
def approve_info_request(request):
    from notifications_app.models import Notification
    from notifications_app.services import process_info_request_notification

    project_uid = (request.POST.get("project_uid") or "").strip()
    if not project_uid:
        return JsonResponse({"ok": False, "error": "Не указан проект."}, status=400)

    project = get_object_or_404(ProjectRegistration.objects.select_related("type"), short_uid=project_uid)

    notifications = list(
        Notification.objects.filter(
            notification_type=Notification.NotificationType.PROJECT_INFO_REQUEST_APPROVAL,
            recipient=request.user,
            project=project,
            is_processed=False,
        )
    )
    if not notifications:
        return JsonResponse({"ok": False, "error": "Нет ожидающих согласования запросов."}, status=400)

    for notification in notifications:
        process_info_request_notification(notification, request.user)

    from django.utils import timezone
    now = timezone.localtime(timezone.now())
    return JsonResponse({
        "ok": True,
        "processed": len(notifications),
        "approved_at": now.strftime("%d.%m.%y %H:%M"),
    })
