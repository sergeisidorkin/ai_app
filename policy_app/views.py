import csv
import copy
import io
import json
import calendar
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from zipfile import BadZipFile

from docx.opc.exceptions import PackageNotFoundError
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException

from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db import transaction
from django.db.models import IntegerField, Max, Q, Value
from django.db.models.functions import Coalesce
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.views.decorators.http import require_http_methods, require_POST

from experts_app.models import ExpertProfile, ExpertSpecialty
from group_app.models import GroupMember, OrgUnit
from .models import (
    ConsultingDirection,
    ConsultingDirectionType,
    ConsultingServiceSubtype,
    ConsultingServiceType,
    Product,
    SYSTEM_DSC_SECTION_CODE,
    TypicalSection,
    TypicalSectionSpecialty,
    SectionStructure,
    ServiceGoalReport,
    TypicalServiceComposition,
    TypicalServiceTerm,
    ExpertiseDirection,
    Grade,
    SpecialtyTariff,
    Tariff,
    MANAGER_GROUPS,
    build_consulting_catalog_meta,
    ensure_system_dsc_section,
    is_system_dsc_code,
)
from .forms import (
    ConsultingDirectionForm,
    ProductForm,
    TypicalSectionForm,
    SectionStructureForm,
    ServiceGoalReportForm,
    TypicalServiceCompositionForm,
    TypicalServiceTermForm,
    ExpertiseDirectionForm,
    GradeForm,
    SpecialtyTariffForm,
    TariffForm,
)
from .docx_service_compositions import (
    build_typical_service_compositions_docx,
    parse_typical_service_compositions_docx,
)

# Вынесенные константы для единообразия шаблонов/заголовков
POLICY_PARTIAL_TEMPLATE = "policy_app/policy_partial.html"
PRODUCT_FORM_TEMPLATE = "policy_app/product_form.html"
SECTION_FORM_TEMPLATE = "policy_app/section_form.html"
STRUCTURE_FORM_TEMPLATE = "policy_app/structure_form.html"
SERVICE_GOAL_REPORT_FORM_TEMPLATE = "policy_app/service_goal_report_form.html"
TYPICAL_SERVICE_COMPOSITION_FORM_TEMPLATE = "policy_app/typical_service_composition_form.html"
TYPICAL_SERVICE_TERM_FORM_TEMPLATE = "policy_app/typical_service_term_form.html"
EXPERTISE_DIR_FORM_TEMPLATE = "policy_app/expertise_direction_form.html"
CONSULTING_DIR_FORM_TEMPLATE = "policy_app/consulting_direction_form.html"
GRADE_FORM_TEMPLATE = "policy_app/grade_form.html"
SPECIALTY_TARIFF_FORM_TEMPLATE = "policy_app/specialty_tariff_form.html"
TARIFF_FORM_TEMPLATE = "policy_app/tariff_form.html"
HX_TRIGGER_HEADER = "HX-Trigger"
HX_POLICY_UPDATED_EVENT = "policy-updated"
PRODUCT_CSV_HEADERS = [
    "Краткое имя",
    "Наименование на английском языке",
    "Наименование на русском языке",
    "Отображаемое в системе имя",
    "Вид консалтинга",
    "Тип услуг",
    "Код",
    "Подтип услуги",
    "Владелец",
]
SERVICE_GOAL_REPORT_CSV_HEADERS = [
    "Продукт",
    "Цели оказания услуг",
    "Цели оказания услуг в родительном падеже",
    "Титул отчета/ТКП",
    "Название продукта",
]
STRUCTURE_CSV_HEADERS = [
    "Продукт",
    "Раздел (услуга)",
    "Подразделы",
]
TYPICAL_SERVICE_COMPOSITION_CSV_HEADERS = [
    "Продукт",
    "Раздел (услуга)",
    "Состав услуг",
]
TYPICAL_SERVICE_COMPOSITION_XLSX_HEADERS = [
    "Продукт",
    "Раздел (услуга)",
    "Состав услуг",
]
TYPICAL_SERVICE_COMPOSITION_EDITOR_STATE_HEADER = "Состояние редактора (JSON)"
TYPICAL_SERVICE_TERM_CSV_HEADERS = [
    "Продукт",
    "Сроки предоставления исходных данных, нед.",
    "Срок подготовки Предварительного отчёта, мес.",
    "Срок подготовки Итогового отчёта, нед.",
]
TYPICAL_SERVICE_TERM_GANTT_VERSION = 1
TYPICAL_SERVICE_TERM_GANTT_SERVICE_SECTION_TYPE = "service_section"
TYPICAL_SERVICE_TERM_GANTT_CALENDAR_KIND_ABSTRACT = "abstract"
TYPICAL_SERVICE_TERM_GANTT_EXECUTOR_DISPLAY_EXECUTOR = "executor"
TYPICAL_SERVICE_TERM_GANTT_EXECUTOR_DISPLAY_RESOURCE = "resource_name"
TYPICAL_SERVICE_TERM_GANTT_SYSTEM_TASK_TEXT = {
    "source_data": "Исходные данные",
    "source_data_asset": "Актив",
    "preliminary_report": "Предварительный отчёт",
    "preliminary_report_asset": "Актив",
    "preliminary_report_submission": "Отправка Предварительного отчёта",
    "final_report": "Итоговый отчёт",
}
TARIFF_CSV_HEADERS = [
    "Продукт",
    "Раздел (услуга)",
    "Базовая ставка в ВПМ",
    "Объем услуг в часах",
    "Объем услуг в днях для ТКП",
    "Руководитель направления",
]
SECTION_CSV_HEADERS = [
    "Продукт",
    "Код",
    "Краткое имя EN",
    "Краткое имя RU",
    "Наименование раздела (услуги) EN",
    "Наименование раздела (услуги) RU",
    "Тип учета",
    "Исполнитель",
    "Экспертиза",
    "Подразделение",
    "ТКП",
]


def _render_form_with_errors(request, template, context):
    response = render(request, template, context)
    response["HX-Retarget"] = "#policy-modal .modal-content"
    response["HX-Reswap"] = "innerHTML"
    return response


def _is_department_head(user):
    return user.groups.filter(name__in=MANAGER_GROUPS).exists()


def _get_grades_for_user(user):
    qs = Grade.objects.select_related("created_by", "created_by__employee_profile", "currency")
    if user.is_superuser:
        return qs
    if _is_department_head(user):
        return qs.filter(created_by=user)
    return qs


def _get_tariffs_for_user(user):
    qs = Tariff.objects.select_related(
        "product",
        "product__consulting_type_ref",
        "product__service_category_ref",
        "product__service_subtype_ref",
        "section",
        "created_by",
        "created_by__employee_profile",
    ).annotate(
        owner_group_position=Coalesce(
            "created_by__employee_profile__position",
            Value(1000000),
            output_field=IntegerField(),
        )
    ).order_by(
        "owner_group_position",
        "created_by__employee_profile__job_title",
        "created_by__username",
        "position",
        "id",
    )
    if user.is_superuser:
        return qs
    if _is_department_head(user):
        return qs.filter(created_by=user)
    return qs


def _get_specialty_tariffs_for_user(user):
    qs = SpecialtyTariff.objects.select_related(
        "currency", "created_by", "created_by__employee_profile"
    ).prefetch_related("specialties")
    if user.is_superuser:
        return qs
    if _is_department_head(user):
        return qs.filter(created_by=user)
    return qs

def staff_required(user):
    return user.is_authenticated and user.is_staff


def _csv_lookup_key(value):
    return str(value or "").strip().lower()


def _split_csv_list_value(value, lookup=None):
    raw = str(value or "").strip()
    if not raw or raw in {"—", "-"}:
        return []
    lookup = lookup or {}
    if _csv_lookup_key(raw) in lookup:
        return [raw]

    values = []
    for line in raw.replace("\r\n", "\n").replace("\r", "\n").split("\n"):
        parts = [line]
        for separator in (",", ";", "|"):
            parts = [
                chunk
                for part in parts
                for chunk in (
                    [part] if _csv_lookup_key(part) in lookup else part.split(separator)
                )
            ]
        values.extend(
            item.strip()
            for item in parts
            if item.strip() and item.strip() not in {"—", "-"}
        )
    return values


def _csv_truthy(value):
    return _csv_lookup_key(value) in {"1", "true", "yes", "y", "да", "д", "on", "checked", "истина", "+", "x", "✓"}

# Вспомогательные функции для устранения дублирования
def _policy_context(request):
    products = Product.objects.select_related(
        "consulting_type_ref", "service_category_ref", "service_subtype_ref"
    ).prefetch_related("owners").all()
    sections = TypicalSection.objects.select_related(
        "product",
        "product__consulting_type_ref",
        "product__service_category_ref",
        "product__service_subtype_ref",
        "expertise_dir",
        "expertise_direction",
    ).prefetch_related(
        "ranked_specialties", "ranked_specialties__specialty"
    ).all()
    structures = SectionStructure.objects.select_related(
        "product",
        "product__consulting_type_ref",
        "product__service_category_ref",
        "product__service_subtype_ref",
        "section",
    ).all()
    service_goal_reports = ServiceGoalReport.objects.select_related(
        "product",
        "product__consulting_type_ref",
        "product__service_category_ref",
        "product__service_subtype_ref",
    ).all()
    typical_service_compositions = TypicalServiceComposition.objects.select_related(
        "product",
        "product__consulting_type_ref",
        "product__service_category_ref",
        "product__service_subtype_ref",
        "section",
    ).all()
    typical_service_terms = TypicalServiceTerm.objects.select_related(
        "product",
        "product__consulting_type_ref",
        "product__service_category_ref",
        "product__service_subtype_ref",
    ).all()
    consulting_directions = ConsultingDirection.objects.prefetch_related(
        "consulting_types",
        "service_types__consulting_type",
        "service_subtypes__service_type__consulting_type",
    ).all()
    expertise_directions = ExpertiseDirection.objects.prefetch_related("owners").all()
    grades = _get_grades_for_user(request.user)
    specialty_tariffs = _get_specialty_tariffs_for_user(request.user)
    tariffs = _get_tariffs_for_user(request.user)
    is_dept_head = _is_department_head(request.user)
    return {
        "products": products,
        "sections": sections,
        "structures": structures,
        "service_goal_reports": service_goal_reports,
        "typical_service_compositions": typical_service_compositions,
        "typical_service_terms": typical_service_terms,
        "consulting_directions": consulting_directions,
        "expertise_directions": expertise_directions,
        "grades": grades,
        "specialty_tariffs": specialty_tariffs,
        "tariffs": tariffs,
        "is_admin": request.user.is_superuser,
        "is_dept_head": is_dept_head,
    }

def _render_policy_updated(request):
    response = render(request, POLICY_PARTIAL_TEMPLATE, _policy_context(request))
    response[HX_TRIGGER_HEADER] = HX_POLICY_UPDATED_EVENT
    return response


def _consulting_catalog_lookup():
    consulting_types = list(ConsultingDirectionType.objects.order_by("position", "id"))
    service_types = list(
        ConsultingServiceType.objects.select_related("consulting_type").order_by(
            "consulting_type__position", "position", "id"
        )
    )
    service_subtypes = list(
        ConsultingServiceSubtype.objects.select_related("service_type", "service_type__consulting_type").order_by(
            "service_type__consulting_type__position", "service_type__position", "position", "id"
        )
    )
    return {
        "consulting_types_by_name": {item.name: item for item in consulting_types},
        "service_types_by_pair": {
            (item.consulting_type.name, item.name): item for item in service_types
        },
        "service_subtypes_by_triple": {
            (item.service_type.consulting_type.name, item.service_type.name, item.name): item
            for item in service_subtypes
        },
    }


def _consulting_direction_form_context(form, action, direction=None):
    ctx = {
        "form": form,
        "action": action,
        "consulting_catalog_initial_json": json.dumps(form.initial_catalog, ensure_ascii=False),
    }
    if direction:
        ctx["direction"] = direction
    return ctx

def _product_form_page_context(extra: dict) -> dict:
    ctx = {
        "product_service_meta_json": json.dumps(
            build_consulting_catalog_meta(),
            ensure_ascii=False,
        ),
    }
    ctx.update(extra)
    return ctx


def _positive_int(value):
    try:
        parsed = int(str(value).strip())
    except (TypeError, ValueError):
        return None
    return parsed if parsed > 0 else None


def _prefill_product_from_request(request):
    product_id = _positive_int(request.GET.get("product"))
    if not product_id:
        return None
    return (
        Product.objects.select_related(
            "consulting_type_ref",
            "service_category_ref",
            "service_subtype_ref",
        )
        .filter(pk=product_id)
        .first()
    )


def _product_field_initial_from_request(request):
    product = _prefill_product_from_request(request)
    return {"product": product.pk} if product else {}


def _product_ref_initial_from_request(request):
    product = _prefill_product_from_request(request)
    if product:
        return {
            "consulting_type_ref": product.consulting_type_ref_id,
            "service_category_ref": product.service_category_ref_id,
            "service_subtype_ref": product.service_subtype_ref_id,
        }
    return {
        key: value
        for key in ("consulting_type_ref", "service_category_ref", "service_subtype_ref")
        if (value := _positive_int(request.GET.get(key)))
    }


def _next_position(model, filters: dict | None = None) -> int:
    """
    Возвращает следующую позицию (last+1) для списка объектов.
    filters — при необходимости ограничивает область (например, внутри продукта).
    """
    qs = model.objects
    if filters:
        qs = qs.filter(**filters)
    last = qs.aggregate(mx=Max("position")).get("mx") or 0
    return last + 1

@login_required
@require_http_methods(["GET"])
def policy_partial(request):
    return render(request, POLICY_PARTIAL_TEMPLATE, _policy_context(request))

@login_required
@user_passes_test(staff_required)
@require_http_methods(["GET", "POST"])
def product_form_create(request):
    if request.method == "GET":
        form = ProductForm(initial=_product_ref_initial_from_request(request))
        return render(request, PRODUCT_FORM_TEMPLATE, _product_form_page_context({"form": form, "action": "create"}))
    # POST
    form = ProductForm(request.POST)
    if not form.is_valid():
        return _render_form_with_errors(
            request, PRODUCT_FORM_TEMPLATE, _product_form_page_context({"form": form, "action": "create"})
        )
    if not form.instance.position:
        form.instance.position = _next_position(Product)
    product = form.save()
    ensure_system_dsc_section(product)
    return _render_policy_updated(request)

@login_required
@user_passes_test(staff_required)
@require_http_methods(["GET", "POST"])
def product_form_edit(request, pk: int):
    product = get_object_or_404(Product, pk=pk)
    if request.method == "GET":
        form = ProductForm(instance=product)
        return render(
            request,
            PRODUCT_FORM_TEMPLATE,
            _product_form_page_context({"form": form, "action": "edit", "product": product}),
        )
    # POST
    form = ProductForm(request.POST, instance=product)
    if not form.is_valid():
        return _render_form_with_errors(
            request,
            PRODUCT_FORM_TEMPLATE,
            _product_form_page_context({"form": form, "action": "edit", "product": product}),
        )
    product = form.save()
    ensure_system_dsc_section(product)
    return _render_policy_updated(request)

@login_required
@user_passes_test(staff_required)
@require_POST
def product_delete(request, pk: int):
    product = get_object_or_404(Product, pk=pk)
    product.delete()
    return _render_policy_updated(request)


@login_required
@user_passes_test(staff_required)
@require_POST
def product_csv_upload(request):
    csv_file = request.FILES.get("csv_file")
    if not csv_file:
        return JsonResponse({"ok": False, "error": "Файл не выбран."}, status=400)
    if not csv_file.name.lower().endswith(".csv"):
        return JsonResponse({"ok": False, "error": "Допустимы только файлы CSV."}, status=400)

    try:
        raw = csv_file.read().decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            csv_file.seek(0)
            raw = csv_file.read().decode("cp1251")
        except Exception:
            return JsonResponse(
                {"ok": False, "error": "Не удалось прочитать файл. Проверьте кодировку (UTF-8 или Windows-1251)."},
                status=400,
            )

    try:
        reader = csv.reader(io.StringIO(raw), delimiter=";")
        rows = list(reader)
        if not rows:
            return JsonResponse({"ok": False, "error": "Файл пуст."}, status=400)
        if len(rows[0]) <= 1:
            reader = csv.reader(io.StringIO(raw), delimiter=",")
            rows = list(reader)
    except csv.Error as exc:
        return JsonResponse({"ok": False, "error": f"Ошибка разбора CSV: {exc}. Проверьте формат и кодировку файла."}, status=400)

    if len(rows) < 2:
        return JsonResponse({"ok": False, "error": "Файл должен содержать заголовок и хотя бы одну строку данных."}, status=400)

    catalog_lookup = _consulting_catalog_lookup()
    valid_consulting_types = set(catalog_lookup["consulting_types_by_name"].keys())
    valid_service_categories = {name for _kind, name in catalog_lookup["service_types_by_pair"].keys()}
    owners_by_short_name = {
        member.short_name.strip().lower(): member
        for member in GroupMember.objects.exclude(short_name="").all()
    }

    created = 0
    warnings = []

    for i, row in enumerate(rows[1:], start=2):
        if not any(cell.strip() for cell in row):
            continue
        if len(row) < 8:
            warnings.append(
                f"Строка {i}: недостаточно столбцов ({len(row)}, ожидается 8-9: "
                "Краткое имя, Наименование EN, Наименование RU, Отображаемое имя, Вид консалтинга, "
                "Тип услуг, Код, Подтип услуги, [Владелец])."
            )
            continue

        short_name = row[0].strip()
        name_en = row[1].strip()
        name_ru = row[2].strip()
        display_name = row[3].strip() if len(row) > 3 else ""
        consulting_type = row[4].strip() if len(row) > 4 else ""
        service_category = row[5].strip() if len(row) > 5 else ""
        csv_code = row[6].strip() if len(row) > 6 else ""
        service_subtype = row[7].strip() if len(row) > 7 else ""
        owner_raw = row[8].strip() if len(row) > 8 else ""

        missing = []
        if not short_name:
            missing.append("Краткое имя")
        if not name_en:
            missing.append("Наименование EN")
        if not name_ru:
            missing.append("Наименование RU")
        if not consulting_type:
            missing.append("Вид консалтинга")
        if not service_category:
            missing.append("Тип услуг")
        if not service_subtype:
            missing.append("Подтип услуги")
        if missing:
            warnings.append(f"Строка {i}: не заполнены обязательные поля: {', '.join(missing)}.")
            continue

        if consulting_type not in valid_consulting_types:
            warnings.append(
                f"Строка {i}: неизвестный вид консалтинга «{consulting_type}». "
                f"Допустимые: {', '.join(valid_consulting_types)}."
            )
            continue

        if service_category not in valid_service_categories:
            warnings.append(
                f"Строка {i}: неизвестный тип услуг «{service_category}». "
                f"Допустимые: {', '.join(valid_service_categories)}."
            )
            continue

        service_type_obj = catalog_lookup["service_types_by_pair"].get((consulting_type, service_category))
        if service_type_obj is None:
            warnings.append(
                f"Строка {i}: тип услуг «{service_category}» недопустим для вида консалтинга "
                f"«{consulting_type}»."
            )
            continue

        service_subtype_obj = catalog_lookup["service_subtypes_by_triple"].get(
            (consulting_type, service_category, service_subtype)
        )
        if service_subtype_obj is None:
            warnings.append(
                f"Строка {i}: подтип услуги «{service_subtype}» недопустим для типа услуг "
                f"«{service_category}»."
            )
            continue

        derived_code = service_type_obj.code or ""
        if csv_code and csv_code != derived_code:
            warnings.append(
                f"Строка {i}: код «{csv_code}» не соответствует типу услуг «{service_category}». "
                f"Использован код «{derived_code}»."
            )

        is_group_owner = not owner_raw or owner_raw == "Группа"
        owner_ids = []
        if not is_group_owner:
            owner_names = [item.strip() for item in owner_raw.split(",") if item.strip()]
            missing_owners = [name for name in owner_names if name.lower() not in owners_by_short_name]
            if missing_owners:
                warnings.append(f"Строка {i}: владельцы не найдены: {', '.join(missing_owners)}.")
                continue
            owner_ids = [owners_by_short_name[name.lower()].pk for name in owner_names]

        try:
            product = Product.objects.create(
                short_name=short_name,
                name_en=name_en,
                display_name=display_name,
                name_ru=name_ru,
                consulting_type_ref=service_type_obj.consulting_type,
                service_category_ref=service_type_obj,
                service_subtype_ref=service_subtype_obj,
                is_group_owner=is_group_owner,
                position=_next_position(Product),
            )
            if owner_ids:
                product.owners.set(owner_ids)
            ensure_system_dsc_section(product)
            created += 1
        except Exception as exc:
            warnings.append(f"Строка {i}: ошибка сохранения — {exc}")

    return JsonResponse({"ok": True, "created": created, "warnings": warnings})


@login_required
@user_passes_test(staff_required)
@require_http_methods(["GET"])
def product_csv_download(request):
    output = io.StringIO()
    output.write("\ufeff")
    writer = csv.writer(output, delimiter=";", lineterminator="\n")
    writer.writerow(PRODUCT_CSV_HEADERS)

    products = _apply_policy_master_filters_to_products(
        Product.objects.select_related(
            "consulting_type_ref", "service_category_ref", "service_subtype_ref"
        ).prefetch_related("owners"),
        request,
    )
    for product in products:
        writer.writerow(
            [
                product.short_name,
                product.name_en,
                product.name_ru,
                product.display_name,
                product.consulting_type_display,
                product.service_category_display,
                product.service_code,
                product.service_subtype_display,
                product.owner_display,
            ]
        )

    response = HttpResponse(output.getvalue(), content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="typical_products.csv"'
    return response

# --- Типовые разделы ---

def _section_specialty_options():
    return [
        {"id": s.pk, "label": s.specialty}
        for s in ExpertSpecialty.objects.exclude(specialty="").order_by("position")
    ]


def _section_form_context(form, action, section=None):
    spec_options = _section_specialty_options()
    ranked = []
    if section and section.pk:
        ranked = [
            {"rank": link.rank, "specialty_id": link.specialty_id}
            for link in TypicalSectionSpecialty.objects.filter(section=section).order_by("rank")
        ]
    ctx = {
        "form": form,
        "action": action,
        "specialty_options": spec_options,
        "specialty_options_json": json.dumps(spec_options, ensure_ascii=False),
        "ranked_specialties": ranked,
        "is_system_section": bool(section and section.is_system_dsc),
    }
    if section:
        ctx["section"] = section
    return ctx


def _save_section_specialties(section, post_data):
    specialty_ids = post_data.getlist("specialty_id")
    TypicalSectionSpecialty.objects.filter(section=section).delete()
    to_create = []
    for rank, raw_id in enumerate(specialty_ids, start=1):
        if raw_id:
            try:
                sid = int(raw_id)
            except (ValueError, TypeError):
                continue
            to_create.append(TypicalSectionSpecialty(
                section=section, specialty_id=sid, rank=rank,
            ))
    if to_create:
        TypicalSectionSpecialty.objects.bulk_create(to_create)


def _typical_sections_by_product_json():
    data = defaultdict(list)
    sections = TypicalSection.objects.select_related("product").order_by("product_id", "position", "id")
    for section in sections:
        data[str(section.product_id)].append({
            "id": section.pk,
            "label": section.name_ru,
        })
    return json.dumps(data, ensure_ascii=False)


def _typical_service_composition_form_context(form, action, composition=None):
    ctx = {
        "form": form,
        "action": action,
        "sections_by_product_json": _typical_sections_by_product_json(),
    }
    if composition:
        ctx["composition"] = composition
    return ctx


def _typical_service_term_form_context(form, action, term=None):
    ctx = {
        "form": form,
        "action": action,
    }
    if term:
        ctx["term"] = term
    return ctx


def _structure_form_context(form, action, structure=None):
    ctx = {
        "form": form,
        "action": action,
        "sections_by_product_json": _typical_sections_by_product_json(),
    }
    if structure:
        ctx["structure"] = structure
    return ctx


def _specialty_tariff_specialty_options():
    return [
        {
            "id": specialty.pk,
            "label": specialty.specialty,
            "expertise_direction": (
                ""
                if (getattr(specialty.expertise_dir, "short_name", "") or "").strip() == "—"
                else (getattr(specialty.expertise_dir, "short_name", "") or "").strip()
            ),
        }
        for specialty in ExpertSpecialty.objects.exclude(specialty="").select_related("expertise_dir").order_by("position", "id")
    ]


def _specialty_tariff_form_context(form, action, tariff=None):
    selected_specialty_ids = []
    if form.is_bound:
        selected_specialty_ids = [str(value) for value in form.data.getlist("specialties") if value]
    elif tariff and tariff.pk:
        selected_specialty_ids = [str(value) for value in tariff.specialties.values_list("pk", flat=True)]

    ctx = {
        "form": form,
        "action": action,
        "specialty_options": _specialty_tariff_specialty_options(),
        "specialty_options_json": json.dumps(_specialty_tariff_specialty_options(), ensure_ascii=False),
        "selected_specialty_ids_json": json.dumps(selected_specialty_ids, ensure_ascii=False),
    }
    if tariff:
        ctx["tariff"] = tariff
    return ctx


def _specialty_tariff_owner(request, form):
    if request.user.is_superuser:
        owner = form.cleaned_data.get("owner")
        if owner:
            return owner
    return request.user


@login_required
@user_passes_test(staff_required)
@require_http_methods(["GET", "POST"])
def section_form_create(request):
    if request.method == "GET":
        form = TypicalSectionForm(initial=_product_field_initial_from_request(request))
        return render(request, SECTION_FORM_TEMPLATE, _section_form_context(form, "create"))
    form = TypicalSectionForm(request.POST)
    if not form.is_valid():
        return render(request, SECTION_FORM_TEMPLATE, _section_form_context(form, "create"))
    obj = form.save(commit=False)
    ensure_system_dsc_section(obj.product)
    if not getattr(obj, "position", 0):
        obj.position = _next_position(TypicalSection, {"product": obj.product})
    obj.save()
    _save_section_specialties(obj, request.POST)
    ensure_system_dsc_section(obj.product)
    return _render_policy_updated(request)

@login_required
@user_passes_test(staff_required)
@require_http_methods(["GET", "POST"])
def section_form_edit(request, pk: int):
    section = get_object_or_404(TypicalSection, pk=pk)
    if request.method == "GET":
        form = TypicalSectionForm(instance=section)
        return render(request, SECTION_FORM_TEMPLATE, _section_form_context(form, "edit", section))
    form = TypicalSectionForm(request.POST, instance=section)
    if not form.is_valid():
        return render(request, SECTION_FORM_TEMPLATE, _section_form_context(form, "edit", section))
    if section.is_system_dsc:
        ensure_system_dsc_section(section.product)
    else:
        form.save()
        _save_section_specialties(section, request.POST)
        ensure_system_dsc_section(section.product)
    return _render_policy_updated(request)

@login_required
@user_passes_test(staff_required)
@require_POST
def section_delete(request, pk: int):
    section = get_object_or_404(TypicalSection, pk=pk)
    if section.is_system_dsc:
        return JsonResponse({"ok": False, "error": "Системный раздел DSC нельзя удалить."}, status=400)
    product = section.product
    section.delete()
    ensure_system_dsc_section(product)
    return _render_policy_updated(request)

def _normalize_product_positions():
    """
    Гарантирует сквозную нумерацию позиций (1..N) в порядке текущей сортировки.
    """
    products = Product.objects.order_by("position", "id").only("id", "position")
    for idx, p in enumerate(products, start=1):
        if p.position != idx:
            Product.objects.filter(pk=p.pk).update(position=idx)

def _normalize_section_positions(product_id: int | None = None):
    """
    Гарантирует сквозную нумерацию позиций разделов внутри каждого продукта.
    Если product_id задан, нормализует только для одного продукта.
    """
    qs = TypicalSection.objects.select_related("product").only("id", "position", "product_id", "code", "is_system")
    if product_id:
        groups = {product_id: list(qs.filter(product_id=product_id).order_by("position", "id"))}
    else:
        # группируем по продукту
        groups = {}
        for sec in qs.order_by("product_id", "position", "id"):
            groups.setdefault(sec.product_id, []).append(sec)
    for pid, items in groups.items():
        items = sorted(items, key=lambda item: (0 if item.is_system_dsc else 1, item.position, item.id))
        for idx, it in enumerate(items, start=1):
            if it.position != idx:
                TypicalSection.objects.filter(pk=it.pk).update(position=idx)

@require_http_methods(["POST", "GET"])
@login_required
def product_move_up(request, pk: int):
    _normalize_product_positions()
    items = list(Product.objects.order_by("position", "id").only("id", "position"))
    idx = next((i for i, it in enumerate(items) if it.id == pk), None)
    if idx is not None and idx > 0:
        cur = items[idx]
        prev = items[idx - 1]
        cur_pos, prev_pos = cur.position, prev.position
        Product.objects.filter(pk=cur.id).update(position=prev_pos)
        Product.objects.filter(pk=prev.id).update(position=cur_pos)
        _normalize_product_positions()
    return _render_policy_updated(request)

@require_http_methods(["POST", "GET"])
@login_required
def product_move_down(request, pk: int):
    _normalize_product_positions()
    items = list(Product.objects.order_by("position", "id").only("id", "position"))
    idx = next((i for i, it in enumerate(items) if it.id == pk), None)
    if idx is not None and idx < len(items) - 1:
        cur = items[idx]
        nxt = items[idx + 1]
        cur_pos, next_pos = cur.position, nxt.position
        Product.objects.filter(pk=cur.id).update(position=next_pos)
        Product.objects.filter(pk=nxt.id).update(position=cur_pos)
        _normalize_product_positions()
    return _render_policy_updated(request)

@require_http_methods(["POST", "GET"])
@login_required
def section_move_up(request, pk: int):
    sec = get_object_or_404(TypicalSection, pk=pk)
    if sec.is_system_dsc:
        return _render_policy_updated(request)
    pid = sec.product_id
    _normalize_section_positions(product_id=pid)
    items = list(TypicalSection.objects.filter(product_id=pid).order_by("position", "id").only("id", "position", "code", "is_system"))
    idx = next((i for i, it in enumerate(items) if it.id == pk), None)
    if idx is not None and idx > 0:
        cur = items[idx]
        prev = items[idx - 1]
        if prev.is_system_dsc:
            return _render_policy_updated(request)
        cur_pos, prev_pos = cur.position, prev.position
        TypicalSection.objects.filter(pk=cur.id).update(position=prev_pos)
        TypicalSection.objects.filter(pk=prev.id).update(position=cur_pos)
        _normalize_section_positions(product_id=pid)
    return _render_policy_updated(request)

@require_http_methods(["POST", "GET"])
@login_required
def section_move_down(request, pk: int):
    sec = get_object_or_404(TypicalSection, pk=pk)
    if sec.is_system_dsc:
        return _render_policy_updated(request)
    pid = sec.product_id
    _normalize_section_positions(product_id=pid)
    items = list(TypicalSection.objects.filter(product_id=pid).order_by("position", "id").only("id", "position"))
    idx = next((i for i, it in enumerate(items) if it.id == pk), None)
    if idx is not None and idx < len(items) - 1:
        cur = items[idx]
        nxt = items[idx + 1]
        cur_pos, next_pos = cur.position, nxt.position
        TypicalSection.objects.filter(pk=cur.id).update(position=next_pos)
        TypicalSection.objects.filter(pk=nxt.id).update(position=cur_pos)
        _normalize_section_positions(product_id=pid)
    return _render_policy_updated(request)


@login_required
@user_passes_test(staff_required)
@require_POST
def section_csv_upload(request):
    csv_file = request.FILES.get("csv_file")
    if not csv_file:
        return JsonResponse({"ok": False, "error": "Файл не выбран."}, status=400)
    if not csv_file.name.lower().endswith(".csv"):
        return JsonResponse({"ok": False, "error": "Допустимы только файлы CSV."}, status=400)

    try:
        raw = csv_file.read().decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            csv_file.seek(0)
            raw = csv_file.read().decode("cp1251")
        except Exception:
            return JsonResponse({"ok": False, "error": "Не удалось прочитать файл. Проверьте кодировку (UTF-8 или Windows-1251)."}, status=400)

    try:
        reader = csv.reader(io.StringIO(raw), delimiter=";")
        rows = list(reader)
        if not rows:
            return JsonResponse({"ok": False, "error": "Файл пуст."}, status=400)
        if len(rows[0]) <= 1:
            reader = csv.reader(io.StringIO(raw), delimiter=",")
            rows = list(reader)
    except csv.Error as exc:
        return JsonResponse({"ok": False, "error": f"Ошибка разбора CSV: {exc}. Проверьте формат и кодировку файла."}, status=400)
    if len(rows) < 2:
        return JsonResponse({"ok": False, "error": "Файл должен содержать заголовок и хотя бы одну строку данных."}, status=400)

    products_by_name = {_csv_lookup_key(p.short_name): p for p in Product.objects.all()}
    expertise_dirs = {}
    for direction in ExpertiseDirection.objects.all():
        for label in (direction.short_name, direction.name):
            key = _csv_lookup_key(label)
            if key:
                expertise_dirs[key] = direction
    expertise_units = {
        key: u
        for u in OrgUnit.objects.filter(Q(unit_type="expertise") | Q(unit_type="administrative", level=1))
        for key in {_csv_lookup_key(u.department_name), _csv_lookup_key(u.short_name)}
        if key
    }
    specialties_by_name = {
        _csv_lookup_key(s.specialty): s
        for s in ExpertSpecialty.objects.exclude(specialty="").all()
    }
    data_rows = rows[1:]
    created = 0
    updated = 0
    warnings = []

    for i, row in enumerate(data_rows, start=2):
        if not any(cell.strip() for cell in row):
            continue
        if len(row) < 7:
            warnings.append(f"Строка {i}: недостаточно столбцов ({len(row)}, ожидается 7-9: Продукт, Код, Краткое имя EN, Краткое имя RU, Наименование EN, Наименование RU, Тип учета, [Направление экспертизы]).")
            continue

        product_name = row[0].strip()
        code = row[1].strip()
        short_name = row[2].strip()
        short_name_ru = row[3].strip() if len(row) > 3 else ""
        name_en = row[4].strip() if len(row) > 4 else ""
        name_ru = row[5].strip() if len(row) > 5 else ""
        accounting_type = row[6].strip() if len(row) > 6 else "Раздел"
        executor_raw = ""
        expertise_dir_name = ""
        tkp_raw = ""
        if len(row) >= len(SECTION_CSV_HEADERS):
            executor_raw = row[7].strip()
            expertise_dir_name = row[8].strip()
            expertise_name = row[9].strip()
            tkp_raw = row[10].strip()
        elif len(row) > 8:
            executor_raw = row[7].strip()
            expertise_name = row[8].strip()
        else:
            expertise_name = row[7].strip() if len(row) > 7 else ""

        product = products_by_name.get(_csv_lookup_key(product_name))
        if not product:
            warnings.append(f"Строка {i}: продукт «{product_name}» не найден. Доступные: {', '.join(products_by_name.keys())}.")
            continue

        if not code:
            warnings.append(f"Строка {i}: отсутствует код раздела.")
            continue

        if is_system_dsc_code(code):
            ensure_system_dsc_section(product)
            warnings.append(
                f"Строка {i}: раздел DSC является системным; строка CSV пропущена, системная запись создана/обновлена автоматически."
            )
            continue

        missing = []
        if not short_name:
            missing.append("Краткое имя EN")
        if not name_en:
            missing.append("Наименование EN")
        if not name_ru:
            missing.append("Наименование RU")
        if missing:
            warnings.append(f"Строка {i}: не заполнены обязательные поля: {', '.join(missing)}.")
            continue

        if accounting_type not in dict(TypicalSection.ACCOUNTING_TYPE_CHOICES):
            warnings.append(f"Строка {i}: неизвестный тип учета «{accounting_type}». Допустимые: {', '.join(dict(TypicalSection.ACCOUNTING_TYPE_CHOICES).keys())}. Установлено «Раздел».")
            accounting_type = "Раздел"

        expertise_dir = None
        if expertise_dir_name:
            expertise_dir = expertise_dirs.get(_csv_lookup_key(expertise_dir_name))
            if not expertise_dir:
                warnings.append(f"Строка {i}: экспертиза «{expertise_dir_name}» не найдена. Поле оставлено пустым.")

        expertise_direction = None
        if expertise_name:
            expertise_direction = expertise_units.get(_csv_lookup_key(expertise_name))
            if not expertise_direction:
                warnings.append(f"Строка {i}: подразделение «{expertise_name}» не найдено. Поле оставлено пустым.")

        try:
            with transaction.atomic():
                ensure_system_dsc_section(product)
                section = TypicalSection.objects.filter(product=product, code=code).first()
                was_created = section is None
                if was_created:
                    section = TypicalSection.objects.create(
                        product=product,
                        code=code,
                        short_name=short_name,
                        short_name_ru=short_name_ru,
                        name_en=name_en,
                        name_ru=name_ru,
                        accounting_type=accounting_type,
                        expertise_dir=expertise_dir,
                        expertise_direction=expertise_direction,
                        exclude_from_tkp_autofill=_csv_truthy(tkp_raw),
                        position=_next_position(TypicalSection, {"product": product}),
                    )
                else:
                    section.short_name = short_name
                    section.short_name_ru = short_name_ru
                    section.name_en = name_en
                    section.name_ru = name_ru
                    section.accounting_type = accounting_type
                    section.expertise_dir = expertise_dir
                    section.expertise_direction = expertise_direction
                    section.exclude_from_tkp_autofill = _csv_truthy(tkp_raw)
                    section.save(
                        update_fields=[
                            "short_name",
                            "short_name_ru",
                            "name_en",
                            "name_ru",
                            "accounting_type",
                            "expertise_dir",
                            "expertise_direction",
                            "exclude_from_tkp_autofill",
                            "updated_at",
                        ]
                    )
                    TypicalSectionSpecialty.objects.filter(section=section).delete()
                missing_specialties = []
                for rank, specialty_name in enumerate(_split_csv_list_value(executor_raw, specialties_by_name), start=1):
                    specialty = specialties_by_name.get(_csv_lookup_key(specialty_name))
                    if not specialty:
                        missing_specialties.append(specialty_name)
                        continue
                    TypicalSectionSpecialty.objects.create(section=section, specialty=specialty, rank=rank)
            if missing_specialties:
                warnings.append(f"Строка {i}: исполнители не найдены: {', '.join(missing_specialties)}.")
            if was_created:
                created += 1
            else:
                updated += 1
            ensure_system_dsc_section(product)
        except Exception as exc:
            warnings.append(f"Строка {i}: ошибка сохранения — {exc}")

    return JsonResponse({"ok": True, "created": created, "updated": updated, "warnings": warnings})


@login_required
@user_passes_test(staff_required)
@require_http_methods(["GET"])
def section_csv_download(request):
    output = io.StringIO()
    output.write("\ufeff")
    writer = csv.writer(output, delimiter=";", lineterminator="\n")
    writer.writerow(SECTION_CSV_HEADERS)

    sections = _apply_policy_master_product_filters(
        TypicalSection.objects.select_related(
            "product",
            "product__consulting_type_ref",
            "product__service_category_ref",
            "product__service_subtype_ref",
            "expertise_dir",
            "expertise_direction",
        ).prefetch_related("ranked_specialties", "ranked_specialties__specialty"),
        request,
    )
    for section in sections:
        executor = "\n".join(
            rs.specialty.specialty
            for rs in section.ranked_specialties.all()
            if rs.specialty and rs.specialty.specialty
        )
        writer.writerow(
            [
                section.product.short_name,
                section.code,
                section.short_name,
                section.short_name_ru,
                section.name_en,
                section.name_ru,
                section.accounting_type,
                executor,
                section.expertise_dir.short_name if section.expertise_dir else "",
                section.expertise_direction.department_name if section.expertise_direction else "",
                "Да" if section.exclude_from_tkp_autofill else "Нет",
            ]
        )

    response = HttpResponse(output.getvalue(), content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="typical_sections.csv"'
    return response


# --- Типовая структура раздела ---

@login_required
@user_passes_test(staff_required)
@require_http_methods(["GET", "POST"])
def structure_form_create(request):
    if request.method == "GET":
        form = SectionStructureForm(initial=_product_field_initial_from_request(request))
        return render(request, STRUCTURE_FORM_TEMPLATE, _structure_form_context(form, "create"))
    form = SectionStructureForm(request.POST)
    if not form.is_valid():
        return render(request, STRUCTURE_FORM_TEMPLATE, _structure_form_context(form, "create"))
    obj = form.save(commit=False)
    if not getattr(obj, "position", 0):
        obj.position = _next_position(SectionStructure)
    obj.save()
    return _render_policy_updated(request)


@login_required
@user_passes_test(staff_required)
@require_http_methods(["GET", "POST"])
def structure_form_edit(request, pk: int):
    structure = get_object_or_404(SectionStructure, pk=pk)
    if request.method == "GET":
        form = SectionStructureForm(instance=structure)
        return render(request, STRUCTURE_FORM_TEMPLATE, _structure_form_context(form, "edit", structure))
    form = SectionStructureForm(request.POST, instance=structure)
    if not form.is_valid():
        return render(request, STRUCTURE_FORM_TEMPLATE, _structure_form_context(form, "edit", structure))
    form.save()
    return _render_policy_updated(request)


@login_required
@user_passes_test(staff_required)
@require_POST
def structure_delete(request, pk: int):
    structure = get_object_or_404(SectionStructure, pk=pk)
    structure.delete()
    return _render_policy_updated(request)


@login_required
@user_passes_test(staff_required)
@require_POST
def structure_csv_upload(request):
    csv_file = request.FILES.get("csv_file")
    if not csv_file:
        return JsonResponse({"ok": False, "error": "Файл не выбран."}, status=400)
    if not csv_file.name.lower().endswith(".csv"):
        return JsonResponse({"ok": False, "error": "Допустимы только файлы CSV."}, status=400)

    try:
        raw = csv_file.read().decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            csv_file.seek(0)
            raw = csv_file.read().decode("cp1251")
        except Exception:
            return JsonResponse(
                {"ok": False, "error": "Не удалось прочитать файл. Проверьте кодировку (UTF-8 или Windows-1251)."},
                status=400,
            )

    try:
        reader = csv.reader(io.StringIO(raw), delimiter=";")
        rows = list(reader)
        if not rows:
            return JsonResponse({"ok": False, "error": "Файл пуст."}, status=400)
        if len(rows[0]) <= 1:
            reader = csv.reader(io.StringIO(raw), delimiter=",")
            rows = list(reader)
    except csv.Error as exc:
        return JsonResponse({"ok": False, "error": f"Ошибка разбора CSV: {exc}. Проверьте формат и кодировку файла."}, status=400)

    if len(rows) < 2:
        return JsonResponse({"ok": False, "error": "Файл должен содержать заголовок и хотя бы одну строку данных."}, status=400)

    products_by_name = {_csv_lookup_key(p.short_name): p for p in Product.objects.all()}
    sections_by_product = defaultdict(dict)
    for section in TypicalSection.objects.select_related("product").all():
        lookup = sections_by_product[section.product_id]
        for label in (section.name_ru, section.name_en, section.code, section.short_name, section.short_name_ru):
            key = _csv_lookup_key(label)
            if key:
                lookup.setdefault(key, section)

    created = 0
    warnings = []

    for i, row in enumerate(rows[1:], start=2):
        if not any(cell.strip() for cell in row):
            continue
        if len(row) < 3:
            warnings.append(
                f"Строка {i}: недостаточно столбцов ({len(row)}, ожидается 3: "
                "Продукт, Раздел (услуга), Подразделы)."
            )
            continue

        product_name = row[0].strip()
        section_name = row[1].strip()
        product = products_by_name.get(_csv_lookup_key(product_name))
        if not product:
            warnings.append(f"Строка {i}: продукт «{product_name}» не найден. Доступные: {', '.join(products_by_name.keys())}.")
            continue

        section = sections_by_product[product.pk].get(_csv_lookup_key(section_name))
        if not section:
            warnings.append(f"Строка {i}: раздел «{section_name}» не найден для продукта «{product.short_name}».")
            continue

        try:
            SectionStructure.objects.create(
                product=product,
                section=section,
                subsections=row[2].strip(),
                position=_next_position(SectionStructure),
            )
            created += 1
        except Exception as exc:
            warnings.append(f"Строка {i}: ошибка сохранения — {exc}")

    return JsonResponse({"ok": True, "created": created, "warnings": warnings})


@login_required
@user_passes_test(staff_required)
@require_http_methods(["GET"])
def structure_csv_download(request):
    output = io.StringIO()
    output.write("\ufeff")
    writer = csv.writer(output, delimiter=";", lineterminator="\n")
    writer.writerow(STRUCTURE_CSV_HEADERS)

    structures = _apply_policy_master_product_filters(
        SectionStructure.objects.select_related(
            "product",
            "product__consulting_type_ref",
            "product__service_category_ref",
            "product__service_subtype_ref",
            "section",
        ),
        request,
    )
    for structure in structures:
        writer.writerow(
            [
                structure.product.short_name,
                structure.section.name_ru or structure.section.name_en,
                structure.subsections,
            ]
        )

    response = HttpResponse(output.getvalue(), content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="section_structures.csv"'
    return response


def _normalize_structure_positions():
    items = SectionStructure.objects.order_by("position", "id").only("id", "position")
    for idx, it in enumerate(items, start=1):
        if it.position != idx:
            SectionStructure.objects.filter(pk=it.pk).update(position=idx)


@require_http_methods(["POST", "GET"])
@login_required
def structure_move_up(request, pk: int):
    _normalize_structure_positions()
    items = list(SectionStructure.objects.order_by("position", "id").only("id", "position"))
    idx = next((i for i, it in enumerate(items) if it.id == pk), None)
    if idx is not None and idx > 0:
        cur = items[idx]
        prev = items[idx - 1]
        cur_pos, prev_pos = cur.position, prev.position
        SectionStructure.objects.filter(pk=cur.id).update(position=prev_pos)
        SectionStructure.objects.filter(pk=prev.id).update(position=cur_pos)
        _normalize_structure_positions()
    return _render_policy_updated(request)


@require_http_methods(["POST", "GET"])
@login_required
def structure_move_down(request, pk: int):
    _normalize_structure_positions()
    items = list(SectionStructure.objects.order_by("position", "id").only("id", "position"))
    idx = next((i for i, it in enumerate(items) if it.id == pk), None)
    if idx is not None and idx < len(items) - 1:
        cur = items[idx]
        nxt = items[idx + 1]
        cur_pos, next_pos = cur.position, nxt.position
        SectionStructure.objects.filter(pk=cur.id).update(position=next_pos)
        SectionStructure.objects.filter(pk=nxt.id).update(position=cur_pos)
        _normalize_structure_positions()
    return _render_policy_updated(request)


# --- Цели услуг и названия отчетов ---

@login_required
@user_passes_test(staff_required)
@require_http_methods(["GET", "POST"])
def service_goal_report_form_create(request):
    if request.method == "GET":
        form = ServiceGoalReportForm(initial=_product_field_initial_from_request(request))
        return render(request, SERVICE_GOAL_REPORT_FORM_TEMPLATE, {"form": form, "action": "create"})
    form = ServiceGoalReportForm(request.POST)
    if not form.is_valid():
        return render(request, SERVICE_GOAL_REPORT_FORM_TEMPLATE, {"form": form, "action": "create"})
    obj = form.save(commit=False)
    if not getattr(obj, "position", 0):
        obj.position = _next_position(ServiceGoalReport)
    obj.save()
    return _render_policy_updated(request)


@login_required
@user_passes_test(staff_required)
@require_http_methods(["GET", "POST"])
def service_goal_report_form_edit(request, pk: int):
    service_goal_report = get_object_or_404(ServiceGoalReport, pk=pk)
    if request.method == "GET":
        form = ServiceGoalReportForm(instance=service_goal_report)
        return render(
            request,
            SERVICE_GOAL_REPORT_FORM_TEMPLATE,
            {"form": form, "action": "edit", "service_goal_report": service_goal_report},
        )
    form = ServiceGoalReportForm(request.POST, instance=service_goal_report)
    if not form.is_valid():
        return render(
            request,
            SERVICE_GOAL_REPORT_FORM_TEMPLATE,
            {"form": form, "action": "edit", "service_goal_report": service_goal_report},
        )
    form.save()
    return _render_policy_updated(request)


@login_required
@user_passes_test(staff_required)
@require_POST
def service_goal_report_delete(request, pk: int):
    service_goal_report = get_object_or_404(ServiceGoalReport, pk=pk)
    service_goal_report.delete()
    return _render_policy_updated(request)


@login_required
@user_passes_test(staff_required)
@require_POST
def service_goal_report_csv_upload(request):
    csv_file = request.FILES.get("csv_file")
    if not csv_file:
        return JsonResponse({"ok": False, "error": "Файл не выбран."}, status=400)
    if not csv_file.name.lower().endswith(".csv"):
        return JsonResponse({"ok": False, "error": "Допустимы только файлы CSV."}, status=400)

    try:
        raw = csv_file.read().decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            csv_file.seek(0)
            raw = csv_file.read().decode("cp1251")
        except Exception:
            return JsonResponse(
                {"ok": False, "error": "Не удалось прочитать файл. Проверьте кодировку (UTF-8 или Windows-1251)."},
                status=400,
            )

    try:
        reader = csv.reader(io.StringIO(raw), delimiter=";")
        rows = list(reader)
        if not rows:
            return JsonResponse({"ok": False, "error": "Файл пуст."}, status=400)
        if len(rows[0]) <= 1:
            reader = csv.reader(io.StringIO(raw), delimiter=",")
            rows = list(reader)
    except csv.Error as exc:
        return JsonResponse({"ok": False, "error": f"Ошибка разбора CSV: {exc}. Проверьте формат и кодировку файла."}, status=400)

    if len(rows) < 2:
        return JsonResponse({"ok": False, "error": "Файл должен содержать заголовок и хотя бы одну строку данных."}, status=400)

    products_by_name = {_csv_lookup_key(p.short_name): p for p in Product.objects.all()}
    created = 0
    updated = 0
    warnings = []

    for i, row in enumerate(rows[1:], start=2):
        if not any(cell.strip() for cell in row):
            continue
        if len(row) < 5:
            warnings.append(
                f"Строка {i}: недостаточно столбцов ({len(row)}, ожидается 5: "
                "Продукт, Цели оказания услуг, Цели оказания услуг в родительном падеже, "
                "Титул отчета/ТКП, Название продукта)."
            )
            continue

        product_name = row[0].strip()
        product = products_by_name.get(_csv_lookup_key(product_name))
        if not product:
            warnings.append(f"Строка {i}: продукт «{product_name}» не найден. Доступные: {', '.join(products_by_name.keys())}.")
            continue

        try:
            item = ServiceGoalReport.objects.filter(product=product).order_by("position", "id").first()
            was_created = item is None
            if was_created:
                item = ServiceGoalReport(product=product, position=_next_position(ServiceGoalReport))
            item.service_goal = row[1].strip()
            item.service_goal_genitive = row[2].strip()
            item.report_title = row[3].strip()
            item.product_name = row[4].strip()
            item.save()
            if was_created:
                created += 1
            else:
                updated += 1
        except Exception as exc:
            warnings.append(f"Строка {i}: ошибка сохранения — {exc}")

    return JsonResponse({"ok": True, "created": created, "updated": updated, "warnings": warnings})


@login_required
@user_passes_test(staff_required)
@require_http_methods(["GET"])
def service_goal_report_csv_download(request):
    output = io.StringIO()
    output.write("\ufeff")
    writer = csv.writer(output, delimiter=";", lineterminator="\n")
    writer.writerow(SERVICE_GOAL_REPORT_CSV_HEADERS)

    items = _apply_policy_master_product_filters(
        ServiceGoalReport.objects.select_related(
            "product",
            "product__consulting_type_ref",
            "product__service_category_ref",
            "product__service_subtype_ref",
        ),
        request,
    )
    for item in items:
        writer.writerow(
            [
                item.product.short_name,
                item.service_goal,
                item.service_goal_genitive,
                item.report_title,
                item.product_name,
            ]
        )

    response = HttpResponse(output.getvalue(), content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="service_goal_reports.csv"'
    return response


def _normalize_service_goal_report_positions():
    items = ServiceGoalReport.objects.order_by("position", "id").only("id", "position")
    for idx, item in enumerate(items, start=1):
        if item.position != idx:
            ServiceGoalReport.objects.filter(pk=item.pk).update(position=idx)


@require_http_methods(["POST", "GET"])
@login_required
@user_passes_test(staff_required)
def service_goal_report_move_up(request, pk: int):
    _normalize_service_goal_report_positions()
    items = list(
        ServiceGoalReport.objects
        .order_by("position", "id")
        .only("id", "position")
    )
    idx = next((i for i, it in enumerate(items) if it.id == pk), None)
    if idx is not None and idx > 0:
        cur = items[idx]
        prev = items[idx - 1]
        cur_pos, prev_pos = cur.position, prev.position
        ServiceGoalReport.objects.filter(pk=cur.id).update(position=prev_pos)
        ServiceGoalReport.objects.filter(pk=prev.id).update(position=cur_pos)
        _normalize_service_goal_report_positions()
    return _render_policy_updated(request)


@require_http_methods(["POST", "GET"])
@login_required
@user_passes_test(staff_required)
def service_goal_report_move_down(request, pk: int):
    _normalize_service_goal_report_positions()
    items = list(
        ServiceGoalReport.objects
        .order_by("position", "id")
        .only("id", "position")
    )
    idx = next((i for i, it in enumerate(items) if it.id == pk), None)
    if idx is not None and idx < len(items) - 1:
        cur = items[idx]
        nxt = items[idx + 1]
        cur_pos, next_pos = cur.position, nxt.position
        ServiceGoalReport.objects.filter(pk=cur.id).update(position=next_pos)
        ServiceGoalReport.objects.filter(pk=nxt.id).update(position=cur_pos)
        _normalize_service_goal_report_positions()
    return _render_policy_updated(request)


# --- Типовой состав услуг ---

@login_required
@user_passes_test(staff_required)
@require_http_methods(["GET", "POST"])
def typical_service_composition_form_create(request):
    if request.method == "GET":
        form = TypicalServiceCompositionForm(initial=_product_field_initial_from_request(request))
        return render(
            request,
            TYPICAL_SERVICE_COMPOSITION_FORM_TEMPLATE,
            _typical_service_composition_form_context(form, "create"),
        )
    form = TypicalServiceCompositionForm(request.POST)
    if not form.is_valid():
        return render(
            request,
            TYPICAL_SERVICE_COMPOSITION_FORM_TEMPLATE,
            _typical_service_composition_form_context(form, "create"),
        )
    obj = form.save(commit=False)
    if not getattr(obj, "position", 0):
        obj.position = _next_position(TypicalServiceComposition)
    obj.save()
    return _render_policy_updated(request)


@login_required
@user_passes_test(staff_required)
@require_http_methods(["GET", "POST"])
def typical_service_composition_form_edit(request, pk: int):
    composition = get_object_or_404(TypicalServiceComposition, pk=pk)
    if request.method == "GET":
        form = TypicalServiceCompositionForm(instance=composition)
        return render(
            request,
            TYPICAL_SERVICE_COMPOSITION_FORM_TEMPLATE,
            _typical_service_composition_form_context(form, "edit", composition),
        )
    form = TypicalServiceCompositionForm(request.POST, instance=composition)
    if not form.is_valid():
        return render(
            request,
            TYPICAL_SERVICE_COMPOSITION_FORM_TEMPLATE,
            _typical_service_composition_form_context(form, "edit", composition),
        )
    form.save()
    return _render_policy_updated(request)


@login_required
@user_passes_test(staff_required)
@require_POST
def typical_service_composition_delete(request, pk: int):
    composition = get_object_or_404(TypicalServiceComposition, pk=pk)
    composition.delete()
    return _render_policy_updated(request)


@login_required
@user_passes_test(staff_required)
@require_POST
def typical_service_composition_csv_upload(request):
    csv_file = request.FILES.get("csv_file")
    if not csv_file:
        return JsonResponse({"ok": False, "error": "Файл не выбран."}, status=400)
    if not csv_file.name.lower().endswith(".csv"):
        return JsonResponse({"ok": False, "error": "Допустимы только файлы CSV."}, status=400)

    try:
        raw = csv_file.read().decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            csv_file.seek(0)
            raw = csv_file.read().decode("cp1251")
        except Exception:
            return JsonResponse(
                {"ok": False, "error": "Не удалось прочитать файл. Проверьте кодировку (UTF-8 или Windows-1251)."},
                status=400,
            )

    try:
        reader = csv.reader(io.StringIO(raw), delimiter=";")
        rows = list(reader)
        if not rows:
            return JsonResponse({"ok": False, "error": "Файл пуст."}, status=400)
        if len(rows[0]) <= 1:
            reader = csv.reader(io.StringIO(raw), delimiter=",")
            rows = list(reader)
    except csv.Error as exc:
        return JsonResponse({"ok": False, "error": f"Ошибка разбора CSV: {exc}. Проверьте формат и кодировку файла."}, status=400)

    if len(rows) < 2:
        return JsonResponse({"ok": False, "error": "Файл должен содержать заголовок и хотя бы одну строку данных."}, status=400)

    products_by_name, sections_by_product = _typical_service_composition_import_lookups()

    created = 0
    warnings = []

    for i, row in enumerate(rows[1:], start=2):
        if not any(cell.strip() for cell in row):
            continue
        if len(row) < 3:
            warnings.append(
                f"Строка {i}: недостаточно столбцов ({len(row)}, ожидается 3: "
                "Продукт, Раздел (услуга), Состав услуг)."
            )
            continue

        product_name = row[0].strip()
        section_name = row[1].strip()
        product = products_by_name.get(_csv_lookup_key(product_name))
        if not product:
            warnings.append(f"Строка {i}: продукт «{product_name}» не найден. Доступные: {', '.join(products_by_name.keys())}.")
            continue

        section = sections_by_product[product.pk].get(_csv_lookup_key(section_name))
        if not section:
            warnings.append(f"Строка {i}: раздел «{section_name}» не найден для продукта «{product.short_name}».")
            continue

        service_composition = row[2].strip()
        try:
            TypicalServiceComposition.objects.create(
                product=product,
                section=section,
                service_composition=service_composition,
                service_composition_editor_state={
                    "html": "",
                    "plain_text": service_composition,
                },
                position=_next_position(TypicalServiceComposition),
            )
            created += 1
        except Exception as exc:
            warnings.append(f"Строка {i}: ошибка сохранения — {exc}")

    return JsonResponse({"ok": True, "created": created, "warnings": warnings})


def _typical_service_composition_import_lookups():
    products_by_name = {_csv_lookup_key(p.short_name): p for p in Product.objects.all()}
    sections_by_product = defaultdict(dict)
    for section in TypicalSection.objects.select_related("product").all():
        lookup = sections_by_product[section.product_id]
        for label in (section.name_ru, section.name_en, section.code, section.short_name, section.short_name_ru):
            key = _csv_lookup_key(label)
            if key:
                lookup.setdefault(key, section)
    return products_by_name, sections_by_product


def _normalize_typical_service_composition_editor_state(value, plain_text=""):
    if isinstance(value, str):
        raw = value.strip()
        if raw:
            try:
                parsed = json.loads(raw)
            except (TypeError, ValueError, json.JSONDecodeError):
                parsed = {}
        else:
            parsed = {}
    elif isinstance(value, dict):
        parsed = value
    else:
        parsed = {}

    fallback_plain_text = str(plain_text or "").strip()
    return {
        "html": str(parsed.get("html") or "").strip(),
        "plain_text": str(parsed.get("plain_text") or fallback_plain_text).strip(),
    }


def _xlsx_cell_value(cell):
    return str(cell.value or "").strip()


def _find_header_column(headers, header_name):
    lookup_name = _csv_lookup_key(header_name)
    for idx, value in enumerate(headers, start=1):
        if _csv_lookup_key(value) == lookup_name:
            return idx
    return None


def _product_display_field_q(char_field, ref_field, values, *, prefix="product__"):
    if not values:
        return Q()
    return Q(**{f"{prefix}{ref_field}__name__in": values}) | Q(
        **{f"{prefix}{ref_field}__isnull": True, f"{prefix}{char_field}__in": values}
    )


def _apply_policy_master_product_filters(qs, request):
    product_ids = [_positive_int(value) for value in request.GET.getlist("product")]
    product_ids = [value for value in product_ids if value]
    if product_ids:
        qs = qs.filter(product_id__in=product_ids)

    consulting = [value.strip() for value in request.GET.getlist("consulting") if value and value.strip()]
    category = [value.strip() for value in request.GET.getlist("category") if value and value.strip()]
    subtype = [value.strip() for value in request.GET.getlist("subtype") if value and value.strip()]

    if consulting:
        qs = qs.filter(_product_display_field_q("consulting_type", "consulting_type_ref", consulting))
    if category:
        qs = qs.filter(_product_display_field_q("service_category", "service_category_ref", category))
    if subtype:
        qs = qs.filter(_product_display_field_q("service_subtype", "service_subtype_ref", subtype))

    return qs


def _apply_policy_master_filters_to_products(qs, request):
    product_ids = [_positive_int(value) for value in request.GET.getlist("product")]
    product_ids = [value for value in product_ids if value]
    if product_ids:
        qs = qs.filter(pk__in=product_ids)

    consulting = [value.strip() for value in request.GET.getlist("consulting") if value and value.strip()]
    category = [value.strip() for value in request.GET.getlist("category") if value and value.strip()]
    subtype = [value.strip() for value in request.GET.getlist("subtype") if value and value.strip()]

    if consulting:
        qs = qs.filter(_product_display_field_q("consulting_type", "consulting_type_ref", consulting, prefix=""))
    if category:
        qs = qs.filter(_product_display_field_q("service_category", "service_category_ref", category, prefix=""))
    if subtype:
        qs = qs.filter(_product_display_field_q("service_subtype", "service_subtype_ref", subtype, prefix=""))

    return qs


def _filter_typical_service_compositions_queryset(request):
    qs = TypicalServiceComposition.objects.select_related(
        "product",
        "product__consulting_type_ref",
        "product__service_category_ref",
        "product__service_subtype_ref",
        "section",
    )
    return _apply_policy_master_product_filters(qs, request)


@login_required
@user_passes_test(staff_required)
@require_http_methods(["GET"])
def typical_service_composition_csv_download(request):
    output = io.StringIO()
    output.write("\ufeff")
    writer = csv.writer(output, delimiter=";", lineterminator="\n")
    writer.writerow(TYPICAL_SERVICE_COMPOSITION_CSV_HEADERS)

    compositions = _filter_typical_service_compositions_queryset(request)
    for composition in compositions:
        writer.writerow(
            [
                composition.product.short_name,
                composition.section.name_ru or composition.section.name_en,
                composition.service_composition,
            ]
        )

    response = HttpResponse(output.getvalue(), content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="typical_service_compositions.csv"'
    return response


@login_required
@user_passes_test(staff_required)
@require_http_methods(["GET"])
def typical_service_composition_docx_download(request):
    rows = []
    compositions = _filter_typical_service_compositions_queryset(request)
    for composition in compositions:
        editor_state = _normalize_typical_service_composition_editor_state(
            composition.service_composition_editor_state,
            composition.service_composition,
        )
        rows.append(
            {
                "id": composition.pk,
                "product": composition.product.short_name,
                "section": composition.section.name_ru or composition.section.name_en,
                "html": editor_state["html"],
                "plain_text": editor_state["plain_text"] or composition.service_composition,
            }
        )

    content = build_typical_service_compositions_docx(rows)
    response = HttpResponse(
        content,
        content_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    )
    response["Content-Disposition"] = 'attachment; filename="typical_service_compositions.docx"'
    return response


@login_required
@user_passes_test(staff_required)
@require_POST
def typical_service_composition_docx_upload(request):
    docx_file = request.FILES.get("docx_file") or request.FILES.get("csv_file")
    if not docx_file:
        return JsonResponse({"ok": False, "error": "Файл не выбран."}, status=400)
    if not docx_file.name.lower().endswith(".docx"):
        return JsonResponse({"ok": False, "error": "Допустимы только файлы DOCX."}, status=400)

    try:
        rows = parse_typical_service_compositions_docx(docx_file)
    except (BadZipFile, PackageNotFoundError, OSError, ValueError) as exc:
        return JsonResponse({"ok": False, "error": f"Не удалось прочитать DOCX: {exc}"}, status=400)

    products_by_name, sections_by_product = _typical_service_composition_import_lookups()
    created = 0
    updated = 0
    warnings = []

    for row in rows:
        row_number = row.get("row_number")
        row_label = f"Строка {row_number}"
        raw_id = str(row.get("id") or "").strip()
        product_name = str(row.get("product") or "").strip()
        section_name = str(row.get("section") or "").strip()
        editor_state = _normalize_typical_service_composition_editor_state(
            row.get("editor_state"),
            "",
        )
        service_composition = editor_state["plain_text"]

        product = products_by_name.get(_csv_lookup_key(product_name))
        if not product:
            warnings.append(f"{row_label}: продукт «{product_name}» не найден. Доступные: {', '.join(products_by_name.keys())}.")
            continue

        section = sections_by_product[product.pk].get(_csv_lookup_key(section_name))
        if not section:
            warnings.append(f"{row_label}: раздел «{section_name}» не найден для продукта «{product.short_name}».")
            continue

        try:
            with transaction.atomic():
                if raw_id:
                    if not raw_id.isdigit():
                        warnings.append(f"{row_label}: некорректный ID «{raw_id}»; строка пропущена.")
                        continue
                    composition = TypicalServiceComposition.objects.filter(pk=int(raw_id)).first()
                    if not composition:
                        warnings.append(f"{row_label}: строка с ID «{raw_id}» не найдена; строка пропущена.")
                        continue
                    composition.product = product
                    composition.section = section
                    composition.service_composition = service_composition
                    composition.service_composition_editor_state = editor_state
                    composition.save(
                        update_fields=[
                            "product",
                            "section",
                            "service_composition",
                            "service_composition_editor_state",
                            "updated_at",
                        ]
                    )
                    updated += 1
                else:
                    TypicalServiceComposition.objects.create(
                        product=product,
                        section=section,
                        service_composition=service_composition,
                        service_composition_editor_state=editor_state,
                        position=_next_position(TypicalServiceComposition),
                    )
                    created += 1
        except Exception as exc:
            warnings.append(f"{row_label}: ошибка сохранения — {exc}")

    return JsonResponse({"ok": True, "created": created, "updated": updated, "warnings": warnings})


@login_required
@user_passes_test(staff_required)
@require_http_methods(["GET"])
def typical_service_composition_xlsx_download(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Типовой состав услуг"

    hidden_state_col = len(TYPICAL_SERVICE_COMPOSITION_XLSX_HEADERS) + 1
    ws.append(TYPICAL_SERVICE_COMPOSITION_XLSX_HEADERS + [TYPICAL_SERVICE_COMPOSITION_EDITOR_STATE_HEADER])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    compositions = _filter_typical_service_compositions_queryset(request)
    for composition in compositions:
        editor_state = _normalize_typical_service_composition_editor_state(
            composition.service_composition_editor_state,
            composition.service_composition,
        )
        ws.append(
            [
                composition.product.short_name,
                composition.section.name_ru or composition.section.name_en,
                editor_state["plain_text"] or composition.service_composition,
                json.dumps(editor_state, ensure_ascii=False),
            ]
        )

    widths = {
        1: 18,
        2: 28,
        3: 80,
        hidden_state_col: 80,
    }
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.column_dimensions[get_column_letter(hidden_state_col)].hidden = True
    ws.freeze_panes = "A2"

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="typical_service_compositions.xlsx"'
    return response


@login_required
@user_passes_test(staff_required)
@require_POST
def typical_service_composition_xlsx_upload(request):
    xlsx_file = request.FILES.get("xlsx_file") or request.FILES.get("csv_file")
    if not xlsx_file:
        return JsonResponse({"ok": False, "error": "Файл не выбран."}, status=400)
    if not xlsx_file.name.lower().endswith(".xlsx"):
        return JsonResponse({"ok": False, "error": "Допустимы только файлы XLSX."}, status=400)

    try:
        wb = load_workbook(xlsx_file, data_only=False)
    except (BadZipFile, InvalidFileException, OSError, ValueError) as exc:
        return JsonResponse({"ok": False, "error": f"Не удалось прочитать XLSX: {exc}"}, status=400)

    ws = wb.active
    if ws.max_row < 2:
        return JsonResponse({"ok": False, "error": "Файл должен содержать заголовок и хотя бы одну строку данных."}, status=400)

    headers = [_xlsx_cell_value(cell) for cell in ws[1]]
    editor_state_col = _find_header_column(headers, TYPICAL_SERVICE_COMPOSITION_EDITOR_STATE_HEADER)

    products_by_name, sections_by_product = _typical_service_composition_import_lookups()
    created = 0
    warnings = []

    for row_idx in range(2, ws.max_row + 1):
        product_name = _xlsx_cell_value(ws.cell(row=row_idx, column=1))
        section_name = _xlsx_cell_value(ws.cell(row=row_idx, column=2))
        service_composition = _xlsx_cell_value(ws.cell(row=row_idx, column=3))
        editor_state_raw = (
            ws.cell(row=row_idx, column=editor_state_col).value
            if editor_state_col
            else None
        )

        if not any([product_name, section_name, service_composition, str(editor_state_raw or "").strip()]):
            continue

        product = products_by_name.get(_csv_lookup_key(product_name))
        if not product:
            warnings.append(f"Строка {row_idx}: продукт «{product_name}» не найден. Доступные: {', '.join(products_by_name.keys())}.")
            continue

        section = sections_by_product[product.pk].get(_csv_lookup_key(section_name))
        if not section:
            warnings.append(f"Строка {row_idx}: раздел «{section_name}» не найден для продукта «{product.short_name}».")
            continue

        try:
            editor_state = _normalize_typical_service_composition_editor_state(
                editor_state_raw,
                service_composition,
            )
            if (
                editor_state_raw
                and service_composition
                and editor_state["plain_text"]
                and editor_state["plain_text"] != service_composition
            ):
                editor_state = {"html": "", "plain_text": service_composition}
                warnings.append(
                    f"Строка {row_idx}: видимый текст отличается от скрытого состояния редактора; "
                    "форматирование для этой строки сброшено."
                )
            service_composition_plain = editor_state["plain_text"] or service_composition
            TypicalServiceComposition.objects.create(
                product=product,
                section=section,
                service_composition=service_composition_plain,
                service_composition_editor_state=editor_state,
                position=_next_position(TypicalServiceComposition),
            )
            created += 1
        except Exception as exc:
            warnings.append(f"Строка {row_idx}: ошибка сохранения — {exc}")

    return JsonResponse({"ok": True, "created": created, "warnings": warnings})


def _normalize_typical_service_composition_positions():
    items = TypicalServiceComposition.objects.order_by("position", "id").only("id", "position")
    for idx, item in enumerate(items, start=1):
        if item.position != idx:
            TypicalServiceComposition.objects.filter(pk=item.pk).update(position=idx)


@require_http_methods(["POST", "GET"])
@login_required
@user_passes_test(staff_required)
def typical_service_composition_move_up(request, pk: int):
    _normalize_typical_service_composition_positions()
    items = list(
        TypicalServiceComposition.objects.order_by("position", "id").only("id", "position")
    )
    idx = next((i for i, it in enumerate(items) if it.id == pk), None)
    if idx is not None and idx > 0:
        cur = items[idx]
        prev = items[idx - 1]
        cur_pos, prev_pos = cur.position, prev.position
        TypicalServiceComposition.objects.filter(pk=cur.id).update(position=prev_pos)
        TypicalServiceComposition.objects.filter(pk=prev.id).update(position=cur_pos)
        _normalize_typical_service_composition_positions()
    return _render_policy_updated(request)


@require_http_methods(["POST", "GET"])
@login_required
@user_passes_test(staff_required)
def typical_service_composition_move_down(request, pk: int):
    _normalize_typical_service_composition_positions()
    items = list(
        TypicalServiceComposition.objects.order_by("position", "id").only("id", "position")
    )
    idx = next((i for i, it in enumerate(items) if it.id == pk), None)
    if idx is not None and idx < len(items) - 1:
        cur = items[idx]
        nxt = items[idx + 1]
        cur_pos, next_pos = cur.position, nxt.position
        TypicalServiceComposition.objects.filter(pk=cur.id).update(position=next_pos)
        TypicalServiceComposition.objects.filter(pk=nxt.id).update(position=cur_pos)
        _normalize_typical_service_composition_positions()
    return _render_policy_updated(request)


# --- Типовые сроки оказания услуг ---

def _typical_service_term_gantt_base_date():
    return date.today().replace(month=1, day=1)


def _add_typical_service_term_gantt_months(start_date, months):
    safe_months = max(Decimal(months or 0), Decimal("0"))
    whole_months = int(safe_months)
    fractional_months = safe_months - Decimal(whole_months)
    month_index = start_date.month - 1 + whole_months
    year = start_date.year + month_index // 12
    month = month_index % 12 + 1
    day = min(start_date.day, calendar.monthrange(year, month)[1])
    whole_date = date(year, month, day)
    extra_days = int((fractional_months * Decimal("30")).quantize(Decimal("1"), rounding=ROUND_HALF_UP))
    return whole_date + timedelta(days=extra_days)


def _serialize_typical_service_term_gantt_date(value):
    return value.isoformat()


def _default_typical_service_term_gantt_data(term):
    base_date = _typical_service_term_gantt_base_date()
    source_data_end = base_date + timedelta(days=int(getattr(term, "source_data_weeks", 0) or 0) * 7)
    preliminary_end = _add_typical_service_term_gantt_months(source_data_end, term.preliminary_report_months)
    final_end = preliminary_end + timedelta(days=int(term.final_report_weeks or 0) * 7)
    prefix = f"typical-service-term-{term.pk}"
    return {
        "data": [
            {
                "id": f"{prefix}-source-data",
                "text": "Исходные данные",
                "start_date": _serialize_typical_service_term_gantt_date(base_date),
                "end_date": _serialize_typical_service_term_gantt_date(source_data_end),
                "progress": 0,
                "system_key": "source_data",
                "type": "project",
                "is_report_bar": True,
                "$open": True,
            },
            {
                "id": f"{prefix}-source-data-asset",
                "text": "Актив",
                "start_date": _serialize_typical_service_term_gantt_date(base_date),
                "end_date": _serialize_typical_service_term_gantt_date(source_data_end),
                "progress": 0,
                "system_key": "source_data_asset",
                "type": "task",
                "parent": f"{prefix}-source-data",
                "is_report_bar": True,
            },
            {
                "id": f"{prefix}-preliminary-report",
                "text": "Предварительный отчёт",
                "start_date": _serialize_typical_service_term_gantt_date(source_data_end),
                "end_date": _serialize_typical_service_term_gantt_date(preliminary_end),
                "progress": 0,
                "system_key": "preliminary_report",
                "type": "project",
                "is_report_bar": True,
                "$open": True,
            },
            {
                "id": f"{prefix}-preliminary-report-asset",
                "text": "Актив",
                "start_date": _serialize_typical_service_term_gantt_date(source_data_end),
                "end_date": _serialize_typical_service_term_gantt_date(preliminary_end),
                "progress": 0,
                "system_key": "preliminary_report_asset",
                "type": "task",
                "parent": f"{prefix}-preliminary-report",
                "is_report_bar": True,
            },
            {
                "id": f"{prefix}-preliminary-report-submission",
                "text": "Отправка Предварительного отчёта",
                "start_date": _serialize_typical_service_term_gantt_date(preliminary_end),
                "end_date": _serialize_typical_service_term_gantt_date(preliminary_end),
                "progress": 0,
                "system_key": "preliminary_report_submission",
                "type": "milestone",
                "is_report_bar": True,
            },
            {
                "id": f"{prefix}-final-report",
                "text": "Итоговый отчёт",
                "start_date": _serialize_typical_service_term_gantt_date(preliminary_end),
                "end_date": _serialize_typical_service_term_gantt_date(final_end),
                "progress": 0,
                "system_key": "final_report",
                "type": "task",
                "is_report_bar": True,
            },
        ],
        "links": [
            {
                "id": f"{prefix}-source-data-to-preliminary",
                "source": f"{prefix}-source-data",
                "target": f"{prefix}-preliminary-report",
                "type": "0",
            },
            {
                "id": f"{prefix}-preliminary-to-submission",
                "source": f"{prefix}-preliminary-report",
                "target": f"{prefix}-preliminary-report-submission",
                "type": "0",
            },
            {
                "id": f"{prefix}-submission-to-final",
                "source": f"{prefix}-preliminary-report-submission",
                "target": f"{prefix}-final-report",
                "type": "0",
            }
        ],
        "meta": {
            "base_date": _serialize_typical_service_term_gantt_date(base_date),
            "project_start": _serialize_typical_service_term_gantt_date(base_date),
            "project_end": _serialize_typical_service_term_gantt_date(final_end),
            "calendar_kind": TYPICAL_SERVICE_TERM_GANTT_CALENDAR_KIND_ABSTRACT,
            "executor_display": TYPICAL_SERVICE_TERM_GANTT_EXECUTOR_DISPLAY_RESOURCE,
            "version": TYPICAL_SERVICE_TERM_GANTT_VERSION,
        },
    }


def _parse_typical_service_term_gantt_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = str(value or "").strip()
    if not raw:
        return None
    if "T" in raw:
        raw = raw.replace("Z", "+00:00")
        try:
            return datetime.fromisoformat(raw).date()
        except ValueError:
            return None
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None


def _task_end_date(task, start_date):
    end_date = _parse_typical_service_term_gantt_date(task.get("end_date"))
    if end_date:
        return end_date
    try:
        duration = int(Decimal(str(task.get("duration", 0))))
    except (InvalidOperation, ValueError, TypeError):
        duration = 0
    return start_date + timedelta(days=max(duration, 0))


def _unique_typical_service_term_gantt_task_id(tasks, preferred_id):
    existing = {str(task.get("id")) for task in tasks if task.get("id") is not None}
    candidate = str(preferred_id or "task").strip() or "task"
    if candidate not in existing:
        return candidate
    suffix = 1
    while f"{candidate}-{suffix}" in existing:
        suffix += 1
    return f"{candidate}-{suffix}"


def _sync_typical_service_term_gantt_asset_task(tasks, parent_system_key="preliminary_report", asset_system_key="preliminary_report_asset"):
    tasks_by_id = {
        str(task.get("id")): task
        for task in tasks
        if task.get("id") is not None
    }
    parent = next(
        (
            task for task in tasks
            if str(task.get("system_key") or "").strip() == parent_system_key
        ),
        None,
    )
    if parent is None or parent.get("id") is None:
        return None

    parent_id = str(parent.get("id"))
    asset = next(
        (
            task for task in tasks
            if str(task.get("system_key") or "").strip() == asset_system_key
        ),
        None,
    )
    if asset is None:
        asset = next(
            (
                task for task in tasks
                if str(task.get("parent") or "").strip() == parent_id
                and str(task.get("text") or "").strip() == "Актив"
            ),
            None,
        )
        if asset is None:
            asset = {
                "id": _unique_typical_service_term_gantt_task_id(
                    tasks,
                    f"{parent_id}-asset",
                ),
                "progress": 0,
                "type": "task",
                "is_report_bar": True,
            }
            try:
                insert_at = tasks.index(parent) + 1
            except ValueError:
                insert_at = len(tasks)
            tasks.insert(insert_at, asset)

    start_date = _parse_typical_service_term_gantt_date(parent.get("start_date"))
    if start_date:
        end_date = _task_end_date(parent, start_date)
        asset["start_date"] = _serialize_typical_service_term_gantt_date(start_date)
        asset["end_date"] = _serialize_typical_service_term_gantt_date(end_date)
        asset["duration"] = max((end_date - start_date).days, 0)
    asset["system_key"] = asset_system_key
    asset["text"] = TYPICAL_SERVICE_TERM_GANTT_SYSTEM_TASK_TEXT[asset_system_key]
    asset["parent"] = parent_id
    asset["type"] = "task"
    asset["is_report_bar"] = True
    parent["$open"] = True
    if str(parent.get("type") or "").strip() != TYPICAL_SERVICE_TERM_GANTT_SERVICE_SECTION_TYPE:
        parent["type"] = "project"
    # Drop impossible self-parenting or stale parent references if the row was
    # repurposed from an existing user task.
    if str(asset.get("id")) == parent_id:
        asset["id"] = _unique_typical_service_term_gantt_task_id(
            [task for task in tasks if task is not asset],
            f"{parent_id}-asset",
        )
    return asset


def _sync_typical_service_term_gantt_system_tasks(tasks):
    preliminary = next(
        (
            task for task in tasks
            if str(task.get("system_key") or "").strip() == "preliminary_report"
            and task.get("id") is not None
        ),
        None,
    )
    source_data = next(
        (
            task for task in tasks
            if str(task.get("system_key") or "").strip() == "source_data"
            and task.get("id") is not None
        ),
        None,
    )
    if source_data is None and preliminary is not None:
        preliminary_start = _parse_typical_service_term_gantt_date(preliminary.get("start_date"))
        preferred_id = f"{preliminary.get('id')}-source-data"
        source_data = {
            "id": _unique_typical_service_term_gantt_task_id(tasks, preferred_id),
            "text": TYPICAL_SERVICE_TERM_GANTT_SYSTEM_TASK_TEXT["source_data"],
            "start_date": _serialize_typical_service_term_gantt_date(preliminary_start) if preliminary_start else preliminary.get("start_date"),
            "end_date": _serialize_typical_service_term_gantt_date(preliminary_start) if preliminary_start else preliminary.get("start_date"),
            "duration": 0,
            "progress": 0,
            "system_key": "source_data",
            "type": "project",
            "is_report_bar": True,
            "$open": True,
        }
        try:
            insert_at = tasks.index(preliminary)
        except ValueError:
            insert_at = 0
        tasks.insert(insert_at, source_data)
    submission = next(
        (
            task for task in tasks
            if str(task.get("system_key") or "").strip() == "preliminary_report_submission"
            and task.get("id") is not None
        ),
        None,
    )
    if submission is None and preliminary is not None:
        preliminary_start = _parse_typical_service_term_gantt_date(preliminary.get("start_date"))
        preliminary_end = _task_end_date(preliminary, preliminary_start) if preliminary_start else None
        preferred_id = f"{preliminary.get('id')}-submission"
        submission = {
            "id": _unique_typical_service_term_gantt_task_id(tasks, preferred_id),
            "text": TYPICAL_SERVICE_TERM_GANTT_SYSTEM_TASK_TEXT["preliminary_report_submission"],
            "start_date": _serialize_typical_service_term_gantt_date(preliminary_end) if preliminary_end else preliminary.get("end_date") or preliminary.get("start_date"),
            "end_date": _serialize_typical_service_term_gantt_date(preliminary_end) if preliminary_end else preliminary.get("end_date") or preliminary.get("start_date"),
            "duration": 0,
            "progress": 0,
            "system_key": "preliminary_report_submission",
            "type": "milestone",
            "is_report_bar": True,
        }
        preliminary_id = str(preliminary.get("id") or "")
        descendant_ids = {preliminary_id}
        changed = True
        while changed:
            changed = False
            for task in tasks:
                task_id = str(task.get("id") or "")
                if task_id and task_id not in descendant_ids and str(task.get("parent") or "") in descendant_ids:
                    descendant_ids.add(task_id)
                    changed = True
        insert_at = max(
            (index for index, task in enumerate(tasks) if str(task.get("id") or "") in descendant_ids),
            default=-1,
        ) + 1
        tasks.insert(insert_at, submission)
    for system_key in ("source_data", "preliminary_report", "preliminary_report_submission", "final_report"):
        task = next(
            (
                item for item in tasks
                if str(item.get("system_key") or "").strip() == system_key
            ),
            None,
        )
        if task is not None:
            task["text"] = TYPICAL_SERVICE_TERM_GANTT_SYSTEM_TASK_TEXT[system_key]
            if system_key == "preliminary_report_submission":
                task["type"] = "milestone"
    _sync_typical_service_term_gantt_asset_task(tasks, "source_data", "source_data_asset")
    _sync_typical_service_term_gantt_asset_task(tasks, "preliminary_report", "preliminary_report_asset")


def _roll_up_typical_service_term_gantt_parent_dates(tasks):
    tasks_by_id = {str(task.get("id")): task for task in tasks if task.get("id") is not None}
    children_by_parent = defaultdict(list)
    for task in tasks:
        parent_id = task.get("parent")
        if parent_id in (None, "", 0, "0"):
            continue
        children_by_parent[str(parent_id)].append(task)

    resolved = {}
    resolving = set()

    def resolve_dates(task):
        task_id = str(task.get("id"))
        if task_id in resolved:
            return resolved[task_id]
        if task_id in resolving:
            raise ValueError("В диаграмме обнаружена циклическая связь родительских задач.")

        resolving.add(task_id)
        try:
            start_date = _parse_typical_service_term_gantt_date(task.get("start_date"))
            if not start_date:
                resolved[task_id] = (None, None)
                return resolved[task_id]
            end_date = _task_end_date(task, start_date)

            child_dates = [
                resolve_dates(child)
                for child in children_by_parent.get(task_id, [])
            ]
            child_dates = [
                (child_start, child_end)
                for child_start, child_end in child_dates
                if child_start and child_end
            ]
            if child_dates:
                start_date = min(child_start for child_start, _ in child_dates)
                end_date = max(child_end for _, child_end in child_dates)
                task["start_date"] = _serialize_typical_service_term_gantt_date(start_date)
                task["end_date"] = _serialize_typical_service_term_gantt_date(end_date)
                task["duration"] = max((end_date - start_date).days, 0)
                if task.get("type") not in {"milestone", TYPICAL_SERVICE_TERM_GANTT_SERVICE_SECTION_TYPE}:
                    task["type"] = "project"

            resolved[task_id] = (start_date, end_date)
            return resolved[task_id]
        finally:
            resolving.discard(task_id)

    for task in tasks_by_id.values():
        resolve_dates(task)


def _normalize_typical_service_term_gantt_payload(
    payload,
    allowed_section_names=None,
    section_specialties_by_name=None,
    allowed_specialties=None,
    allowed_executors=None,
):
    if not isinstance(payload, dict):
        raise ValueError("Некорректный формат диаграммы.")
    tasks = payload.get("data", payload.get("tasks", []))
    links = payload.get("links", [])
    if not isinstance(tasks, list):
        raise ValueError("Список задач диаграммы должен быть массивом.")
    if not isinstance(links, list):
        raise ValueError("Список связей диаграммы должен быть массивом.")

    allowed_section_names = set(allowed_section_names or [])
    section_specialties_by_name = section_specialties_by_name or {}
    allowed_specialties = set(allowed_specialties or [])
    executor_specialties_by_key = {}
    executor_legacy_labels = defaultdict(list)
    for item in allowed_executors or []:
        if isinstance(item, dict):
            label = str(item.get("label") or "").strip()
            key = str(item.get("value") or item.get("id") or label).strip()
            specialties = {
                str(value or "").strip()
                for value in item.get("specialties", [])
                if str(value or "").strip()
            }
        else:
            label = str(item or "").strip()
            key = label
            specialties = set()
        if key:
            executor_specialties_by_key[key] = specialties
        if label and key:
            executor_legacy_labels[label].append(key)
    ambiguous_executor_labels = {
        label
        for label, count in Counter(
            str(item.get("label") or "").strip()
            for item in allowed_executors or []
            if isinstance(item, dict) and str(item.get("label") or "").strip()
        ).items()
        if count > 1
    }
    for label, keys in executor_legacy_labels.items():
        if label not in ambiguous_executor_labels and len(keys) == 1:
            executor_specialties_by_key[label] = executor_specialties_by_key[keys[0]]
    allowed_executor_keys = set(executor_specialties_by_key)
    normalized_tasks = [dict(task) for task in tasks if isinstance(task, dict)]
    normalized_links = [dict(link) for link in links if isinstance(link, dict)]
    _sync_typical_service_term_gantt_system_tasks(normalized_tasks)
    normalized_task_ids = {
        str(task.get("id")).strip()
        for task in normalized_tasks
        if task.get("id") is not None and str(task.get("id")).strip()
    }
    parent_task_ids = set()
    for task in normalized_tasks:
        parent_id = str(task.get("parent") or "").strip()
        if parent_id and parent_id not in {"0", "null", "None"} and parent_id in normalized_task_ids:
            parent_task_ids.add(parent_id)
    for link in normalized_links:
        lag_mode = str(link.get("lag_mode") or "").strip().lower()
        link["lag_mode"] = "auto" if lag_mode == "auto" else "fixed"
    meta = dict(payload.get("meta") or {}) if isinstance(payload.get("meta"), dict) else {}
    def validate_executor_specialty(executor, specialty):
        if executor and (not specialty or specialty not in executor_specialties_by_key.get(executor, set())):
            raise ValueError("Выберите исполнителя, связанного с выбранной специальностью.")

    for task in normalized_tasks:
        task_id = str(task.get("id")).strip() if task.get("id") is not None else ""
        is_parent_task = task_id in parent_task_ids
        system_key = str(task.get("system_key") or "").strip()
        if system_key in TYPICAL_SERVICE_TERM_GANTT_SYSTEM_TASK_TEXT:
            task["system_key"] = system_key
            task["text"] = TYPICAL_SERVICE_TERM_GANTT_SYSTEM_TASK_TEXT[system_key]
        specialty = str(task.get("specialty") or "").strip()
        executor = str(task.get("executor") or "").strip()
        if is_parent_task:
            specialty = ""
            executor = ""
            task.pop("resource_id", None)
            task.pop("resource_name", None)
        if specialty and specialty not in allowed_specialties:
            raise ValueError("Выберите специальность из списка.")
        if executor and executor not in allowed_executor_keys:
            raise ValueError("Выберите исполнителя из списка.")
        task["specialty"] = specialty
        task["executor"] = executor
        if str(task.get("type") or "").strip() != TYPICAL_SERVICE_TERM_GANTT_SERVICE_SECTION_TYPE:
            if not is_parent_task:
                validate_executor_specialty(executor, specialty)
            continue
        section_name = str(task.get("service_section_name") or task.get("section_name") or "").strip()
        display_name = str(task.get("text") or "").strip()
        if not section_name and display_name in allowed_section_names:
            section_name = display_name
        if not section_name or section_name not in allowed_section_names:
            raise ValueError("Выберите раздел (услугу) из списка для выбранного продукта.")
        task["service_section_name"] = section_name
        task["text"] = display_name or section_name
        if is_parent_task:
            task["specialty"] = ""
            task["executor"] = ""
            continue
        section_specialties = section_specialties_by_name.get(section_name, [])
        section_specialty_labels = [item["label"] for item in section_specialties if item.get("label")]
        if section_specialty_labels:
            if specialty and specialty not in section_specialty_labels:
                raise ValueError("Выберите специальность из списка раздела (услуги).")
            task["specialty"] = specialty or section_specialty_labels[0]
        else:
            task["specialty"] = ""
        validate_executor_specialty(executor, task["specialty"])

    task_ids = normalized_task_ids
    tasks_by_id = {
        str(task.get("id")).strip(): task
        for task in normalized_tasks
        if task.get("id") is not None and str(task.get("id")).strip()
    }

    def task_section_name(task):
        if not task:
            return ""
        explicit = str(task.get("service_section_name") or task.get("section_name") or "").strip()
        if explicit in allowed_section_names:
            return explicit
        text = str(task.get("text") or "").strip()
        if text in allowed_section_names:
            return text
        visited = set()
        parent_id = str(task.get("parent") or "").strip()
        while parent_id and parent_id not in {"0", "null", "None"} and parent_id not in visited:
            visited.add(parent_id)
            parent = tasks_by_id.get(parent_id)
            if not parent:
                return ""
            explicit = str(parent.get("service_section_name") or parent.get("section_name") or "").strip()
            if explicit in allowed_section_names:
                return explicit
            text = str(parent.get("text") or "").strip()
            if text in allowed_section_names:
                return text
            parent_id = str(parent.get("parent") or "").strip()
        return ""

    raw_resources = meta.get("resources", [])
    if raw_resources in (None, ""):
        raw_resources = []
    if not isinstance(raw_resources, list):
        raise ValueError("Список ресурсов проекта должен быть массивом.")
    normalized_resources = []
    seen_resource_ids = set()
    seen_resource_task_ids = set()
    seen_resource_pairs = set()
    resource_numbers_by_executor = {}
    next_resource_number = 1
    for index, resource in enumerate(raw_resources, start=1):
        if not isinstance(resource, dict):
            continue
        resource_id = str(resource.get("id") or f"resource-{index}").strip()
        if not resource_id:
            resource_id = f"resource-{index}"
        if resource_id in seen_resource_ids:
            base_resource_id = resource_id
            suffix = 2
            while resource_id in seen_resource_ids:
                resource_id = f"{base_resource_id}-{suffix}"
                suffix += 1
        seen_resource_ids.add(resource_id)
        specialty = str(resource.get("specialty") or "").strip()
        executor = str(resource.get("executor") or "").strip()
        if specialty and specialty not in allowed_specialties:
            raise ValueError("Выберите специальность ресурса из списка.")
        if executor and executor not in allowed_executor_keys:
            raise ValueError("Выберите исполнителя ресурса из списка.")
        validate_executor_specialty(executor, specialty)
        resource_pair = (specialty, executor)
        if specialty and executor:
            if resource_pair in seen_resource_pairs:
                raise ValueError("Ресурс с такой специальностью и ФИО уже есть в таблице.")
            seen_resource_pairs.add(resource_pair)
        resource_number_key = f"executor:{executor}" if executor else f"resource:{resource_id}"
        if resource_number_key not in resource_numbers_by_executor:
            resource_numbers_by_executor[resource_number_key] = next_resource_number
            next_resource_number += 1
        resource_name = f"Сотрудник {resource_numbers_by_executor[resource_number_key]}"
        raw_task_ids = resource.get("task_ids", resource.get("taskIds", []))
        if raw_task_ids in (None, ""):
            raw_task_ids = []
        if not isinstance(raw_task_ids, list):
            raise ValueError("Список задач ресурса должен быть массивом.")
        task_id_list = []
        for raw_task_id in raw_task_ids:
            task_id = str(raw_task_id or "").strip()
            if not task_id:
                continue
            if task_id not in task_ids:
                raise ValueError("Выберите задачу ресурса из списка задач диаграммы.")
            if task_id in parent_task_ids:
                raise ValueError("Родительскую задачу нельзя назначить ресурсу проекта.")
            task = tasks_by_id.get(task_id)
            section_name = task_section_name(task)
            section_specialty_labels = [
                item["label"]
                for item in section_specialties_by_name.get(section_name, [])
                if item.get("label")
            ]
            if specialty and section_specialty_labels and specialty not in section_specialty_labels:
                raise ValueError("Выберите задачу из раздела (услуги), доступного выбранной специальности ресурса.")
            if task_id in seen_resource_task_ids:
                raise ValueError("Одна задача не может быть назначена нескольким ресурсам.")
            seen_resource_task_ids.add(task_id)
            task_id_list.append(task_id)
        normalized_resources.append({
            "id": resource_id,
            "specialty": specialty,
            "executor": executor,
            "resource_name": resource_name,
            "task_ids": task_id_list,
            "position": index,
        })
    for resource in normalized_resources:
        for task_id in resource["task_ids"]:
            task = tasks_by_id.get(task_id)
            if not task:
                continue
            task["specialty"] = resource["specialty"]
            task["executor"] = resource["executor"]
            task["resource_id"] = resource["id"]
            task["resource_name"] = resource["resource_name"]
    meta["resources"] = normalized_resources
    _roll_up_typical_service_term_gantt_parent_dates(normalized_tasks)
    _sync_typical_service_term_gantt_system_tasks(normalized_tasks)
    _roll_up_typical_service_term_gantt_parent_dates(normalized_tasks)

    dated_tasks = []
    for task in normalized_tasks:
        start_date = _parse_typical_service_term_gantt_date(task.get("start_date"))
        if not start_date:
            continue
        end_date = _task_end_date(task, start_date)
        task["start_date"] = _serialize_typical_service_term_gantt_date(start_date)
        task["end_date"] = _serialize_typical_service_term_gantt_date(end_date)
        dated_tasks.append((task, start_date, end_date))

    if not dated_tasks:
        raise ValueError("В диаграмме должна быть хотя бы одна задача с датой начала.")

    base_date = _parse_typical_service_term_gantt_date(meta.get("base_date"))
    if not base_date:
        base_date = min(start_date for _, start_date, _ in dated_tasks)
    meta["base_date"] = _serialize_typical_service_term_gantt_date(base_date)
    project_start = _parse_typical_service_term_gantt_date(meta.get("project_start"))
    if not project_start:
        project_start = min(start_date for _, start_date, _ in dated_tasks)
    project_end = _parse_typical_service_term_gantt_date(meta.get("project_end"))
    if not project_end:
        project_end = max(end_date for _, _, end_date in dated_tasks)
    if project_end < project_start:
        project_end = project_start
    meta["project_start"] = _serialize_typical_service_term_gantt_date(project_start)
    meta["project_end"] = _serialize_typical_service_term_gantt_date(project_end)
    executor_display = str(meta.get("executor_display") or "").strip()
    if executor_display not in {
        TYPICAL_SERVICE_TERM_GANTT_EXECUTOR_DISPLAY_EXECUTOR,
        TYPICAL_SERVICE_TERM_GANTT_EXECUTOR_DISPLAY_RESOURCE,
    }:
        executor_display = (
            TYPICAL_SERVICE_TERM_GANTT_EXECUTOR_DISPLAY_RESOURCE
            if str(meta.get("calendar_kind") or "").strip() == "abstract"
            else TYPICAL_SERVICE_TERM_GANTT_EXECUTOR_DISPLAY_EXECUTOR
        )
    meta["executor_display"] = executor_display
    meta["version"] = TYPICAL_SERVICE_TERM_GANTT_VERSION

    return {"data": normalized_tasks, "links": normalized_links, "meta": meta}, dated_tasks, base_date


def _find_typical_service_term_system_task(dated_tasks, system_key):
    for task, start_date, end_date in dated_tasks:
        if task.get("system_key") == system_key:
            return task, start_date, end_date
    return None


def _calculate_typical_service_term_durations(dated_tasks):
    source_data_task = _find_typical_service_term_system_task(dated_tasks, "source_data")
    preliminary_task = _find_typical_service_term_system_task(dated_tasks, "preliminary_report")
    final_task = _find_typical_service_term_system_task(dated_tasks, "final_report")
    if not source_data_task:
        raise ValueError("Не найдена задача «Исходные данные».")
    if not preliminary_task:
        raise ValueError("Не найдена задача «Предварительный отчёт».")
    if not final_task:
        raise ValueError("Не найдена задача «Итоговый отчёт».")

    _, source_data_start, source_data_end = source_data_task
    _, preliminary_start, preliminary_end = preliminary_task
    _, final_start, final_end = final_task
    source_data_days = max((source_data_end - source_data_start).days, 0)
    preliminary_days = max((preliminary_end - preliminary_start).days, 0)
    final_days = max((final_end - final_start).days, 0)
    source_data_weeks = int((Decimal(source_data_days) / Decimal("7")).quantize(Decimal("1"), rounding=ROUND_HALF_UP))
    preliminary_months = (Decimal(preliminary_days) / Decimal("30")).quantize(
        Decimal("0.1"),
        rounding=ROUND_HALF_UP,
    )
    final_weeks = int((Decimal(final_days) / Decimal("7")).quantize(Decimal("1"), rounding=ROUND_HALF_UP))
    return max(source_data_weeks, 0), preliminary_months, max(final_weeks, 0)


def _typical_service_term_specialty_options():
    return [
        value
        for value in ExpertSpecialty.objects.order_by("position", "id").values_list("specialty", flat=True)
        if str(value or "").strip()
    ]


def _format_typical_service_term_executor_name(profile):
    user = profile.employee.user
    last_name = str(user.last_name or "").strip()
    first_name = str(user.first_name or "").strip()
    middle_name = str(profile.employee.patronymic or "").strip()
    initials = "".join(part[:1] + "." for part in (first_name, middle_name) if part)
    if last_name and initials:
        return f"{last_name} {initials}"
    return str(profile.full_name or "").strip()


def _typical_service_term_executor_value(profile):
    return f"expert-profile:{profile.pk}"


def _typical_service_term_executor_options():
    options = []
    for profile in (
        ExpertProfile.objects.select_related("employee__user")
        .prefetch_related("ranked_specialties", "ranked_specialties__specialty")
        .order_by("position", "id")
    ):
        label = _format_typical_service_term_executor_name(profile)
        if not label:
            continue
        option = {
            "id": profile.pk,
            "value": _typical_service_term_executor_value(profile),
            "label": label,
            "specialties": [],
        }
        seen_specialties = set()
        for link in profile.ranked_specialties.all():
            specialty = str(link.specialty.specialty or "").strip()
            if specialty and specialty not in seen_specialties:
                option["specialties"].append(specialty)
                seen_specialties.add(specialty)
        options.append(option)
    return options


def _typical_service_term_section_options(product_id):
    sections = (
        TypicalSection.objects.filter(product_id=product_id)
        .exclude(Q(is_system=True) | Q(code__iexact=SYSTEM_DSC_SECTION_CODE))
        .prefetch_related("ranked_specialties", "ranked_specialties__specialty")
        .order_by("position", "id")
    )
    options = []
    for section in sections:
        specialties = []
        seen = set()
        for link in section.ranked_specialties.all():
            label = str(link.specialty.specialty or "").strip()
            if not label or label in seen:
                continue
            seen.add(label)
            specialties.append({"label": label, "rank": link.rank})
        options.append({"id": section.id, "label": section.name_ru, "specialties": specialties})
    return options


def _typical_service_term_gantt_response_payload(term):
    gantt_data = copy.deepcopy(term.gantt_data) if isinstance(term.gantt_data, dict) and term.gantt_data.get("data") else None
    if gantt_data is not None:
        meta = gantt_data.setdefault("meta", {})
        meta.setdefault("calendar_kind", TYPICAL_SERVICE_TERM_GANTT_CALENDAR_KIND_ABSTRACT)
        meta.setdefault("executor_display", TYPICAL_SERVICE_TERM_GANTT_EXECUTOR_DISPLAY_RESOURCE)
        tasks = gantt_data.get("data")
        if isinstance(tasks, list):
            try:
                _sync_typical_service_term_gantt_system_tasks(tasks)
                _roll_up_typical_service_term_gantt_parent_dates(tasks)
                _sync_typical_service_term_gantt_system_tasks(tasks)
                _roll_up_typical_service_term_gantt_parent_dates(tasks)
            except ValueError:
                pass
    section_options = _typical_service_term_section_options(term.product_id)
    return {
        "ok": True,
        "gantt": gantt_data or _default_typical_service_term_gantt_data(term),
        "section_options": section_options,
        "specialty_options": _typical_service_term_specialty_options(),
        "executor_options": _typical_service_term_executor_options(),
        "term": {
            "id": term.pk,
            "product": term.product.short_name,
            "source_data_weeks": term.source_data_weeks,
            "preliminary_report_months": term.preliminary_report_months_display,
            "final_report_weeks": term.final_report_weeks,
        },
    }


@login_required
@user_passes_test(staff_required)
@require_http_methods(["GET", "POST"])
def typical_service_term_form_create(request):
    if request.method == "GET":
        form = TypicalServiceTermForm(initial=_product_field_initial_from_request(request))
        return render(
            request,
            TYPICAL_SERVICE_TERM_FORM_TEMPLATE,
            _typical_service_term_form_context(form, "create"),
        )
    form = TypicalServiceTermForm(request.POST)
    if not form.is_valid():
        return render(
            request,
            TYPICAL_SERVICE_TERM_FORM_TEMPLATE,
            _typical_service_term_form_context(form, "create"),
        )
    obj = form.save(commit=False)
    if not getattr(obj, "position", 0):
        obj.position = _next_position(TypicalServiceTerm)
    obj.save()
    return _render_policy_updated(request)


@login_required
@user_passes_test(staff_required)
@require_http_methods(["GET", "POST"])
def typical_service_term_form_edit(request, pk: int):
    term = get_object_or_404(TypicalServiceTerm, pk=pk)
    if request.method == "GET":
        form = TypicalServiceTermForm(instance=term)
        return render(
            request,
            TYPICAL_SERVICE_TERM_FORM_TEMPLATE,
            _typical_service_term_form_context(form, "edit", term),
        )
    form = TypicalServiceTermForm(request.POST, instance=term)
    if not form.is_valid():
        return render(
            request,
            TYPICAL_SERVICE_TERM_FORM_TEMPLATE,
            _typical_service_term_form_context(form, "edit", term),
        )
    form.save()
    return _render_policy_updated(request)


@login_required
@user_passes_test(staff_required)
@require_http_methods(["GET", "POST"])
def typical_service_term_gantt(request, pk: int):
    term = get_object_or_404(TypicalServiceTerm.objects.select_related("product"), pk=pk)
    if request.method == "GET":
        return JsonResponse(_typical_service_term_gantt_response_payload(term))

    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
        for task in (payload.get("data") or []) if isinstance(payload, dict) else []:
            if not isinstance(task, dict):
                continue
            for key in (
                "managed_source",
                "managed_scope",
                "work_volume_id",
                "performer_id",
                "typical_section_id",
                "template_task_id",
                "asset_name",
            ):
                task.pop(key, None)
            if task.get("system_key") == "project_asset":
                task["system_key"] = "preliminary_report_asset"
            if task.get("system_key") == "source_data_project_asset":
                task["system_key"] = "source_data_asset"
        meta = payload.setdefault("meta", {}) if isinstance(payload, dict) else {}
        if isinstance(meta, dict):
            calendar_kind = str(meta.get("calendar_kind") or TYPICAL_SERVICE_TERM_GANTT_CALENDAR_KIND_ABSTRACT).strip()
            if calendar_kind != TYPICAL_SERVICE_TERM_GANTT_CALENDAR_KIND_ABSTRACT:
                raise ValueError("Типовая диаграмма должна сохраняться только в условном календаре.")
            meta["calendar_kind"] = TYPICAL_SERVICE_TERM_GANTT_CALENDAR_KIND_ABSTRACT
            meta["executor_display"] = TYPICAL_SERVICE_TERM_GANTT_EXECUTOR_DISPLAY_RESOURCE
        section_options = _typical_service_term_section_options(term.product_id)
        section_names = [item["label"] for item in section_options]
        section_specialties_by_name = {
            item["label"]: item.get("specialties", [])
            for item in section_options
        }
        specialty_options = _typical_service_term_specialty_options()
        executor_options = _typical_service_term_executor_options()
        gantt_data, dated_tasks, base_date = _normalize_typical_service_term_gantt_payload(
            payload,
            section_names,
            section_specialties_by_name,
            specialty_options,
            executor_options,
        )
        source_data_weeks, preliminary_months, final_weeks = _calculate_typical_service_term_durations(dated_tasks)
    except (json.JSONDecodeError, UnicodeDecodeError, ValueError) as exc:
        return JsonResponse({"ok": False, "error": str(exc)}, status=400)

    term.gantt_data = gantt_data
    term.source_data_weeks = source_data_weeks
    term.preliminary_report_months = preliminary_months
    term.final_report_weeks = final_weeks
    term.save(update_fields=["gantt_data", "source_data_weeks", "preliminary_report_months", "final_report_weeks", "updated_at"])
    return JsonResponse(_typical_service_term_gantt_response_payload(term))


@login_required
@user_passes_test(staff_required)
@require_POST
def typical_service_term_delete(request, pk: int):
    term = get_object_or_404(TypicalServiceTerm, pk=pk)
    term.delete()
    return _render_policy_updated(request)


@login_required
@user_passes_test(staff_required)
@require_POST
def typical_service_term_csv_upload(request):
    csv_file = request.FILES.get("csv_file")
    if not csv_file:
        return JsonResponse({"ok": False, "error": "Файл не выбран."}, status=400)
    if not csv_file.name.lower().endswith(".csv"):
        return JsonResponse({"ok": False, "error": "Допустимы только файлы CSV."}, status=400)

    try:
        raw = csv_file.read().decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            csv_file.seek(0)
            raw = csv_file.read().decode("cp1251")
        except Exception:
            return JsonResponse(
                {"ok": False, "error": "Не удалось прочитать файл. Проверьте кодировку (UTF-8 или Windows-1251)."},
                status=400,
            )

    try:
        reader = csv.reader(io.StringIO(raw), delimiter=";")
        rows = list(reader)
        if not rows:
            return JsonResponse({"ok": False, "error": "Файл пуст."}, status=400)
        if len(rows[0]) <= 1:
            reader = csv.reader(io.StringIO(raw), delimiter=",")
            rows = list(reader)
    except csv.Error as exc:
        return JsonResponse({"ok": False, "error": f"Ошибка разбора CSV: {exc}. Проверьте формат и кодировку файла."}, status=400)

    if len(rows) < 2:
        return JsonResponse({"ok": False, "error": "Файл должен содержать заголовок и хотя бы одну строку данных."}, status=400)

    products_by_name = {_csv_lookup_key(p.short_name): p for p in Product.objects.all()}
    created = 0
    updated = 0
    warnings = []

    for i, row in enumerate(rows[1:], start=2):
        if not any(cell.strip() for cell in row):
            continue
        if len(row) < 3:
            warnings.append(
                f"Строка {i}: недостаточно столбцов ({len(row)}, ожидается 3 или 4: "
                "Продукт, Сроки предоставления исходных данных, нед., "
                "Срок подготовки Предварительного отчёта, мес., Срок подготовки Итогового отчёта, нед.)."
            )
            continue

        product_name = row[0].strip()
        product = products_by_name.get(_csv_lookup_key(product_name))
        if not product:
            warnings.append(f"Строка {i}: продукт «{product_name}» не найден. Доступные: {', '.join(products_by_name.keys())}.")
            continue

        has_source_data_column = len(row) >= 4
        source_data_raw = row[1].strip() if has_source_data_column else "0"
        preliminary_raw = row[2].strip() if has_source_data_column else row[1].strip()
        final_raw = row[3].strip() if has_source_data_column else row[2].strip()

        try:
            source_data_weeks = int(source_data_raw or "0")
        except (TypeError, ValueError):
            warnings.append(f"Строка {i}: некорректный срок исходных данных «{source_data_raw}».")
            continue
        if source_data_weeks < 0:
            warnings.append(f"Строка {i}: срок исходных данных не может быть отрицательным.")
            continue

        try:
            preliminary_report_months = Decimal(preliminary_raw.replace(",", "."))
        except (InvalidOperation, ValueError):
            warnings.append(f"Строка {i}: некорректный срок предварительного отчёта «{preliminary_raw}».")
            continue
        if preliminary_report_months < 0:
            warnings.append(f"Строка {i}: срок предварительного отчёта не может быть отрицательным.")
            continue

        try:
            final_report_weeks = int(final_raw)
        except (TypeError, ValueError):
            warnings.append(f"Строка {i}: некорректный срок итогового отчёта «{final_raw}».")
            continue
        if final_report_weeks < 0:
            warnings.append(f"Строка {i}: срок итогового отчёта не может быть отрицательным.")
            continue

        try:
            item = TypicalServiceTerm.objects.filter(product=product).order_by("position", "id").first()
            was_created = item is None
            if was_created:
                item = TypicalServiceTerm(product=product, position=_next_position(TypicalServiceTerm))
            item.source_data_weeks = source_data_weeks
            item.preliminary_report_months = preliminary_report_months
            item.final_report_weeks = final_report_weeks
            item.save()
            if was_created:
                created += 1
            else:
                updated += 1
        except Exception as exc:
            warnings.append(f"Строка {i}: ошибка сохранения — {exc}")

    return JsonResponse({"ok": True, "created": created, "updated": updated, "warnings": warnings})


@login_required
@user_passes_test(staff_required)
@require_http_methods(["GET"])
def typical_service_term_csv_download(request):
    output = io.StringIO()
    output.write("\ufeff")
    writer = csv.writer(output, delimiter=";", lineterminator="\n")
    writer.writerow(TYPICAL_SERVICE_TERM_CSV_HEADERS)

    terms = _apply_policy_master_product_filters(
        TypicalServiceTerm.objects.select_related(
            "product",
            "product__consulting_type_ref",
            "product__service_category_ref",
            "product__service_subtype_ref",
        ),
        request,
    )
    for term in terms:
        writer.writerow(
            [
                term.product.short_name,
                term.source_data_weeks,
                term.preliminary_report_months_display,
                term.final_report_weeks,
            ]
        )

    response = HttpResponse(output.getvalue(), content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="typical_service_terms.csv"'
    return response


def _normalize_typical_service_term_positions():
    items = TypicalServiceTerm.objects.order_by("position", "id").only("id", "position")
    for idx, item in enumerate(items, start=1):
        if item.position != idx:
            TypicalServiceTerm.objects.filter(pk=item.pk).update(position=idx)


@require_http_methods(["POST", "GET"])
@login_required
@user_passes_test(staff_required)
def typical_service_term_move_up(request, pk: int):
    _normalize_typical_service_term_positions()
    items = list(TypicalServiceTerm.objects.order_by("position", "id").only("id", "position"))
    idx = next((i for i, it in enumerate(items) if it.id == pk), None)
    if idx is not None and idx > 0:
        cur = items[idx]
        prev = items[idx - 1]
        cur_pos, prev_pos = cur.position, prev.position
        TypicalServiceTerm.objects.filter(pk=cur.id).update(position=prev_pos)
        TypicalServiceTerm.objects.filter(pk=prev.id).update(position=cur_pos)
        _normalize_typical_service_term_positions()
    return _render_policy_updated(request)


@require_http_methods(["POST", "GET"])
@login_required
@user_passes_test(staff_required)
def typical_service_term_move_down(request, pk: int):
    _normalize_typical_service_term_positions()
    items = list(TypicalServiceTerm.objects.order_by("position", "id").only("id", "position"))
    idx = next((i for i, it in enumerate(items) if it.id == pk), None)
    if idx is not None and idx < len(items) - 1:
        cur = items[idx]
        nxt = items[idx + 1]
        cur_pos, next_pos = cur.position, nxt.position
        TypicalServiceTerm.objects.filter(pk=cur.id).update(position=next_pos)
        TypicalServiceTerm.objects.filter(pk=nxt.id).update(position=cur_pos)
        _normalize_typical_service_term_positions()
    return _render_policy_updated(request)


# --- Направления консалтинга ---

@login_required
@user_passes_test(staff_required)
@require_http_methods(["GET", "POST"])
def consulting_dir_form_create(request):
    if request.method == "GET":
        form = ConsultingDirectionForm()
        return render(
            request,
            CONSULTING_DIR_FORM_TEMPLATE,
            _consulting_direction_form_context(form, "create"),
        )
    form = ConsultingDirectionForm(request.POST)
    if not form.is_valid():
        return _render_form_with_errors(
            request,
            CONSULTING_DIR_FORM_TEMPLATE,
            _consulting_direction_form_context(form, "create"),
        )
    if not form.instance.position:
        form.instance.position = _next_position(ConsultingDirection)
    form.save()
    return _render_policy_updated(request)


@login_required
@user_passes_test(staff_required)
@require_http_methods(["GET", "POST"])
def consulting_dir_form_edit(request, pk: int):
    direction = get_object_or_404(ConsultingDirection, pk=pk)
    if request.method == "GET":
        form = ConsultingDirectionForm(instance=direction)
        return render(
            request,
            CONSULTING_DIR_FORM_TEMPLATE,
            _consulting_direction_form_context(form, "edit", direction),
        )
    form = ConsultingDirectionForm(request.POST, instance=direction)
    if not form.is_valid():
        return _render_form_with_errors(
            request,
            CONSULTING_DIR_FORM_TEMPLATE,
            _consulting_direction_form_context(form, "edit", direction),
        )
    form.save()
    return _render_policy_updated(request)


@login_required
@user_passes_test(staff_required)
@require_POST
def consulting_dir_delete(request, pk: int):
    direction = get_object_or_404(ConsultingDirection, pk=pk)
    direction.delete()
    return _render_policy_updated(request)


def _normalize_consulting_dir_positions():
    items = ConsultingDirection.objects.order_by("position", "id").only("id", "position")
    for idx, it in enumerate(items, start=1):
        if it.position != idx:
            ConsultingDirection.objects.filter(pk=it.pk).update(position=idx)


@require_http_methods(["POST", "GET"])
@login_required
def consulting_dir_move_up(request, pk: int):
    _normalize_consulting_dir_positions()
    items = list(ConsultingDirection.objects.order_by("position", "id").only("id", "position"))
    idx = next((i for i, it in enumerate(items) if it.id == pk), None)
    if idx is not None and idx > 0:
        cur = items[idx]
        prev = items[idx - 1]
        cur_pos, prev_pos = cur.position, prev.position
        ConsultingDirection.objects.filter(pk=cur.id).update(position=prev_pos)
        ConsultingDirection.objects.filter(pk=prev.id).update(position=cur_pos)
        _normalize_consulting_dir_positions()
    return _render_policy_updated(request)


@require_http_methods(["POST", "GET"])
@login_required
def consulting_dir_move_down(request, pk: int):
    _normalize_consulting_dir_positions()
    items = list(ConsultingDirection.objects.order_by("position", "id").only("id", "position"))
    idx = next((i for i, it in enumerate(items) if it.id == pk), None)
    if idx is not None and idx < len(items) - 1:
        cur = items[idx]
        nxt = items[idx + 1]
        cur_pos, next_pos = cur.position, nxt.position
        ConsultingDirection.objects.filter(pk=cur.id).update(position=next_pos)
        ConsultingDirection.objects.filter(pk=nxt.id).update(position=cur_pos)
        _normalize_consulting_dir_positions()
    return _render_policy_updated(request)


# --- Направления экспертизы ---

@login_required
@user_passes_test(staff_required)
@require_http_methods(["GET", "POST"])
def expertise_dir_form_create(request):
    if request.method == "GET":
        form = ExpertiseDirectionForm()
        return render(request, EXPERTISE_DIR_FORM_TEMPLATE, {"form": form, "action": "create"})
    form = ExpertiseDirectionForm(request.POST)
    if not form.is_valid():
        return _render_form_with_errors(request, EXPERTISE_DIR_FORM_TEMPLATE, {"form": form, "action": "create"})
    if not form.instance.position:
        form.instance.position = _next_position(ExpertiseDirection)
    form.save()
    return _render_policy_updated(request)


@login_required
@user_passes_test(staff_required)
@require_http_methods(["GET", "POST"])
def expertise_dir_form_edit(request, pk: int):
    direction = get_object_or_404(ExpertiseDirection, pk=pk)
    if request.method == "GET":
        form = ExpertiseDirectionForm(instance=direction)
        return render(request, EXPERTISE_DIR_FORM_TEMPLATE, {"form": form, "action": "edit", "direction": direction})
    form = ExpertiseDirectionForm(request.POST, instance=direction)
    if not form.is_valid():
        return render(request, EXPERTISE_DIR_FORM_TEMPLATE, {"form": form, "action": "edit", "direction": direction})
    form.save()
    return _render_policy_updated(request)


@login_required
@user_passes_test(staff_required)
@require_POST
def expertise_dir_delete(request, pk: int):
    direction = get_object_or_404(ExpertiseDirection, pk=pk)
    direction.delete()
    return _render_policy_updated(request)


def _normalize_expertise_dir_positions():
    items = ExpertiseDirection.objects.order_by("position", "id").only("id", "position")
    for idx, it in enumerate(items, start=1):
        if it.position != idx:
            ExpertiseDirection.objects.filter(pk=it.pk).update(position=idx)


@require_http_methods(["POST", "GET"])
@login_required
def expertise_dir_move_up(request, pk: int):
    _normalize_expertise_dir_positions()
    items = list(ExpertiseDirection.objects.order_by("position", "id").only("id", "position"))
    idx = next((i for i, it in enumerate(items) if it.id == pk), None)
    if idx is not None and idx > 0:
        cur = items[idx]
        prev = items[idx - 1]
        cur_pos, prev_pos = cur.position, prev.position
        ExpertiseDirection.objects.filter(pk=cur.id).update(position=prev_pos)
        ExpertiseDirection.objects.filter(pk=prev.id).update(position=cur_pos)
        _normalize_expertise_dir_positions()
    return _render_policy_updated(request)


@require_http_methods(["POST", "GET"])
@login_required
def expertise_dir_move_down(request, pk: int):
    _normalize_expertise_dir_positions()
    items = list(ExpertiseDirection.objects.order_by("position", "id").only("id", "position"))
    idx = next((i for i, it in enumerate(items) if it.id == pk), None)
    if idx is not None and idx < len(items) - 1:
        cur = items[idx]
        nxt = items[idx + 1]
        cur_pos, next_pos = cur.position, nxt.position
        ExpertiseDirection.objects.filter(pk=cur.id).update(position=next_pos)
        ExpertiseDirection.objects.filter(pk=nxt.id).update(position=cur_pos)
        _normalize_expertise_dir_positions()
    return _render_policy_updated(request)


@login_required
@user_passes_test(staff_required)
@require_POST
def products_apply_defaults(request):
    """
    Пакетное применение флагов is_default по отмеченным чекбоксам.
    В POST приходит несколько defaults=<id>. Для всех id из списка ставим True, для остальных — False.
    """
    ids_checked = request.POST.getlist("defaults")
    ids_checked = [int(x) for x in ids_checked if str(x).isdigit()]

    # Обновляем — сначала всем False, затем отмеченным True
    Product.objects.update(is_default=False)
    if ids_checked:
        Product.objects.filter(id__in=ids_checked).update(is_default=True)

    return _render_policy_updated(request)


# --- Грейды ---

def _grade_owner(request, form):
    if request.user.is_superuser:
        owner = form.cleaned_data.get("owner")
        if owner:
            return owner
    return request.user


@login_required
@require_http_methods(["GET", "POST"])
def grade_form_create(request):
    if request.method == "GET":
        form = GradeForm(request_user=request.user)
        return render(request, GRADE_FORM_TEMPLATE, {
            "form": form, "action": "create",
            "is_admin": request.user.is_superuser,
        })
    form = GradeForm(request.POST, request_user=request.user)
    if not form.is_valid():
        return render(request, GRADE_FORM_TEMPLATE, {
            "form": form, "action": "create",
            "is_admin": request.user.is_superuser,
        })
    obj = form.save(commit=False)
    obj.created_by = _grade_owner(request, form)
    obj.position = _next_position(Grade, {"created_by": obj.created_by})
    obj.save()
    if form.cleaned_data.get("qualification_levels"):
        Grade.objects.filter(created_by=obj.created_by).exclude(pk=obj.pk).update(
            qualification_levels=obj.qualification_levels
        )
    if obj.is_base_rate:
        Grade.objects.filter(created_by=obj.created_by).exclude(pk=obj.pk).update(is_base_rate=False)
    return _render_policy_updated(request)


@login_required
@require_http_methods(["GET", "POST"])
def grade_form_edit(request, pk: int):
    grade = get_object_or_404(Grade, pk=pk)
    if not request.user.is_superuser and grade.created_by != request.user:
        return _render_policy_updated(request)
    if request.method == "GET":
        form = GradeForm(instance=grade, request_user=request.user)
        return render(request, GRADE_FORM_TEMPLATE, {
            "form": form, "action": "edit", "grade": grade,
            "is_admin": request.user.is_superuser,
        })
    form = GradeForm(request.POST, instance=grade, request_user=request.user)
    if not form.is_valid():
        return render(request, GRADE_FORM_TEMPLATE, {
            "form": form, "action": "edit", "grade": grade,
            "is_admin": request.user.is_superuser,
        })
    obj = form.save(commit=False)
    if request.user.is_superuser:
        owner = form.cleaned_data.get("owner")
        if owner:
            obj.created_by = owner
    obj.save()
    if form.cleaned_data.get("qualification_levels"):
        Grade.objects.filter(created_by=obj.created_by).exclude(pk=obj.pk).update(
            qualification_levels=obj.qualification_levels
        )
    if obj.is_base_rate:
        Grade.objects.filter(created_by=obj.created_by).exclude(pk=obj.pk).update(is_base_rate=False)
    return _render_policy_updated(request)


@login_required
@require_POST
def grade_delete(request, pk: int):
    grade = get_object_or_404(Grade, pk=pk)
    if not request.user.is_superuser and grade.created_by != request.user:
        return _render_policy_updated(request)
    grade.delete()
    return _render_policy_updated(request)


@login_required
@require_http_methods(["POST", "GET"])
def grade_move_up(request, pk: int):
    obj = get_object_or_404(Grade, pk=pk)
    if not request.user.is_superuser and obj.created_by != request.user:
        return _render_policy_updated(request)
    qs = Grade.objects.filter(created_by=obj.created_by)
    prev = qs.filter(position__lt=obj.position).order_by("-position").first()
    if prev:
        obj.position, prev.position = prev.position, obj.position
        Grade.objects.filter(pk=obj.pk).update(position=obj.position)
        Grade.objects.filter(pk=prev.pk).update(position=prev.position)
    return _render_policy_updated(request)


@login_required
@require_http_methods(["POST", "GET"])
def grade_move_down(request, pk: int):
    obj = get_object_or_404(Grade, pk=pk)
    if not request.user.is_superuser and obj.created_by != request.user:
        return _render_policy_updated(request)
    qs = Grade.objects.filter(created_by=obj.created_by)
    nxt = qs.filter(position__gt=obj.position).order_by("position").first()
    if nxt:
        obj.position, nxt.position = nxt.position, obj.position
        Grade.objects.filter(pk=obj.pk).update(position=obj.position)
        Grade.objects.filter(pk=nxt.pk).update(position=nxt.position)
    return _render_policy_updated(request)


# --- Тарифы специальностей ---

@login_required
@user_passes_test(staff_required)
@require_http_methods(["GET", "POST"])
def specialty_tariff_form_create(request):
    if request.method == "GET":
        form = SpecialtyTariffForm(request_user=request.user)
        return render(
            request,
            SPECIALTY_TARIFF_FORM_TEMPLATE,
            {
                **_specialty_tariff_form_context(form, "create"),
                "is_admin": request.user.is_superuser,
            },
        )
    form = SpecialtyTariffForm(request.POST, request_user=request.user)
    if not form.is_valid():
        return render(
            request,
            SPECIALTY_TARIFF_FORM_TEMPLATE,
            {
                **_specialty_tariff_form_context(form, "create"),
                "is_admin": request.user.is_superuser,
            },
        )
    obj = form.save(commit=False)
    obj.created_by = _specialty_tariff_owner(request, form)
    if not getattr(obj, "position", 0):
        obj.position = _next_position(SpecialtyTariff, {"created_by": obj.created_by})
    obj.save()
    form.save_m2m()
    return _render_policy_updated(request)


@login_required
@user_passes_test(staff_required)
@require_http_methods(["GET", "POST"])
def specialty_tariff_form_edit(request, pk: int):
    specialty_tariff = get_object_or_404(SpecialtyTariff, pk=pk)
    if not request.user.is_superuser and specialty_tariff.created_by != request.user:
        return _render_policy_updated(request)
    if request.method == "GET":
        form = SpecialtyTariffForm(instance=specialty_tariff, request_user=request.user)
        return render(
            request,
            SPECIALTY_TARIFF_FORM_TEMPLATE,
            {
                **_specialty_tariff_form_context(form, "edit", specialty_tariff),
                "is_admin": request.user.is_superuser,
            },
        )
    form = SpecialtyTariffForm(request.POST, instance=specialty_tariff, request_user=request.user)
    if not form.is_valid():
        return render(
            request,
            SPECIALTY_TARIFF_FORM_TEMPLATE,
            {
                **_specialty_tariff_form_context(form, "edit", specialty_tariff),
                "is_admin": request.user.is_superuser,
            },
        )
    obj = form.save(commit=False)
    if request.user.is_superuser:
        owner = form.cleaned_data.get("owner")
        if owner:
            obj.created_by = owner
    obj.save()
    form.save_m2m()
    return _render_policy_updated(request)


@login_required
@user_passes_test(staff_required)
@require_POST
def specialty_tariff_delete(request, pk: int):
    specialty_tariff = get_object_or_404(SpecialtyTariff, pk=pk)
    if not request.user.is_superuser and specialty_tariff.created_by != request.user:
        return _render_policy_updated(request)
    specialty_tariff.delete()
    return _render_policy_updated(request)


def _normalize_specialty_tariff_positions():
    items = SpecialtyTariff.objects.order_by("position", "id").only("id", "position")
    for idx, item in enumerate(items, start=1):
        if item.position != idx:
            SpecialtyTariff.objects.filter(pk=item.pk).update(position=idx)


def _normalize_tariff_positions(created_by_id: int | None = None):
    qs = Tariff.objects.only("id", "position", "created_by_id")
    if created_by_id is not None:
        items = qs.filter(created_by_id=created_by_id).order_by("position", "id")
        for idx, item in enumerate(items, start=1):
            if item.position != idx:
                Tariff.objects.filter(pk=item.pk).update(position=idx)
        return
    items = qs.order_by("created_by_id", "position", "id")
    current_owner_id = object()
    idx = 0
    for item in items:
        if item.created_by_id != current_owner_id:
            current_owner_id = item.created_by_id
            idx = 1
        else:
            idx += 1
        if item.position != idx:
            Tariff.objects.filter(pk=item.pk).update(position=idx)


@require_http_methods(["POST", "GET"])
@login_required
def specialty_tariff_move_up(request, pk: int):
    obj = get_object_or_404(SpecialtyTariff, pk=pk)
    if not request.user.is_superuser and obj.created_by != request.user:
        return _render_policy_updated(request)
    qs = SpecialtyTariff.objects.filter(created_by=obj.created_by)
    prev = qs.filter(position__lt=obj.position).order_by("-position").first()
    if prev:
        obj.position, prev.position = prev.position, obj.position
        SpecialtyTariff.objects.filter(pk=obj.pk).update(position=obj.position)
        SpecialtyTariff.objects.filter(pk=prev.pk).update(position=prev.position)
    return _render_policy_updated(request)


@require_http_methods(["POST", "GET"])
@login_required
def specialty_tariff_move_down(request, pk: int):
    obj = get_object_or_404(SpecialtyTariff, pk=pk)
    if not request.user.is_superuser and obj.created_by != request.user:
        return _render_policy_updated(request)
    qs = SpecialtyTariff.objects.filter(created_by=obj.created_by)
    nxt = qs.filter(position__gt=obj.position).order_by("position").first()
    if nxt:
        obj.position, nxt.position = nxt.position, obj.position
        SpecialtyTariff.objects.filter(pk=obj.pk).update(position=obj.position)
        SpecialtyTariff.objects.filter(pk=nxt.pk).update(position=nxt.position)
    return _render_policy_updated(request)


# --- Тарифы ---

def _tariff_owner(request, form):
    if request.user.is_superuser:
        owner = form.cleaned_data.get("owner")
        if owner:
            return owner
    return request.user


def _tariff_owner_label(user):
    employee = getattr(user, "employee_profile", None)
    if employee and employee.job_title:
        return employee.job_title
    return user.get_full_name() or user.username


def _tariff_form_context(request, form, action, tariff=None):
    ctx = {
        "form": form,
        "action": action,
        "is_admin": request.user.is_superuser,
        "sections_by_product_json": _typical_sections_by_product_json(),
    }
    if tariff:
        ctx["tariff"] = tariff
    return ctx


@login_required
@require_http_methods(["GET", "POST"])
def tariff_form_create(request):
    if request.method == "GET":
        form = TariffForm(initial=_product_field_initial_from_request(request), request_user=request.user)
        return render(request, TARIFF_FORM_TEMPLATE, _tariff_form_context(request, form, "create"))
    form = TariffForm(request.POST, request_user=request.user)
    if not form.is_valid():
        return render(request, TARIFF_FORM_TEMPLATE, _tariff_form_context(request, form, "create"))
    obj = form.save(commit=False)
    obj.created_by = _tariff_owner(request, form)
    obj.position = _next_position(Tariff, {"created_by": obj.created_by})
    obj.save()
    return _render_policy_updated(request)


@login_required
@require_http_methods(["GET", "POST"])
def tariff_form_edit(request, pk: int):
    tariff = get_object_or_404(Tariff, pk=pk)
    if not request.user.is_superuser and tariff.created_by != request.user:
        return _render_policy_updated(request)
    if request.method == "GET":
        form = TariffForm(instance=tariff, request_user=request.user)
        return render(request, TARIFF_FORM_TEMPLATE, _tariff_form_context(request, form, "edit", tariff))
    form = TariffForm(request.POST, instance=tariff, request_user=request.user)
    if not form.is_valid():
        return render(request, TARIFF_FORM_TEMPLATE, _tariff_form_context(request, form, "edit", tariff))
    obj = form.save(commit=False)
    if request.user.is_superuser:
        owner = form.cleaned_data.get("owner")
        if owner:
            obj.created_by = owner
    obj.save()
    return _render_policy_updated(request)


@login_required
@require_POST
def tariff_delete(request, pk: int):
    tariff = get_object_or_404(Tariff, pk=pk)
    if not request.user.is_superuser and tariff.created_by != request.user:
        return _render_policy_updated(request)
    tariff.delete()
    return _render_policy_updated(request)


@login_required
@require_POST
def tariff_csv_upload(request):
    csv_file = request.FILES.get("csv_file")
    if not csv_file:
        return JsonResponse({"ok": False, "error": "Файл не выбран."}, status=400)
    if not csv_file.name.lower().endswith(".csv"):
        return JsonResponse({"ok": False, "error": "Допустимы только файлы CSV."}, status=400)

    try:
        raw = csv_file.read().decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            csv_file.seek(0)
            raw = csv_file.read().decode("cp1251")
        except Exception:
            return JsonResponse(
                {"ok": False, "error": "Не удалось прочитать файл. Проверьте кодировку (UTF-8 или Windows-1251)."},
                status=400,
            )

    try:
        reader = csv.reader(io.StringIO(raw), delimiter=";")
        rows = list(reader)
        if not rows:
            return JsonResponse({"ok": False, "error": "Файл пуст."}, status=400)
        if len(rows[0]) <= 1:
            reader = csv.reader(io.StringIO(raw), delimiter=",")
            rows = list(reader)
    except csv.Error as exc:
        return JsonResponse({"ok": False, "error": f"Ошибка разбора CSV: {exc}. Проверьте формат и кодировку файла."}, status=400)

    if len(rows) < 2:
        return JsonResponse({"ok": False, "error": "Файл должен содержать заголовок и хотя бы одну строку данных."}, status=400)

    products_by_name = {_csv_lookup_key(p.short_name): p for p in Product.objects.all()}
    sections_by_product = defaultdict(dict)
    for section in TypicalSection.objects.select_related("product").all():
        lookup = sections_by_product[section.product_id]
        for label in (section.name_ru, section.name_en, section.code, section.short_name, section.short_name_ru):
            key = _csv_lookup_key(label)
            if key:
                lookup.setdefault(key, section)

    owners_by_label = {}
    if request.user.is_superuser:
        for user in get_user_model().objects.select_related("employee_profile").all():
            labels = [
                user.username,
                user.get_full_name(),
                _tariff_owner_label(user),
            ]
            for label in labels:
                key = _csv_lookup_key(label)
                if key:
                    owners_by_label.setdefault(key, user)

    created = 0
    updated = 0
    warnings = []

    for i, row in enumerate(rows[1:], start=2):
        if not any(cell.strip() for cell in row):
            continue
        if len(row) < 5:
            warnings.append(
                f"Строка {i}: недостаточно столбцов ({len(row)}, ожидается 5-6: "
                "Продукт, Раздел (услуга), Базовая ставка в ВПМ, Объем услуг в часах, "
                "Объем услуг в днях для ТКП, [Руководитель направления])."
            )
            continue

        product_name = row[0].strip()
        section_name = row[1].strip()
        product = products_by_name.get(_csv_lookup_key(product_name))
        if not product:
            warnings.append(f"Строка {i}: продукт «{product_name}» не найден. Доступные: {', '.join(products_by_name.keys())}.")
            continue

        section = sections_by_product[product.pk].get(_csv_lookup_key(section_name))
        if not section:
            warnings.append(f"Строка {i}: раздел «{section_name}» не найден для продукта «{product.short_name}».")
            continue

        try:
            base_rate_vpm = Decimal(row[2].strip().replace(",", "."))
        except (InvalidOperation, ValueError):
            warnings.append(f"Строка {i}: некорректная базовая ставка «{row[2].strip()}».")
            continue
        if base_rate_vpm < 0:
            warnings.append(f"Строка {i}: базовая ставка не может быть отрицательной.")
            continue

        try:
            service_hours = int(row[3].strip())
        except (TypeError, ValueError):
            warnings.append(f"Строка {i}: некорректный объем услуг в часах «{row[3].strip()}».")
            continue
        if service_hours < 0:
            warnings.append(f"Строка {i}: объем услуг в часах не может быть отрицательным.")
            continue

        try:
            service_days_tkp = int(row[4].strip())
        except (TypeError, ValueError):
            warnings.append(f"Строка {i}: некорректный объем услуг в днях для ТКП «{row[4].strip()}».")
            continue
        if service_days_tkp < 0:
            warnings.append(f"Строка {i}: объем услуг в днях для ТКП не может быть отрицательным.")
            continue

        owner = request.user
        owner_name = row[5].strip() if len(row) > 5 else ""
        if request.user.is_superuser and owner_name:
            owner = owners_by_label.get(_csv_lookup_key(owner_name))
            if not owner:
                warnings.append(f"Строка {i}: руководитель «{owner_name}» не найден.")
                continue

        try:
            tariff = (
                Tariff.objects
                .filter(product=product, section=section, created_by=owner)
                .order_by("position", "id")
                .first()
            )
            was_created = tariff is None
            if was_created:
                tariff = Tariff(
                    product=product,
                    section=section,
                    created_by=owner,
                    position=_next_position(Tariff, {"created_by": owner}),
                )
            tariff.base_rate_vpm = base_rate_vpm
            tariff.service_hours = service_hours
            tariff.service_days_tkp = service_days_tkp
            tariff.save()
            if was_created:
                created += 1
            else:
                updated += 1
        except Exception as exc:
            warnings.append(f"Строка {i}: ошибка сохранения — {exc}")

    return JsonResponse({"ok": True, "created": created, "updated": updated, "warnings": warnings})


@login_required
@require_http_methods(["GET"])
def tariff_csv_download(request):
    output = io.StringIO()
    output.write("\ufeff")
    writer = csv.writer(output, delimiter=";", lineterminator="\n")
    writer.writerow(TARIFF_CSV_HEADERS)

    for tariff in _apply_policy_master_product_filters(_get_tariffs_for_user(request.user), request):
        writer.writerow(
            [
                tariff.product.short_name,
                tariff.section.name_ru or tariff.section.name_en,
                str(tariff.base_rate_vpm).replace(".", ","),
                tariff.service_hours,
                tariff.service_days_tkp,
                _tariff_owner_label(tariff.created_by),
            ]
        )

    response = HttpResponse(output.getvalue(), content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="section_tariffs.csv"'
    return response


@login_required
@require_http_methods(["POST", "GET"])
def tariff_move_up(request, pk: int):
    obj = get_object_or_404(Tariff, pk=pk)
    if not request.user.is_superuser and obj.created_by != request.user:
        return _render_policy_updated(request)
    _normalize_tariff_positions(created_by_id=obj.created_by_id)
    items = list(Tariff.objects.filter(created_by_id=obj.created_by_id).order_by("position", "id").only("id", "position"))
    idx = next((i for i, item in enumerate(items) if item.id == pk), None)
    if idx is not None and idx > 0:
        cur = items[idx]
        prev = items[idx - 1]
        cur_pos, prev_pos = cur.position, prev.position
        Tariff.objects.filter(pk=cur.id).update(position=prev_pos)
        Tariff.objects.filter(pk=prev.id).update(position=cur_pos)
        _normalize_tariff_positions(created_by_id=obj.created_by_id)
    return _render_policy_updated(request)


@login_required
@require_http_methods(["POST", "GET"])
def tariff_move_down(request, pk: int):
    obj = get_object_or_404(Tariff, pk=pk)
    if not request.user.is_superuser and obj.created_by != request.user:
        return _render_policy_updated(request)
    _normalize_tariff_positions(created_by_id=obj.created_by_id)
    items = list(Tariff.objects.filter(created_by_id=obj.created_by_id).order_by("position", "id").only("id", "position"))
    idx = next((i for i, item in enumerate(items) if item.id == pk), None)
    if idx is not None and idx < len(items) - 1:
        cur = items[idx]
        nxt = items[idx + 1]
        cur_pos, next_pos = cur.position, nxt.position
        Tariff.objects.filter(pk=cur.id).update(position=next_pos)
        Tariff.objects.filter(pk=nxt.id).update(position=cur_pos)
        _normalize_tariff_positions(created_by_id=obj.created_by_id)
    return _render_policy_updated(request)
