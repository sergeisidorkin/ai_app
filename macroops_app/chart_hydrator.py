# macroops_app/chart_hydrator.py
from __future__ import annotations
import os, io, json, zipfile, uuid
from typing import Dict, Any, List, Optional, Tuple
import base64
from lxml import etree
from io import BytesIO
from zipfile import ZipFile
from openpyxl.utils.cell import range_boundaries

try:
    # openpyxl >= 3.0
    from openpyxl.workbook.defined_name import DefinedName
except Exception:
    DefinedName = None

# openpyxl — для работы с embedded workbook
try:
    import openpyxl
except Exception as e:
    raise RuntimeError("openpyxl is required for chart hydration: pip install openpyxl") from e

try:
    from logs_app import utils as plog
except Exception:
    plog = None

CHARTS_DIR = os.path.join(os.path.dirname(__file__), "assets", "charts")

# XML namespace map для chart*.xml
NS = {
    "c": "http://schemas.openxmlformats.org/drawingml/2006/chart",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
}

def _dn_get(wb, name):
    """
    Вернуть объект DefinedName (или None), совместимо с openpyxl 2.x/3.x.
    """
    dn = None
    try:
        dn = wb.defined_names.get(name)
        # В новых версиях get() может вернуть список (скоуповые имена)
        if isinstance(dn, list):
            dn = dn[0] if dn else None
    except Exception:
        dn = None

    if dn is not None:
        return dn

    # Легаси-доступ: в старых версиях это был .definedName (list)
    try:
        seq = getattr(wb.defined_names, 'definedName', None)
        if seq:
            for x in seq:
                if getattr(x, 'name', None) == name:
                    return x
    except Exception:
        pass
    return None

def _dn_destinations(wb, dn):
    """
    Разрешить destinations именованного диапазона в список (ws, coord).
    Совместимо с openpyxl 2.x/3.x.
    """
    if dn is None:
        return []
    # Нормальный путь
    try:
        dests = list(getattr(dn, 'destinations', []) or [])
        if dests:
            out = []
            for title, coord in dests:
                try:
                    ws = wb[title]
                    out.append((ws, coord))
                except Exception:
                    continue
            return out
    except Exception:
        pass

    # Легаси: парсим attr_text вроде 'Лист1'!$B$2:$B$8
    txt = getattr(dn, 'attr_text', None) or getattr(dn, 'value', None) or ''
    if '!' in str(txt):
        try:
            title, coord = str(txt).split('!', 1)
            title = title.strip().strip("'")
            coord = coord.strip()
            ws = wb[title]
            return [(ws, coord)]
        except Exception:
            return []
    return []

def _write_1d(ws, coord, values, *, clear: bool = False):
    """
    Пишет values в указанный 1D-диапазон.
    clear=False (по умолч.) — патч: заполняем только первые len(values) ячеек, остальное не трогаем.
    clear=True  — полностью перезаписываем диапазон, «лишние» ячейки очищаем.
    """
    min_col, min_row, max_col, max_row = range_boundaries(coord)
    height = max_row - min_row + 1
    width  = max_col - min_col + 1
    vals   = list(values or [])

    if width == 1:  # вертикаль
        if clear:
            for i in range(height):
                cell = ws.cell(row=min_row + i, column=min_col)
                cell.value = vals[i] if i < len(vals) else ""
        else:
            for i in range(min(height, len(vals))):
                ws.cell(row=min_row + i, column=min_col).value = vals[i]
    elif height == 1:  # горизонталь
        if clear:
            for j in range(width):
                cell = ws.cell(row=min_row, column=min_col + j)
                cell.value = vals[j] if j < len(vals) else ""
        else:
            for j in range(min(width, len(vals))):
                ws.cell(row=min_row, column=min_col + j).value = vals[j]
    else:
        # Матрица
        if clear:
            idx = 0
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    cell = ws.cell(row=r, column=c)
                    cell.value = vals[idx] if idx < len(vals) else ""
                    idx += 1
        else:
            idx = 0
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    if idx >= len(vals):
                        return
                    ws.cell(row=r, column=c).value = vals[idx]
                    idx += 1

def _write_scalar(ws, coord, value):
    min_col, min_row, max_col, max_row = range_boundaries(coord)
    ws.cell(row=min_row, column=min_col).value = value

def _log(level: str, event: str, *, message: str = "", data: Dict[str, Any] | None = None, log_ctx: Dict[str, Any] | None = None):
    """
    Безопасное логирование для этапов гидрации диаграмм:
    - никогда не выбрасывает исключения;
    - гарантирует ненулевой project_code6 (если нет — 'UNK'), чтобы не бить NOT NULL в БД.
    """
    if not plog:
        return
    fn = getattr(plog, level, getattr(plog, "info", None)) or (lambda *a, **k: None)
    ctx = dict(log_ctx or {})
    pc6 = (ctx.get("project_code6")
           or ctx.get("code6")
           or ctx.get("project")
           or "UNK")
    # жёстко прокидываем ожидаемые поля
    kwargs = {
        "phase": "chart",
        "event": event,
        "message": message or "",
        "trace_id": ctx.get("trace_id"),
        "email": ctx.get("email"),
        "project_code6": str(pc6),   # <— гарантированно НЕ None
        "data": (data or {}),
    }
    try:
        fn(None, **kwargs)
    except Exception:
        # Логгер никогда не должен ломать основную логику
        try:
            # попытаемся записать максимально «пусто»
            fn(None, phase="chart", event=event, message=message or "",
               project_code6="UNK", data=(data or {}))
        except Exception:
            pass

def _ensure_dir():
    if not os.path.isdir(CHARTS_DIR):
        raise FileNotFoundError(f"Charts dir not found: {CHARTS_DIR}")

def _resolve_template_paths(template: str) -> Tuple[str, Optional[str]]:
    """
    Находит radar_esg_v1.docx и (опционально) radar_esg_v1.meta.json.
    """
    _ensure_dir()
    base = template.strip()
    if not base.lower().endswith(".docx"):
        docx_path = os.path.join(CHARTS_DIR, f"{base}.docx")
    else:
        docx_path = os.path.join(CHARTS_DIR, base)

    if not os.path.exists(docx_path):
        # Попробуем точечный поиск среди файлов (чтобы не зависеть от расширения в имени)
        for fn in os.listdir(CHARTS_DIR):
            if fn.lower().endswith(".docx") and os.path.splitext(fn)[0].lower() == os.path.splitext(base)[0].lower():
                docx_path = os.path.join(CHARTS_DIR, fn)
                break

    if not os.path.exists(docx_path):
        raise FileNotFoundError(f"Chart template not found: {docx_path}")

    meta_guess = os.path.splitext(docx_path)[0] + ".meta.json"
    return docx_path, (meta_guess if os.path.exists(meta_guess) else None)

def _pick_first(path_list: List[str], suffix: str) -> Optional[str]:
    for p in path_list:
        if p.lower().endswith(suffix.lower()):
            return p
    return None

def _find_embedded_workbook(zf: zipfile.ZipFile, meta: Dict[str, Any]) -> str:
    # Ищем /word/embeddings/*.xlsx
    xlsx_files = [n for n in zf.namelist() if n.startswith("word/embeddings/") and n.lower().endswith(".xlsx")]
    if not xlsx_files:
        raise RuntimeError("No embedded workbook (.xlsx) found in template DOCX")
    # meta["workbook"] может быть именем файла (хвост), иначе — auto (берём первый)
    hint = str(meta.get("workbook") or "auto").strip().lower()
    if hint != "auto":
        for n in xlsx_files:
            if os.path.basename(n).lower() == hint or n.lower().endswith("/" + hint):
                return n
    return xlsx_files[0]

def _find_chart_xml(zf: zipfile.ZipFile, meta: Dict[str, Any]) -> str:
    # Ищем /word/charts/chart*.xml (первый)
    charts = [n for n in zf.namelist() if n.startswith("word/charts/") and n.lower().endswith(".xml")]
    if not charts:
        raise RuntimeError("No chart XML found in template DOCX")
    # можно расширить по meta["chart_xml_hint"], но на фазе 1 — auto
    return charts[0]

def _col_row(col: str, row: int) -> str:
    return f"{col}{int(row)}"

def _read_1d(ws, coord):
    """Читает 1D-диапазон как список значений (без преобразований типов)."""
    min_col, min_row, max_col, max_row = range_boundaries(coord)
    out = []
    if min_col == max_col:  # вертикаль
        for i in range(min_row, max_row + 1):
            out.append(ws.cell(row=i, column=min_col).value)
    elif min_row == max_row:  # горизонталь
        for j in range(min_col, max_col + 1):
            out.append(ws.cell(row=min_row, column=j).value)
    else:
        # матрицу разворачиваем по строкам
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                out.append(ws.cell(row=r, column=c).value)
    return out

def _read_scalar(wb, fml: str):
    """Читает левую верхнюю ячейку диапазона, заданного именем/ссылкой fml."""
    for ws, coord in _parse_formula_target(wb, fml):
        min_col, min_row, _, _ = range_boundaries(coord)
        return ws.cell(row=min_row, column=min_col).value
    return None

def _parse_formula_target(wb, formula: str):
    """
    Возвращает список (ws, coord) для формулы:
      • 'Лист1'!$B$2:$B$9  → [(ws, '$B$2:$B$9')]
      • _S1 / Labels (имя) → по defined name
    """
    f = (formula or "").strip()
    if not f:
        return []
    if '!' in f:
        try:
            sheet, coord = f.split('!', 1)
            sheet = sheet.strip().strip("'")
            coord = coord.strip()
            ws = wb[sheet]
            return [(ws, coord)]
        except Exception:
            return []
    dn = _dn_get(wb, f)
    return _dn_destinations(wb, dn)

def _update_chart_cache_from_wb(chart_xml_bytes: bytes, wb) -> bytes:
    """
    Для КАЖДОЙ серии в chart*.xml:
      • читает реальные ranges из c:f (tx/cat/val),
      • берёт значения из книги,
      • пишет их в strCache/numCache.
    """
    root = etree.fromstring(chart_xml_bytes)
    plot = root.find(".//c:plotArea", namespaces=NS)
    if plot is None:
        return chart_xml_bytes

    series = plot.findall(".//c:ser", namespaces=NS)
    for ser in series:
        # Название серии
        tx_f = (ser.findtext("./c:tx/c:strRef/c:f", namespaces=NS) or "").strip()
        if tx_f:
            tx_val = _read_scalar(wb, tx_f)
            if tx_val is not None:
                tx = ser.find("./c:tx", namespaces=NS) or etree.SubElement(ser, f"{{{NS['c']}}}tx")
                tx_sr = tx.find("./c:strRef", namespaces=NS) or etree.SubElement(tx, f"{{{NS['c']}}}strRef")
                _set_str_cache(tx_sr, [str(tx_val)])

        # Категории
        cat_f = (ser.findtext("./c:cat/c:strRef/c:f", namespaces=NS) or "").strip()
        if cat_f:
            labels = []
            for ws, coord in _parse_formula_target(wb, cat_f):
                labels = [("" if v is None else str(v)) for v in _read_1d(ws, coord)]
                break
            cat = ser.find("./c:cat", namespaces=NS) or etree.SubElement(ser, f"{{{NS['c']}}}cat")
            str_ref = cat.find("./c:strRef", namespaces=NS) or etree.SubElement(cat, f"{{{NS['c']}}}strRef")
            _set_str_cache(str_ref, labels)

        # Значения
        val_f = (ser.findtext("./c:val/c:numRef/c:f", namespaces=NS) or "").strip()
        if val_f:
            nums: List[float] = []
            for ws, coord in _parse_formula_target(wb, val_f):
                raw = _read_1d(ws, coord)
                nums = []
                for v in raw:
                    try:
                        nums.append(float(str(v).replace(",", ".").strip()))
                    except Exception:
                        nums.append(0.0 if (v is None or (isinstance(v, str) and v.strip() == "")) else float("nan"))
                break
            val = ser.find("./c:val", namespaces=NS) or etree.SubElement(ser, f"{{{NS['c']}}}val")
            num_ref = val.find("./c:numRef", namespaces=NS) or etree.SubElement(val, f"{{{NS['c']}}}numRef")
            _set_num_cache(num_ref, nums)

    return etree.tostring(root, encoding="UTF-8", standalone=True)

def _apply_targets(wb, targets: List[Dict[str, Any]]):
    """
    Пишет произвольное число целей:
      { "target": <defined name | 'Sheet'!A1:B9>, "values": [...], "scalar": "...", "clear": false }
    """
    if not isinstance(targets, list):
        return
    for i, t in enumerate(targets):
        try:
            if not isinstance(t, dict):
                continue
            tgt = str(t.get("target") or "").strip()
            if not tgt:
                continue
            clear = bool(t.get("clear") or False)
            if "scalar" in t and t["scalar"] is not None:
                for ws, coord in _parse_formula_target(wb, tgt):
                    _write_scalar(ws, coord, t["scalar"])
                    _log("debug", "target.write", data={"i": i, "target": tgt, "kind": "scalar", "coord": coord})
                    break
                continue
            vals = t.get("values")
            if not isinstance(vals, (list, tuple)):
                continue
            dests = _parse_formula_target(wb, tgt)
            if not dests:
                _log("warn", "target.resolve.fail", data={"i": i, "target": tgt})
                continue
            for ws, coord in dests:
                _write_1d(ws, coord, list(vals), clear=clear)
                _log("debug", "target.write", data={"i": i, "target": tgt, "kind": "values", "coord": coord, "n": len(vals)})
                break
        except Exception as e:
            _log("warn", "target.write.fail", message=str(e), data={"i": i, "target": t})

def _safe_clear(el: etree._Element, tag: str):
    for n in list(el.findall(tag, namespaces=NS)):
        el.remove(n)

def _set_str_cache(str_ref_el: etree._Element, labels: List[str]):
    # Удалим старый кэш, создадим новый
    cache = str_ref_el.find("c:strCache", namespaces=NS)
    if cache is None:
        cache = etree.SubElement(str_ref_el, f"{{{NS['c']}}}strCache")
    # очистка
    for child in list(cache):
        cache.remove(child)
    ptCount = etree.SubElement(cache, f"{{{NS['c']}}}ptCount")
    ptCount.set("val", str(len(labels)))
    for idx, txt in enumerate(labels):
        pt = etree.SubElement(cache, f"{{{NS['c']}}}pt")
        pt.set("idx", str(idx))
        v = etree.SubElement(pt, f"{{{NS['c']}}}v")
        # Текст допускает переводы строк
        v.text = txt

def _set_num_cache(num_ref_el: etree._Element, values: List[float]):
    cache = num_ref_el.find("c:numCache", namespaces=NS)
    if cache is None:
        cache = etree.SubElement(num_ref_el, f"{{{NS['c']}}}numCache")
    for child in list(cache):
        cache.remove(child)
    ptCount = etree.SubElement(cache, f"{{{NS['c']}}}ptCount")
    ptCount.set("val", str(len(values)))
    # Можно задать formatCode, но не обязательно
    for idx, val in enumerate(values):
        pt = etree.SubElement(cache, f"{{{NS['c']}}}pt")
        pt.set("idx", str(idx))
        v = etree.SubElement(pt, f"{{{NS['c']}}}v")
        v.text = str(float(val))

def build_chart_snippet_base64(template: str, data: dict, *, log_ctx=None) -> str:
    """
    Собирает .docx с диаграммой на основе шаблона template и данных data.
    ЕДИНСТВЕННЫЙ вход — data.targets:
      {
        "targets": [
          { "target": "_S1", "values": [4,5,4,4,4,2,4,4], "clear": false },
          { "target": "'ChartData'!$B$1", "scalar": "Итоговый балл (0–5)" },
          { "target": "_CATS", "values": ["Политика","Риски…","…"], "clear": false }
        ]
      }
    После записи в книгу обновляет strCache/numCache у всех серий из фактических данных.
    """
    from django.conf import settings
    import os

    template = str(template or "").strip()
    data = data or {}

    _log("info", "hydrate.begin", data={"template": template, "keys": list(data.keys())}, log_ctx=log_ctx)

    # 1) найти файл-шаблон docx
    def _resolve_docx_path(tmpl: str) -> str:
        names = [tmpl] if tmpl.lower().endswith(".docx") else [tmpl, f"{tmpl}.docx"]
        tried: list[str] = []

        # a) абсолютный путь
        for nm in names:
            if os.path.isabs(nm):
                tried.append(nm)
                if os.path.isfile(nm):
                    return nm

        # b) каталоги-кандидаты
        candidates = []
        cfg_dir = getattr(settings, "CHART_TEMPLATES_DIR", None)
        if cfg_dir:
            candidates.append(cfg_dir)
        candidates.append(CHARTS_DIR)

        for base in candidates:
            for nm in names:
                p = os.path.join(base, nm)
                tried.append(p)
                if os.path.isfile(p):
                    return p

        # c) мягкий поиск по стему внутри CHARTS_DIR
        try:
            stem = os.path.splitext(tmpl)[0].lower()
            for fn in os.listdir(CHARTS_DIR):
                if fn.lower().endswith(".docx") and os.path.splitext(fn)[0].lower() == stem:
                    return os.path.join(CHARTS_DIR, fn)
        except Exception:
            pass

        _log("error", "template.resolve.fail",
             message="chart template not found",
             data={"template": tmpl, "tried": tried}, log_ctx=log_ctx)
        raise FileNotFoundError(f"chart template not found, tried: {tried}")

    docx_path = _resolve_docx_path(template)
    _log("info", "template.used", data={"path": docx_path}, log_ctx=log_ctx)

    # 2) распаковать docx, найти embedded xlsx
    with open(docx_path, "rb") as f:
        docx_bytes = f.read()

    with ZipFile(BytesIO(docx_bytes), "r") as zin:
        names = zin.namelist()
        xlsx_name = None
        for nm in names:
            if nm.lower().startswith("word/embeddings/") and nm.lower().endswith(".xlsx"):
                xlsx_name = nm
                break
        if not xlsx_name:
            raise RuntimeError("embedded workbook (.xlsx) not found in docx")
        xlsx_bytes = zin.read(xlsx_name)

    # 3) загрузить xlsx через openpyxl и применить targets
    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(xlsx_bytes))

    targets = data.get("targets") or []
    if targets:
        _apply_targets(wb, targets)

    # 4) собрать новый docx: подменить xlsx и обновить кэш диаграмм
    out_doc = BytesIO()
    with ZipFile(BytesIO(docx_bytes), "r") as zin, ZipFile(out_doc, "w") as zout:
        names = zin.namelist()
        chart_xml_names = [n for n in names if n.startswith("word/charts/") and n.lower().endswith(".xml")]

        out_xlsx = BytesIO()
        wb.save(out_xlsx)
        out_xlsx.seek(0)
        new_xlsx_bytes = out_xlsx.read()

        updated_cnt = 0
        for nm in names:
            if nm == xlsx_name:
                zout.writestr(nm, new_xlsx_bytes)  # подменили xlsx
            elif nm in chart_xml_names:
                try:
                    orig_xml = zin.read(nm)
                    upd_xml = _update_chart_cache_from_wb(orig_xml, wb)
                    zout.writestr(nm, upd_xml)
                    updated_cnt += 1
                except Exception:
                    zout.writestr(nm, orig_xml)  # fallback: оставить как было
            else:
                zout.writestr(nm, zin.read(nm))

        _log("info", "cache.update.summary", data={"charts_updated": updated_cnt}, log_ctx=log_ctx)

    result_b64 = base64.b64encode(out_doc.getvalue()).decode("ascii")
    _log("info", "hydrate.ok", data={"template": template, "b64_len": len(result_b64)}, log_ctx=log_ctx)
    return result_b64