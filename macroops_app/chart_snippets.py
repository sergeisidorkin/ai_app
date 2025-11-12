# /Users/sergei/PycharmProjects/ai_app/macroops_app/chart_snippets.py
from __future__ import annotations
import os, zipfile
from io import BytesIO
from base64 import b64encode
from typing import Any, Dict, Optional, List

try:
    from logs_app import utils as plog
except Exception:
    plog = None

def _log(level: str, event: str, *, message: str = "", data: Dict[str, Any] | None = None, log_ctx: Dict[str, Any] | None = None):
    if not plog:
        return
    ctx = log_ctx or {}
    fn = getattr(plog, level, plog.info)
    try:
        fn(None, phase="chart", event=event, message=message,
           trace_id=ctx.get("trace_id"), email=ctx.get("email"), data=(data or {}))
    except Exception:
        pass

def _charts_dir() -> str:
    return os.path.join(os.path.dirname(__file__), "assets", "charts")

def build_chart_snippet_base64(template: str, data: Dict[str, Any] | None = None, *, log_ctx: Dict[str, Any] | None = None) -> str:
    """
    Загружает шаблон DOCX с диаграммой и (по возможности) обновляет значения серии
    в embedded Excel (sheet 'Data', ряд 2 начиная с B2). Если не удаётся — возвращает
    исходный файл без изменений (фолбэк).
    """
    data = data or {}
    path = os.path.join(_charts_dir(), f"{template}.docx")
    _log("info", "hydrate.begin", data={"template": template, "path": path}, log_ctx=log_ctx)

    if not os.path.exists(path):
        _log("error", "template.not_found", message="chart template not found", data={"path": path}, log_ctx=log_ctx)
        raise FileNotFoundError(path)

    with open(path, "rb") as f:
        docx_bytes = f.read()

    # Пытаемся обновить embedded workbook
    try:
        from openpyxl import load_workbook  # optional dep
        zin = zipfile.ZipFile(BytesIO(docx_bytes), "r")
        embed_name: Optional[str] = None
        for name in zin.namelist():
            if name.startswith("word/embeddings/") and name.endswith(".xlsx"):
                embed_name = name
                break

        if not embed_name:
            _log("warn", "hydrate.no_workbook", message="no embedded workbook found", log_ctx=log_ctx)
            return b64encode(docx_bytes).decode("ascii")

        xlsx_bytes = zin.read(embed_name)
        wb = load_workbook(BytesIO(xlsx_bytes))
        ws = wb.active  # convention: 'Data' или первый лист

        values = list((data.get("values") or []))
        # пишем значения во 2-ю строку, начиная с B2
        for i, v in enumerate(values, start=2):
            try:
                ws.cell(row=2, column=i).value = float(v)
            except Exception:
                ws.cell(row=2, column=i).value = v

        bio_x = BytesIO()
        wb.save(bio_x)
        new_xlsx = bio_x.getvalue()

        bio_out = BytesIO()
        zout = zipfile.ZipFile(bio_out, "w", zipfile.ZIP_DEFLATED)
        for name in zin.namelist():
            if name == embed_name:
                zout.writestr(name, new_xlsx)
            else:
                zout.writestr(name, zin.read(name))
        zin.close(); zout.close()

        out_bytes = bio_out.getvalue()
        _log("info", "hydrate.ok", data={"template": template, "points": len(values)}, log_ctx=log_ctx)
        return b64encode(out_bytes).decode("ascii")

    except Exception as e:
        # мягкий фолбэк — возвращаем исходный файл (диаграмма вставится с дефолтными данными)
        _log("warn", "hydrate.failed", message=str(e), data={"template": template}, log_ctx=log_ctx)
        return b64encode(docx_bytes).decode("ascii")